import time
import io
import os
import json
import glob
import difflib
import threading
import re
import hashlib
from datetime import datetime
from html import escape
from urllib.parse import urlparse, quote

import requests
import pandas as pd
import streamlit as st
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES Y DICCIONARIOS
# ══════════════════════════════════════════════════════════════════════════════
TRADUCCIONES_ES_EN = {
    "órdenes de jefe": "Boss's Orders", "ordenes de jefe": "Boss's Orders",
    "órdenes del jefe": "Boss's Orders", "sabio": "Cynthia",
    "juicio de lillie": "Lillie", "investigación del profesor": "Professor's Research",
    "investigación del profesor juniper": "Professor's Research",
    "naranja académica": "Iono", "iono": "Iono", "miriam": "Miriam",
    "arven": "Arven", "penny": "Penny", "jacq": "Jacq", "tulip": "Tulip",
    "larry": "Larry", "ultra ball": "Ultra Ball", "ultra bola": "Ultra Ball",
    "nido ball": "Nest Ball", "nest ball": "Nest Ball", "poké ball": "Poké Ball",
    "gran ball": "Great Ball", "bola rápida": "Quick Ball", "quick ball": "Quick Ball",
    "poción": "Potion", "súper poción": "Super Potion", "hiperpoción": "Hyper Potion",
    "cinturón luchador": "Choice Belt", "banda elegida": "Choice Band",
    "banda de combate": "Choice Band", "guante de combate": "Counter Catcher",
    "escudo mental": "Collapsed Stadium", "energizador": "Energy Recycler",
    "parche turbo": "Turbo Patch", "roca lisa": "Smooth Stone",
    "bucle brillante": "Pal Pad", "repetición de película": "Lure Ball",
    "estadio roto": "Collapsed Stadium", "academia naranja": "Orange Academy",
    "gimnasio pokémon": "Pokémon League Headquarters",
    "energía agua": "Water Energy", "energía fuego": "Fire Energy",
    "energía eléctrica": "Lightning Energy", "energía planta": "Grass Energy",
    "energía psíquica": "Psychic Energy", "energía lucha": "Fighting Energy",
    "energía oscuridad": "Darkness Energy", "energía metal": "Metal Energy",
    "energía hada": "Fairy Energy", "energía doble": "Double Colorless Energy",
}

KEYWORDS_LIGA = ["promo", "league", "player cup", "professor program"]

# ── Normalización de apóstrofes ───────────────────────────────────────────────
def _normalizar_apostrofe(texto: str) -> str:
    return texto.replace("\u2019", "'").replace("\u2018", "'").replace("\u02bc", "'")

TIPOS_API = {
    "pokemon": "Pokémon", "pokémon": "Pokémon",
    "trainer": "Trainer", "entrenador": "Trainer",
    "energy": "Energy", "energía": "Energy", "energia": "Energy",
}

# Colores para tipos de carta adaptados para combinar con la paleta clara de NexoGeek
TIPO_COLORES = {
    "Pokémon":  {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#BEE3F8", "text": "#1E3A8A"},
    "pokemon":  {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#BEE3F8", "text": "#1E3A8A"},
    "Trainer":  {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#E2E8F0", "text": "#475569"},
    "trainer":  {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#E2E8F0", "text": "#475569"},
    "Energy":   {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#FEEBC8", "text": "#C05621"},
    "energy":   {"bg": "#FFFFFF", "border": "#CBD5E1", "badge": "#FEEBC8", "text": "#C05621"},
}
COLOR_DEFAULT = {"bg": "#FFFFFF", "border": "#E2E8F0", "badge": "#F1F5F9", "text": "#334155"}

BASE_URL               = "https://api.pokemontcg.io/v2/cards"
DELAY_ENTRE_CONSULTAS  = 1.2   # segundos

# ── Caché global de consultas (segura entre hilos) ────────────────────────────
# A diferencia de st.session_state (que NO es accesible de forma fiable desde los
# hilos del ThreadPoolExecutor), un dict de módulo + Lock sí es thread-safe.
# Persiste mientras el servidor esté encendido; se vacía al reiniciar la app.
_QUERY_CACHE: dict = {}
_QUERY_LOCK = threading.Lock()

# Caché de precios compartida por el proceso de Streamlit. La base local se usa
# para identificar cartas e imágenes; esta caché evita volver a consultar el
# mismo Card ID durante seis horas.
_PRICE_CACHE: dict = {}
_PRICE_CACHE_LOCK = threading.Lock()
_PRICE_CACHE_TTL = 6 * 60 * 60

# ── Métricas de diagnóstico (para entender lentitud / rate limit en vivo) ─────
_STATS = {"requests": 0, "cache_hits": 0, "rate_limited": 0, "errores": 0, "tiempo_red": 0.0}
_STATS_LOCK = threading.Lock()

# Config ajustable desde el sidebar (la lee _raw_query y el loop de proceso)
_CFG = {"throttle": 0.2, "max_workers": 4}

# Persistencia ligera para el piloto. En un despliegue definitivo debe migrarse
# a una base de datos, pero para un grupo reducido permite recopilar feedback.
FEEDBACK_FILE = os.getenv("NEXOGEEK_FEEDBACK_FILE", "feedback_nexogeek.csv")
_FEEDBACK_LOCK = threading.Lock()

# ── Base de datos LOCAL (opcional, recomendada) ───────────────────────────────
# Si el usuario descarga la base de pokemontcg.io (repo PokemonTCG/pokemon-tcg-data,
# carpeta cards/en renombrada a card_data), la cargamos en memoria y buscamos ahí:
# instantáneo, completo y sin depender de que la API esté en pie.
_DB_CARDS: list = []
_DB_NAMES: list = []       # nombres únicos, para sugerencias "¿quisiste decir…?"
_DB_LOADED = False


def cargar_base_local(carpeta: str) -> int:
    """Carga todos los .json de una carpeta (cada uno es una lista de cartas)."""
    global _DB_CARDS, _DB_NAMES, _DB_LOADED
    cards = []
    for path in sorted(glob.glob(os.path.join(carpeta, "*.json"))):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                cards.extend(data)
            elif isinstance(data, dict) and "data" in data:
                cards.extend(data["data"])
        except Exception:
            continue
    _DB_CARDS = cards
    _DB_NAMES = sorted({c.get("name", "") for c in cards if c.get("name")})
    _DB_LOADED = bool(cards)
    return len(cards)


def _buscar_local(nombre_en: str) -> list:
    """Match parcial por nombre en la base local (imita el comportamiento de la API)."""
    q = nombre_en.strip().lower()
    if not q:
        return []
    return [c for c in _DB_CARDS if q in c.get("name", "").lower()]


def _pool_nombre(nombre_en: str, api_key: str | None) -> list:
    """Pool de cartas por nombre: de la base local si está cargada, si no de la API."""
    if _DB_LOADED:
        return _buscar_local(nombre_en)
    return _raw_query(f'name:"{nombre_en}"', api_key, page_size=50)


def sugerir_nombres(nombre: str, api_key: str | None = None, n: int = 4) -> list:
    """Devuelve nombres de cartas parecidos a 'nombre' (para corregir typos).

    Usa la base local si está cargada (instantáneo y completo). Si no, intenta
    una búsqueda suelta en la API como aproximación.
    """
    q = (nombre or "").strip()
    if not q:
        return []
    universo = _DB_NAMES
    if not universo:
        # Sin base local: aproximación vía API con la primera palabra
        try:
            primera = q.split()[0]
            res = _raw_query(f'name:"{primera}*"', api_key, page_size=25)
            universo = sorted({c.get("name", "") for c in res if c.get("name")})
        except Exception:
            universo = []
    if not universo:
        return []
    # Coincidencias por similitud (typos) + por subcadena (prefijos)
    cercanos = difflib.get_close_matches(q, universo, n=n, cutoff=0.6)
    ql = q.lower()
    subcadena = [x for x in universo if ql in x.lower() and x not in cercanos]
    salida = (cercanos + subcadena)[:n]
    return salida


def _reset_stats():
    with _STATS_LOCK:
        for k in _STATS:
            _STATS[k] = 0 if k != "tiempo_red" else 0.0


def obtener_api_key_streamlit() -> str:
    """Obtiene la API key sin exponerla en el repositorio.

    Prioridad:
      1. Streamlit Community Cloud Secrets.
      2. Variables de entorno locales.

    Acepta varios nombres para facilitar la migración de configuraciones previas.
    """
    nombres = (
        "POKEMONTCG_API_KEY",
        "POKEMON_TCG_API_KEY",
        "pokemontcg_api_key",
    )
    try:
        for nombre in nombres:
            valor = st.secrets.get(nombre, "")
            if valor:
                return str(valor).strip()
    except Exception:
        pass

    for nombre in nombres:
        valor = os.getenv(nombre, "")
        if valor:
            return str(valor).strip()
    return ""


def limpiar_cache_precios() -> None:
    """Vacía únicamente la caché de precios por Card ID."""
    with _PRICE_CACHE_LOCK:
        _PRICE_CACHE.clear()


def diagnostico_api_pokemon(api_key: str | None) -> tuple[bool, str]:
    """Prueba una consulta mínima y devuelve un mensaje legible."""
    if not api_key:
        return False, "No hay una API Key configurada."
    try:
        t0 = time.time()
        resp = requests.get(
            BASE_URL,
            headers={"X-Api-Key": api_key},
            params={"q": 'name:"Pikachu"', "pageSize": 1},
            timeout=12,
        )
        elapsed = int((time.time() - t0) * 1000)
        if resp.status_code == 200:
            return True, f"Conexión correcta · {elapsed} ms"
        if resp.status_code in (401, 403):
            return False, "La API Key fue rechazada. Revisa el secreto configurado."
        if resp.status_code == 429:
            return False, "La API respondió con límite temporal (429). Intenta nuevamente en unos minutos."
        return False, f"La API respondió con código {resp.status_code}."
    except requests.RequestException as exc:
        return False, f"No se pudo conectar con Pokémon TCG API: {exc}"

# ── Estado/condición de la carta (clave para una herramienta de venta) ────────
# Cada estado tiene una descripción y un multiplicador sobre el precio NM.
# Estos multiplicadores son una convención de mercado aproximada y editables.
ESTADOS = {
    "NM":  ("Near Mint",          1.00),
    "LP":  ("Lightly Played",     0.85),
    "MP":  ("Moderately Played",  0.70),
    "HP":  ("Heavily Played",     0.50),
    "DMG": ("Damaged",            0.35),
}
# Alias aceptados en el Excel (español/inglés) → código estándar
ESTADO_ALIASES = {
    "nm": "NM", "near mint": "NM", "mint": "NM", "m": "NM", "nuevo": "NM", "nueva": "NM",
    "lp": "LP", "lightly played": "LP", "light": "LP", "casi nuevo": "LP", "casi nueva": "LP",
    "mp": "MP", "moderately played": "MP", "moderate": "MP", "jugada": "MP", "jugado": "MP",
    "hp": "HP", "heavily played": "HP", "heavy": "HP", "muy jugada": "HP", "muy jugado": "HP",
    "dmg": "DMG", "damaged": "DMG", "dañada": "DMG", "danada": "DMG", "dañado": "DMG", "poor": "DMG",
}
# Colores de badge por estado legibles sobre el fondo claro
ESTADO_COLOR = {
    "NM":  ("#DEF7EC", "#03543F"), "LP": ("#E1EFFE", "#1E429F"),
    "MP":  ("#FEF08A", "#713F12"), "HP": ("#FFEDD5", "#7C2D12"),
    "DMG": ("#FDE8E8", "#9B1C1C"),
}


def traducir_nombre(nombre: str) -> str:
    return TRADUCCIONES_ES_EN.get(nombre.strip().lower(), nombre.strip())


def normalizar_estado(raw) -> str:
    """Convierte cualquier variante de estado a su código estándar (def. NM)."""
    s = str(raw or "").strip().lower()
    if s in ("", "nan", "none"):
        return "NM"
    if s in ESTADO_ALIASES:
        return ESTADO_ALIASES[s]
    return s.upper() if s.upper() in ESTADOS else "NM"


def normalizar_numero(raw) -> str:
    """Normaliza números leídos desde Excel sin romper valores como TG01 o 001."""
    if raw is None:
        return ""
    texto = str(raw).strip()
    if texto.lower() in ("", "nan", "none"):
        return ""
    return re.sub(r"\.0$", "", texto)


def normalizar_columnas(columnas) -> list[str]:
    """Convierte encabezados comunes a los nombres esperados por la aplicación."""
    salida = []
    reemplazos = str.maketrans("áéíóúñ", "aeioun")
    for columna in columnas:
        nombre = str(columna).strip().lower().translate(reemplazos)
        nombre = re.sub(r"[^a-z0-9]+", "_", nombre).strip("_")
        alias = {
            "nombre_carta": "nombre", "card_name": "nombre",
            "clasificacion": "tipo", "card_type": "tipo",
            "n": "numero", "nro": "numero", "numero_carta": "numero",
            "regulation": "regulation_mark", "marca_regulacion": "regulation_mark",
            "condicion": "estado", "condition": "estado",
            "qty": "cantidad", "quantity": "cantidad",
        }
        salida.append(alias.get(nombre, nombre))
    return salida


def _safe_text(value) -> str:
    return escape(str(value if value is not None else ""), quote=True)


def _safe_image_url(value) -> str:
    url = str(value or "").strip()
    try:
        parsed = urlparse(url)
        return _safe_text(url) if parsed.scheme in {"http", "https"} and parsed.netloc else ""
    except Exception:
        return ""


def guardar_feedback(registro: dict, ruta: str = FEEDBACK_FILE) -> None:
    """Guarda feedback del piloto en CSV de forma segura entre hilos."""
    fila = pd.DataFrame([{
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "nombre": str(registro.get("nombre", "")).strip(),
        "perfil": str(registro.get("perfil", "")).strip(),
        "nota": int(registro.get("nota", 0)),
        "utilidad": str(registro.get("utilidad", "")).strip(),
        "mejoraria": str(registro.get("mejoraria", "")).strip(),
        "usaria": str(registro.get("usaria", "")).strip(),
        "contacto": str(registro.get("contacto", "")).strip(),
    }])
    with _FEEDBACK_LOCK:
        existe = os.path.exists(ruta)
        fila.to_csv(ruta, mode="a", header=not existe, index=False, encoding="utf-8-sig")


def leer_feedback(ruta: str = FEEDBACK_FILE) -> pd.DataFrame:
    if not os.path.exists(ruta):
        return pd.DataFrame()
    try:
        return pd.read_csv(ruta)
    except Exception:
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# CAPA API — BÚSQUEDA EN CASCADA
# ══════════════════════════════════════════════════════════════════════════════

def _raw_query(query: str, api_key: str | None, page_size: int = 10) -> list:
    # Normalizar apóstrofes ANTES de todo (caché y petición real)
    query = _normalizar_apostrofe(query)

    # ── Caché global thread-safe (NO usa session_state, que falla entre hilos) ─
    ckey = (query, page_size)
    with _QUERY_LOCK:
        if ckey in _QUERY_CACHE:
            with _STATS_LOCK:
                _STATS["cache_hits"] += 1
            return _QUERY_CACHE[ckey]

    # ── Throttle configurable desde el sidebar (s entre peticiones reales) ────
    time.sleep(_CFG.get("throttle", 0.2) if api_key else max(_CFG.get("throttle", 0.2), 0.6))

    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["X-Api-Key"] = api_key
    params = {"q": query, "pageSize": page_size, "orderBy": "-set.releaseDate"}
    data = []
    t0 = time.time()
    try:
        resp = requests.get(BASE_URL, headers=headers, params=params, timeout=10)
        with _STATS_LOCK:
            _STATS["requests"] += 1
        if resp.status_code == 429:
            with _STATS_LOCK:
                _STATS["rate_limited"] += 1
            time.sleep(3.0)
            resp = requests.get(BASE_URL, headers=headers, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json().get("data", [])
    except Exception:
        with _STATS_LOCK:
            _STATS["errores"] += 1
        data = []
    finally:
        with _STATS_LOCK:
            _STATS["tiempo_red"] += time.time() - t0

    # Solo cacheamos resultados válidos: así un fallo de red puntual no queda
    # "pegado" como si la carta no existiera.
    if data:
        with _QUERY_LOCK:
            _QUERY_CACHE[ckey] = data
    return data

def buscar_carta(nombre_en, tipo, regulation_mark, numero, api_key) -> tuple[list, str]:
    """
    Cascada de 5 niveles:
      1a. nombre + número + bloque    → arte preciso con set confirmado (más exacto)
      1b. nombre + número             → arte por número sin bloque
      2.  nombre + bloque (+ tipo)    → fallback estándar
      3.  nombre + tipo (sin bloque)
      4.  solo nombre                 → salvavidas final
    """
    tiene_numero = bool(numero and numero not in ("", "nan", "-"))
    tiene_mark   = bool(regulation_mark and regulation_mark.lower() not in ("", "nan"))
    tipo_api     = TIPOS_API.get(tipo.strip().lower(), "") if tipo.strip() else ""
    nn = nombre_en.strip().lower()

    def _exactos(lst):
        return [c for c in lst if c.get("name", "").strip().lower() == nn]

    # ── VÍA RÁPIDA: pool por nombre (base local si está cargada; si no, API) ──
    amplia = _pool_nombre(nombre_en, api_key)
    ex = _exactos(amplia)
    if ex:  # encontramos el nombre EXACTO
        if tiene_numero:
            por_num = [c for c in ex if str(c.get("number", "")).lower() == numero.lower()]
            if por_num:
                return por_num, "número exacto"
        if tiene_mark:
            pm = [c for c in ex if str(c.get("regulationMark", "")).upper() == regulation_mark.upper()]
            if tipo_api:
                pm = [c for c in pm if c.get("supertype", "") == tipo_api] or pm
            if pm:
                return pm, "nombre + bloque"
        if tipo_api:
            pt = [c for c in ex if c.get("supertype", "") == tipo_api]
            if pt:
                return pt, "nombre + tipo"
        return ex, "solo nombre"

    # ── FALLBACK DIRIGIDO (solo en modo API; en modo local el pool ya es total)
    if not _DB_LOADED:
        if tiene_numero:
            r = _raw_query(f'name:"{nombre_en}" number:"{numero}"', api_key)
            rex = _exactos(r)
            if rex:
                return rex, "número exacto"
            if r:
                return r, "número exacto"
        if tiene_mark:
            partes = [f'name:"{nombre_en}"', f"regulationMark:{regulation_mark.upper()}"]
            if tipo_api:
                partes.append(f"supertype:{tipo_api}")
            r = _raw_query(" ".join(partes), api_key)
            rex = _exactos(r) or r
            if rex:
                return rex, "nombre + bloque"
        if tipo_api:
            r = _raw_query(f'name:"{nombre_en}" supertype:{tipo_api}', api_key)
            rex = _exactos(r) or r
            if rex:
                return rex, "nombre + tipo"

    # Último recurso: lo que haya devuelto el pool por nombre
    if amplia:
        return amplia, "solo nombre"
    return [], "sin resultados"


# ══════════════════════════════════════════════════════════════════════════════
# SELECCIÓN INTELIGENTE (liga vs. regular)
# ══════════════════════════════════════════════════════════════════════════════

def _tiene_precio(c: dict) -> bool:
    """¿Esta impresión trae algún precio de mercado en tcgplayer?"""
    prices = (c.get("tcgplayer", {}) or {}).get("prices", {}) or {}
    for p in prices.values():
        if p and (p.get("market") or p.get("mid")):
            return True
    return False


def seleccionar_carta(resultados: list, es_de_liga: bool, nombre_en: str = "") -> dict:
    if not resultados:
        return {}

    # Filtro de nombre EXACTO: la API hace match parcial, así "Iono" devuelve
    # también "Iono's Bellibolt ex". Si hay resultados con nombre exacto, solo
    # usamos esos; si no hay ninguno, usamos todos.
    nombre_norm = nombre_en.strip().lower()
    if nombre_norm:
        exactos = [c for c in resultados if c.get("name", "").strip().lower() == nombre_norm]
        if exactos:
            resultados = exactos

    def sn(c):
        return c.get("set", {}).get("name", "").lower()

    # Candidatos según preferencia de liga/promo (o no), con fallback a todos.
    if es_de_liga:
        candidatos = [c for c in resultados if any(kw in sn(c) for kw in KEYWORDS_LIGA)] or resultados
    # CORREGIDO AQUÍ: Se eliminó el texto basura "or whitespaces = resultados" que gatillaba el SyntaxError
    else:
        candidatos = [c for c in resultados if not any(kw in sn(c) for kw in KEYWORDS_LIGA)] or resultados

    con_precio = [c for c in candidatos if _tiene_precio(c)]
    return (con_precio or candidatos)[0]


# ══════════════════════════════════════════════════════════════════════════════
# CONFIANZA DEL MATCH (clave para vender sin equivocarse de versión)
# ══════════════════════════════════════════════════════════════════════════════

def calcular_confianza(metodo: str, tiene_numero: bool, numero_coincide: bool) -> tuple[str, bool]:
    """
    Devuelve (confianza, necesita_revision).
    'necesita_revision' = True SOLO cuando el match de carta es dudoso
    (se pidió un número pero la API trajo otra versión).
    La ausencia de precio NO activa revisión — eso es dato faltante, no match incorrecto.
    """
    if metodo == "sin resultados":
        return "ninguna", True
    if tiene_numero:
        if metodo == "número exacto" and numero_coincide:
            return "alta", False
        # Se pidió un número pero la cascada trajo OTRA versión → revisar arte
        return "baja", True
    # Sin número: no hay certeza de arte exacto, pero no es un error crítico
    if metodo == "nombre + bloque":
        return "media", False
    return "baja", False   # sin número no hay nada que "revisar"


# ══════════════════════════════════════════════════════════════════════════════
# PRECIO DE REFERENCIA (tcgplayer, viene en la misma respuesta de la API)
# ══════════════════════════════════════════════════════════════════════════════

def extraer_precio(card_data: dict) -> tuple:
    """
    Devuelve (precio_usd_mercado, variante, fecha_actualizacion).
    Importante: el precio NO es en tiempo real; tcgplayer lo actualiza ~1 vez al
    día, por eso devolvemos también 'updatedAt' para mostrarlo con honestidad.
    Dentro de cada variante se prueba market → mid → low; solo se pasa a la
    siguiente variante si los tres son None o 0.
    """
    tcg = card_data.get("tcgplayer", {}) or {}
    prices = tcg.get("prices", {}) or {}
    fecha = tcg.get("updatedAt", "-")
    orden = ("holofoil", "normal", "reverseHolofoil", "1stEditionHolofoil",
             "unlimitedHolofoil", "1stEdition", "unlimited")
    for v in orden:
        p = prices.get(v)
        if not p:
            continue
        valor = p.get("market") or p.get("mid") or p.get("low")
        if valor and valor > 0:
            return valor, v, fecha
    if prices:  # fallback: cualquier variante disponible
        for v, p in prices.items():
            valor = p.get("market") or p.get("mid") or p.get("low")
            if valor and valor > 0:
                return valor, v, fecha
    return None, None, fecha


# ══════════════════════════════════════════════════════════════════════════════
# PROCESAMIENTO DE CARTA
# ══════════════════════════════════════════════════════════════════════════════

def _candidato_liviano(c: dict) -> dict:
    """Versión recortada de una carta para guardar como opción de variante."""
    return {
        "id": c.get("id", "-"),
        "name": c.get("name", ""),
        "number": c.get("number", "-"),
        "rarity": c.get("rarity", "-"),
        "regulationMark": c.get("regulationMark", ""),
        "set": {"name": c.get("set", {}).get("name", "-")},
        "images": {"small": c.get("images", {}).get("small", "")},
        "tcgplayer": c.get("tcgplayer", {}) or {},
    }


def _resultado_desde_carta(base: dict, carta: dict, numero: str, tiene_numero: bool,
                           mult: float, clp_rate: float) -> dict:
    """Construye los campos de salida a partir de una carta elegida.
    Reutilizable tanto en el procesamiento automático como en la selección
    manual de variante desde la interfaz."""
    numero_carta = str(carta.get("number", "-"))
    numero_coincide = bool(tiene_numero and numero_carta.lower() == numero.lower())
    conf, revisar = calcular_confianza(base.get("Método Búsqueda", ""), tiene_numero, numero_coincide)

    precio_usd, variante, fecha_precio = extraer_precio(carta)
    precio_aj  = round(precio_usd * mult, 2) if precio_usd else None
    precio_clp = int(round((precio_aj * clp_rate) / 100.0)) * 100 if (precio_aj and clp_rate) else None

    return {
        **base,
        "Card ID":       carta.get("id", "-"),
        "Set":           carta.get("set", {}).get("name", "-"),
        "Número Carta":  numero_carta,
        "Número Coincide": "Sí" if numero_coincide else ("No" if tiene_numero else "-"),
        "Rareza":        carta.get("rarity", "-"),
        "Confianza":     conf,
        "Revisar":       "Sí" if revisar else "No",
        "Precio USD Mercado":  precio_usd,
        "Precio USD Ajustado": precio_aj,
        "Precio CLP Sugerido": precio_clp,
        "Variante Precio":     variante or "-",
        "Fecha Precio":        fecha_precio,
        "URL Imagen":    carta.get("images", {}).get("small", ""),
    }


def fetch_precio_por_id(
    card_id: str,
    api_key: str | None,
    force_refresh: bool = False,
) -> dict | None:
    """Consulta precios de una impresión exacta usando su Card ID.

    La carta y su imagen ya provienen de ``card_data``. Esta función consulta
    solamente el objeto actualizado de la API, donde vienen los precios de
    TCGplayer. Los resultados exitosos se conservan seis horas en memoria.
    """
    card_id = str(card_id or "").strip()
    if not card_id or card_id == "-" or not api_key:
        return None

    ahora = time.time()
    if not force_refresh:
        with _PRICE_CACHE_LOCK:
            cached = _PRICE_CACHE.get(card_id)
            if cached and ahora - cached[0] < _PRICE_CACHE_TTL:
                with _STATS_LOCK:
                    _STATS["cache_hits"] += 1
                return cached[1]

    headers = {
        "Content-Type": "application/json",
        "X-Api-Key": api_key,
    }
    url = f"{BASE_URL}/{card_id}"
    time.sleep(max(0.0, float(_CFG.get("throttle", 0.2))))
    t0 = time.time()

    try:
        resp = requests.get(url, headers=headers, timeout=12)
        with _STATS_LOCK:
            _STATS["requests"] += 1

        if resp.status_code == 429:
            with _STATS_LOCK:
                _STATS["rate_limited"] += 1
            time.sleep(2.5)
            resp = requests.get(url, headers=headers, timeout=12)
            with _STATS_LOCK:
                _STATS["requests"] += 1

        resp.raise_for_status()
        carta = resp.json().get("data")
        if carta:
            with _PRICE_CACHE_LOCK:
                _PRICE_CACHE[card_id] = (time.time(), carta)
            return carta
    except requests.RequestException:
        with _STATS_LOCK:
            _STATS["errores"] += 1
    except (TypeError, ValueError, KeyError):
        with _STATS_LOCK:
            _STATS["errores"] += 1
    finally:
        with _STATS_LOCK:
            _STATS["tiempo_red"] += time.time() - t0
    return None


def enriquecer_precios_en_lote(
    df: "pd.DataFrame",
    api_key: str | None,
    clp_rate: float,
    progress_bar=None,
    status_box=None,
    force_refresh: bool = False,
) -> "pd.DataFrame":
    """Completa precios consultando una sola vez cada Card ID único.

    El motor local identifica cartas e imágenes instantáneamente. Después esta
    función pide a la API únicamente los IDs que aún no tienen precio, reutiliza
    el resultado en filas repetidas y aplica el multiplicador de condición.
    """
    df = df.copy()
    required = {"Card ID", "Precio USD Mercado"}
    if not required.issubset(df.columns):
        if status_box:
            status_box.warning("Faltan columnas necesarias para sincronizar precios.")
        return df

    if not api_key:
        if status_box:
            status_box.warning(
                "Las cartas fueron identificadas con la base local, pero falta configurar "
                "POKEMONTCG_API_KEY en Streamlit Secrets para obtener precios."
            )
        return df

    mask_sin_precio = df["Precio USD Mercado"].apply(
        lambda value: value is None or pd.isna(value)
    )
    card_ids = df["Card ID"].astype(str).str.strip()
    mask_con_id = ~card_ids.isin(["", "-", "None", "nan"])
    indices = df[mask_sin_precio & mask_con_id].index.tolist()

    if not indices:
        if status_box:
            status_box.success("✅ Todas las cartas identificadas ya tienen precio.")
        return df

    ids_unicos = list(dict.fromkeys(card_ids.loc[indices].tolist()))
    total_ids = len(ids_unicos)
    max_workers = max(1, min(int(_CFG.get("max_workers", 4)), 4, total_ids))
    resultados: dict[str, dict | None] = {}

    if status_box:
        status_box.info(
            f"Consultando {total_ids} impresión(es) única(s) para completar "
            f"{len(indices)} fila(s)..."
        )

    def consultar(card_id: str):
        carta = fetch_precio_por_id(card_id, api_key, force_refresh=force_refresh)
        if not carta:
            return card_id, None
        precio_usd, variante, fecha = extraer_precio(carta)
        return card_id, {
            "precio_usd": precio_usd,
            "variante": variante or "-",
            "fecha": fecha,
        }

    completadas = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(consultar, cid): cid for cid in ids_unicos}
        for future in as_completed(futures):
            cid = futures[future]
            try:
                result_id, datos = future.result()
                resultados[result_id] = datos
            except Exception:
                resultados[cid] = None
                with _STATS_LOCK:
                    _STATS["errores"] += 1
            completadas += 1
            if progress_bar:
                progress_bar.progress(completadas / total_ids)

    filas_actualizadas = 0
    for idx in indices:
        card_id = str(df.at[idx, "Card ID"]).strip()
        datos = resultados.get(card_id)
        if not datos or not datos.get("precio_usd"):
            continue

        estado = str(df.at[idx, "Estado"]) if "Estado" in df.columns else "NM"
        _, multiplicador = ESTADOS.get(estado, ("Near Mint", 1.0))
        precio_usd = float(datos["precio_usd"])
        precio_ajustado = round(precio_usd * multiplicador, 2)
        precio_clp = (
            int(round((precio_ajustado * float(clp_rate)) / 100.0)) * 100
            if clp_rate else None
        )

        df.at[idx, "Precio USD Mercado"] = precio_usd
        df.at[idx, "Precio USD Ajustado"] = precio_ajustado
        df.at[idx, "Precio CLP Sugerido"] = precio_clp
        df.at[idx, "Variante Precio"] = datos["variante"]
        df.at[idx, "Fecha Precio"] = datos["fecha"]
        filas_actualizadas += 1

    if progress_bar:
        progress_bar.progress(1.0)
    if status_box:
        ids_con_precio = sum(
            1 for datos in resultados.values()
            if datos and datos.get("precio_usd")
        )
        if filas_actualizadas:
            status_box.success(
                f"✅ {filas_actualizadas}/{len(indices)} fila(s) actualizadas · "
                f"{ids_con_precio}/{total_ids} impresión(es) con precio."
            )
        else:
            status_box.warning(
                "La API respondió, pero no encontró precios TCGplayer para estas impresiones. "
                "Puedes completar precios manualmente o seleccionar otra versión."
            )

    return df

def procesar_carta(fila: dict, api_key: str | None = None, clp_rate: float = 0) -> dict:
    nombre_original = str(fila.get("nombre", "")).strip()
    tipo_carta      = str(fila.get("tipo", "")).strip()
    regulation_mark = str(fila.get("regulation_mark", "")).strip()
    es_liga_raw     = str(fila.get("es_de_liga", "")).strip().lower()
    es_de_liga      = es_liga_raw in ("sí", "si", "yes", "true", "1")

    numero = normalizar_numero(fila.get("numero", ""))

    set_forzado = str(fila.get("set_forzado", "")).strip()
    set_forzado = "" if set_forzado.lower() in ("nan", "none", "") else set_forzado

    estado  = normalizar_estado(fila.get("estado", ""))
    _, mult = ESTADOS.get(estado, ("Near Mint", 1.0))

    try:
        c = fila.get("cantidad", 1)
        cantidad = int(c) if str(c).strip() not in ("", "nan") else 1
    except (ValueError, TypeError):
        cantidad = 1

    nombre_en = traducir_nombre(nombre_original)
    resultados, metodo = buscar_carta(nombre_en, tipo_carta, regulation_mark, numero, api_key)
    tiene_numero = bool(numero)
    mark_disp = regulation_mark if (regulation_mark and regulation_mark.lower() != "nan") else "-"

    base = {
        "Cantidad":        cantidad,
        "Estado":          estado,
        "Nombre Original": nombre_original,
        "Nombre EN":       nombre_en,
        "Tipo":            tipo_carta,
        "Número Buscado":  numero if numero else "-",
        "Regulation Mark": mark_disp,
        "Es de Liga":      "Sí" if es_de_liga else "No",
        "Método Búsqueda": metodo,
    }

    if not resultados:
        conf, revisar = calcular_confianza(metodo, tiene_numero, False)
        return {
            **base,
            "Card ID": "-",
            "Set": "No encontrado", "Número Carta": "-", "Número Coincide": "-",
            "Rareza": "-", "Confianza": conf, "Revisar": "Sí" if revisar else "No",
            "Precio USD Mercado": None, "Precio USD Ajustado": None,
            "Precio CLP Sugerido": None, "Variante Precio": "-", "Fecha Precio": "-",
            "URL Imagen": "", "_candidatos": [],
        }

    nn = nombre_en.strip().lower()
    exactos = [c for c in resultados if c.get("name", "").strip().lower() == nn] or resultados

    carta = None
    if set_forzado:
        sf = set_forzado.lower()
        for c in exactos:
            if sf in c.get("set", {}).get("name", "").lower():
                carta = c
                base["Método Búsqueda"] = "set forzado"
                break
    if carta is None:
        carta = seleccionar_carta(resultados, es_de_liga, nombre_en)

    resultado = _resultado_desde_carta(base, carta, numero, tiene_numero, mult, clp_rate)
    resultado["_candidatos"] = [_candidato_liviano(c) for c in exactos[:8]]
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# FORMATO DE PRECIOS
# ══════════════════════════════════════════════════════════════════════════════

def _has(v) -> bool:
    try:
        return v is not None and pd.notna(v)
    except Exception:
        return v is not None


def _fmt_usd(v) -> str:
    return f"${v:,.2f}" if _has(v) else "-"


def _fmt_clp(v) -> str:
    return f"${v:,.0f}".replace(",", ".") if _has(v) else "-"


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def exportar_excel(df: pd.DataFrame) -> bytes:
    df = df.drop(columns=["Card ID"], errors="ignore")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Catálogo TCG")
        ws = writer.sheets["Catálogo TCG"]

        hf = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hfill = PatternFill("solid", start_color="5B2A86")
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bf = Font(name="Arial", size=10)
        fp = PatternFill("solid", start_color="FFF8ED")
        fi = PatternFill("solid", start_color="FFFFFF")
        fv = PatternFill("solid", start_color="D4EDDA")
        fr = PatternFill("solid", start_color="F8D7DA")
        fl = PatternFill("solid", start_color="FFF3CD")
        fm = PatternFill("solid", start_color="F3E9FF")
        fn = PatternFill("solid", start_color="DDF8F3")
        
        fconf_alta  = PatternFill("solid", start_color="DDF8F3")
        fconf_media = PatternFill("solid", start_color="FFF0CC")
        fconf_baja  = PatternFill("solid", start_color="FFE1D7")
        frev        = PatternFill("solid", start_color="FFE0E0")
        borde = Border(
            left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
        )
        centro = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = borde
        ws.row_dimensions[1].height = 32

        cols = {name: idx + 1 for idx, name in enumerate(df.columns)}
        usd_cols = {cols.get("Precio USD Mercado"), cols.get("Precio USD Ajustado")}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            fb = fp if row_idx % 2 == 0 else fi
            for cell in row:
                cell.font = bf; cell.border = borde
                cell.alignment = Alignment(vertical="center")
                c = cell.column
                if c == cols.get("Confianza"):
                    cell.alignment = centro
                    v = str(cell.value)
                    cell.fill = (fconf_alta if v == "alta" else fconf_media if v == "media" else fconf_baja if v == "baja" else fb)
                elif c == cols.get("Revisar"):
                    cell.alignment = centro
                    cell.fill = frev if str(cell.value) == "Sí" else (fv if str(cell.value) == "No" else fb)
                elif c == cols.get("Es de Liga"):
                    cell.alignment = centro; cell.fill = fl
                elif c == cols.get("Método Búsqueda"):
                    cell.alignment = centro; cell.fill = fm
                    cell.font = Font(name="Arial", size=9, italic=True)
                elif c == cols.get("Número Buscado"):
                    cell.alignment = centro; cell.fill = fn
                elif c in usd_cols:
                    cell.alignment = centro; cell.fill = fb
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0.00'
                elif c == cols.get("Precio CLP Sugerido"):
                    cell.alignment = centro; cell.fill = fb
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0'
                elif c in (cols.get("Cantidad"), cols.get("Estado"), cols.get("Número Coincide"), cols.get("Número Carta")):
                    cell.alignment = centro; cell.fill = fb
                else:
                    cell.fill = fb

        anchos = {
            "Cantidad": 9, "Estado": 9, "Nombre Original": 22, "Nombre EN": 22, "Tipo": 12,
            "Número Buscado": 14, "Regulation Mark": 14, "Es de Liga": 11,
            "Método Búsqueda": 16, "Set": 28, "Número Carta": 13, "Número Coincide": 12,
            "Rareza": 24, "Confianza": 12, "Revisar": 10,
            "Precio USD Mercado": 13, "Precio USD Ajustado": 13, "Precio CLP Sugerido": 15,
            "Variante Precio": 14, "Fecha Precio": 14, "URL Imagen": 48,
        }
        for i, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = anchos.get(col_name, 16)
        ws.freeze_panes = "A2"

    buffer.seek(0)
    return buffer.read()


# ══════════════════════════════════════════════════════════════════════════════
# CSS BASE DE LA INTERFAZ
# ══════════════════════════════════════════════════════════════════════════════
DARK_CSS = """
<style>
html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"], .stApp {
    background-color: #FAFAFA !important; color: #1E293B !important; font-family: 'Inter', sans-serif;
}
/* Barra lateral oscura. No aplicamos color blanco a TODOS los descendientes,
   porque los expanders y campos usan tarjetas claras. */
[data-testid="stSidebar"] { background-color: #0F172A !important; border-right: 1px solid #E2E8F0; color:#F1F5F9; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4 { color:#F8FAFC !important; }
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p { color:#CBD5E1 !important; }
[data-testid="stSidebar"] [data-testid="stExpander"] {
    background:#FFFFFF !important; color:#0F172A !important;
    border:1px solid #CBD5E1 !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] p,
[data-testid="stSidebar"] [data-testid="stExpander"] label,
[data-testid="stSidebar"] [data-testid="stExpander"] summary,
[data-testid="stSidebar"] [data-testid="stExpander"] input,
[data-testid="stSidebar"] [data-testid="stExpander"] textarea,
[data-testid="stSidebar"] [data-testid="stExpander"] [data-baseweb="select"] div {
    color:#0F172A !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] input,
[data-testid="stSidebar"] [data-testid="stExpander"] textarea,
[data-testid="stSidebar"] [data-testid="stExpander"] [data-baseweb="select"] > div {
    background:#F8FAFC !important;
}

/* El Banner Principal (Hero Section) */
.hero-banner {
    background: linear-gradient(135deg, #0B132B 0%, #1C2541 60%, #3A506B 100%);
    padding: 3rem 2.5rem; border-radius: 20px; margin-bottom: 2.5rem; box-shadow: 0 10px 30px rgba(11, 19, 43, 0.15);
    border-left: 6px solid #2563EB;
}
.hero-title { color: #FFFFFF !important; font-size: 2.4rem; font-weight: 800; margin: 0; line-height: 1.2; }
.hero-subtitle { color: #CBD5E1 !important; font-size: 1.05rem; margin-top: 10px; font-weight: 400; }

/* Caja de marcas de juegos (Hover responsivo) */
.brand-box {
    background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 14px; padding: 20px;
    text-align: center; box-shadow: 0 4px 10px rgba(0,0,0,0.02); transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    display: flex; flex-direction: column; align-items: center; justify-content: center; height: 100px; cursor: pointer;
}
.brand-box:hover { transform: translateY(-4px); box-shadow: 0 12px 20px rgba(0,0,0,0.08); border-color: #3B82F6; }
.brand-title { color: #1E293B !important; font-weight: 800; font-size: 1.1rem; letter-spacing: 0.03em; margin: 0; }

/* Tarjetas Blancas Flotantes de Productos */
.product-card {
    background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 16px; padding: 18px;
    text-align: center; box-shadow: 0 4px 15px rgba(0,0,0,0.03); transition: all 0.25s ease; height: 100%;
    display: flex; flex-direction: column; justify-content: space-between;
}
.product-card:hover { transform: translateY(-5px); box-shadow: 0 15px 30px rgba(0,0,0,0.09); border-color: #CBD5E1; }
.product-title { font-size: 0.95rem; font-weight: 700; color: #0F172A; margin: 10px 0 4px 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.product-meta { font-size: 0.78rem; color: #64748B; margin: 0 0 14px 0; }
.product-price { background: #F1F5F9; color: #1E3A8A; font-weight: 800; font-size: 1.15rem; padding: 8px; border-radius: 10px; border: 1px solid #E2E8F0; }

/* Ajustes de componentes nativos de Streamlit para que combinen con el fondo claro */
div[data-testid="stExpander"] { background-color: #FFFFFF !important; border: 1px solid #E2E8F0 !important; border-radius: 12px !important; }
.stTabs [data-baseweb="tab-list"] { background-color: transparent !important; }
button[data-baseweb="tab"] { font-weight: 600 !important; font-size: 0.95rem !important; padding: 12px 20px !important; color: #475569 !important; }
button[data-baseweb="tab"][aria-selected="true"] { color: #2563EB !important; border-bottom-color: #2563EB !important; }
</style>
"""

def _clean(html: str) -> str:
    return "".join(linea.strip() for linea in html.splitlines())

def texto_publicacion(row: dict, mi_precio=None) -> tuple[str, str]:
    nom = row.get("Nombre Original", "-")
    set_ = row.get("Set", "-")
    num = row.get("Número Carta", "-")
    est = row.get("Estado", "NM")
    est_desc = ESTADOS.get(est, ("", 0))[0]
    rar = row.get("Rareza", "-")
    mark = row.get("Regulation Mark", "-")
    precio = mi_precio if _has(mi_precio) else row.get("Precio CLP Sugerido")

    titulo = f"{nom} — {set_} #{num} ({est})"
    lineas = [
        f"🃏 {nom}",
        f"📦 Set: {set_}  ·  N° {num}  ·  Bloque {mark}",
        f"⭐ Rareza: {rar}",
        f"🛡️ Estado: {est} ({est_desc})",
    ]
    if _has(precio):
        lineas.append(f"💲 Precio: {_fmt_clp(precio)} CLP")
    return titulo, "\n".join(lineas)

def validar_entrada(df) -> tuple[list, list]:
    errores, advert = [], []
    if df is None or len(df) == 0:
        errores.append("El archivo está vacío.")
        return errores, advert

    cols = set(df.columns)
    for req in ("nombre", "tipo"):
        if req not in cols:
            errores.append(f"Falta la columna obligatoria «{req}».")
    if errores:
        return errores, advert

    vacios = df["nombre"].fillna("").astype(str).str.strip().eq("")
    if vacios.any():
        errores.append(f"Hay {int(vacios.sum())} fila(s) sin nombre de carta.")

    tipos_validos = {"pokemon", "pokémon", "trainer", "entrenador", "energy", "energia", "energía"}
    tipos = df["tipo"].fillna("").astype(str).str.strip().str.lower()
    invalidos = ~tipos.isin(tipos_validos)
    if invalidos.any():
        ejemplos = ", ".join(sorted(set(tipos[invalidos].tolist()))[:4]) or "vacío"
        errores.append(f"Hay tipos de carta no reconocidos: {ejemplos}.")

    if "numero" not in cols:
        advert.append("No se incluyó la columna «numero»; algunos artes quedarán con confianza baja.")
    elif df["numero"].fillna("").astype(str).str.strip().eq("").any():
        advert.append("Algunas filas no tienen número de carta; revisa visualmente esas coincidencias.")

    if "cantidad" in cols:
        cantidades = pd.to_numeric(df["cantidad"], errors="coerce")
        if cantidades.isna().any() or (cantidades < 1).any():
            advert.append("Las cantidades vacías o inválidas se reemplazarán por 1.")

    return errores, advert

def _badge_metodo(metodo: str) -> str:
    paleta = {
        "número exacto":   ("#EBF8FF", "#2B6CB0", "🎯"), "nombre + bloque": ("#E6FFFA", "#319795", "📦"),
        "nombre + tipo":   ("#EBF4FF", "#5A67D8", "🔤"), "solo nombre":     ("#FEF3C7", "#D97706", "🔍"),
        "set forzado":     ("#E2E8F0", "#4A5568", "📌"), "selección manual":("#EBF8FF", "#2B6CB0", "✋"),
        "sin resultados":  ("#FEE2E2", "#EF4444", "⚠️"),
    }
    bg, fg, ico = paleta.get(metodo, ("#F1F5F9", "#475569", "•"))
    return f'<span style="background:{bg};color:{fg};font-size:0.65rem;font-weight:600;padding:3px 8px;border-radius:20px;margin-right:4px;">{ico} {metodo}</span>'

def _badge_confianza(conf: str) -> str:
    m = {"alta": ("#DEF7EC", "#03543F", "🟢 Alta"), "media": ("#FEF08A", "#713F12", "🟡 Media"), "baja": ("#FDE8E8", "#9B1C1C", "🔴 Baja"), "ninguna": ("#F1F5F9", "#475569", "⚪ Sin datos")}
    bg, fg, txt = m.get(conf, m["ninguna"])
    return f'<span style="background:{bg};color:{fg};font-size:0.65rem;font-weight:700;padding:3px 8px;border-radius:20px;">{txt}</span>'

def _badge_estado(est: str) -> str:
    bg, fg = ESTADO_COLOR.get(est, ("#F1F5F9", "#475569"))
    return f'<span style="background:{bg};color:{fg};font-size:0.65rem;font-weight:700;padding:3px 8px;border-radius:20px;">{est}</span>'

def _precio_html(row: dict) -> str:
    usd = row.get("Precio USD Mercado")
    clp = row.get("Precio CLP Sugerido")
    if not _has(usd):
        return '<div style="text-align:center;font-size:0.7rem;color:#64748B;margin-bottom:6px;">— sin precio —</div>'
    return f'<div class="product-price">{_fmt_clp(clp)}<br><span style="color:#64748B;font-size:0.68rem;font-weight:400;">Ref: {_fmt_usd(usd)} USD</span></div>'

def _tarjeta_html(row: dict) -> str:
    img = _safe_image_url(row.get("URL Imagen", ""))
    met = str(row.get("Método Búsqueda", ""))
    cant = _safe_text(row.get("Cantidad", 1))
    set_ = _safe_text(row.get("Set", "-"))
    nom = _safe_text(row.get("Nombre Original", row.get("Nombre EN", "")))
    num = _safe_text(row.get("Número Carta", "-"))
    mark = _safe_text(row.get("Regulation Mark", "-"))
    conf = str(row.get("Confianza", "ninguna"))
    est = str(row.get("Estado", "NM"))
    revisar = row.get("Revisar") == "Sí"
    borde = "#EF4444" if revisar else "#E2E8F0"
    ribbon = '<div style="background:#FEE2E2;color:#9B1C1C;font-size:0.65rem;font-weight:700;text-align:center;border-radius:6px;padding:4px;margin-bottom:10px;">⚠️ REVISAR VERSIÓN</div>' if revisar else ''
    img_html = f'<img src="{img}" alt="Carta {nom}" style="width:100%;max-width:130px;border-radius:8px;display:block;margin:0 auto 10px;box-shadow: 0 4px 10px rgba(0,0,0,0.06);">' if img else '<div style="width:110px;height:150px;background:#F1F5F9;color:#94A3B8;border-radius:8px;margin:0 auto 10px;display:flex;align-items:center;justify-content:center;font-size:1.5rem;">🃏</div>'

    return _clean(f"""
<div class="product-card" style="border-color: {borde};">
    <div>
        {ribbon}
        {img_html}
        <div class="product-title" title="{nom}">{nom}</div>
        <div class="product-meta">{set_} (#{num})</div>
        <div style="text-align:center;margin-bottom:8px;">{_badge_estado(est)}</div>
        <div style="background:#F8FAFC;border-radius:8px;padding:6px;font-size:0.7rem;color:#475569;line-height:1.5;margin-bottom:10px;border:1px solid #F1F5F9;">
            <span style="color:#64748B;">Cantidad:</span> <b>×{cant}</b> | <span style="color:#64748B;">Bloque:</span> {mark}
        </div>
    </div>
    <div>
        {_precio_html(row)}
        <div style="text-align:center;margin-top:8px;">{_badge_metodo(met)}{_badge_confianza(conf)}</div>
    </div>
</div>
""")

def _rebuild_row(row: dict, carta: dict, clp_rate: float) -> dict:
    """Recalcula una fila a partir de una carta elegida manualmente."""
    estado = row.get("Estado", "NM")
    _, mult = ESTADOS.get(estado, ("", 1.0))
    numero = row.get("Número Buscado", "-")
    numero = "" if numero in ("-", "") else numero
    base = {k: row.get(k) for k in ["Cantidad", "Estado", "Nombre Original", "Nombre EN",
                                     "Tipo", "Número Buscado", "Regulation Mark", "Es de Liga"]}
    base["Método Búsqueda"] = "selección manual"
    nuevo = _resultado_desde_carta(base, carta, numero, bool(numero), mult, clp_rate)
    # La selección manual valida la versión visualmente, pero no falsea si el
    # número elegido es distinto al solicitado originalmente.
    nuevo["Revisar"] = "No"
    nuevo["Confianza"] = "alta"
    return nuevo


def _inventory_publication_key(row: dict) -> str:
    """Identificador estable para agrupar la misma impresión y evitar duplicados."""
    card_id = str(row.get("Card ID", "") or "").strip()
    if card_id.lower() in {"", "-", "nan", "none"}:
        card_id = "|".join([
            str(row.get("Nombre EN", row.get("Nombre Original", ""))).strip().lower(),
            str(row.get("Set", "")).strip().lower(),
            str(row.get("Número Carta", "")).strip().lower(),
        ])
    estado = str(row.get("Estado", "NM") or "NM").strip().upper()
    variante = str(row.get("Variante Precio", "-") or "-").strip().lower()
    return f"{card_id}|{estado}|{variante}"


def _inventory_signature(df: pd.DataFrame) -> str:
    """Firma del inventario para reiniciar el editor solo cuando cambian los resultados."""
    cols = [c for c in [
        "Card ID", "Nombre Original", "Set", "Número Carta", "Estado", "Cantidad",
        "Precio CLP Sugerido", "Variante Precio", "Revisar", "URL Imagen", "Publicado",
    ] if c in df.columns]
    if not cols:
        return "empty"
    normalized = df[cols].fillna("").astype(str)
    payload = normalized.to_csv(index=True).encode("utf-8")
    return hashlib.sha1(payload).hexdigest()[:16]


def _prepare_bulk_publication_table(df_result: pd.DataFrame) -> pd.DataFrame:
    """Agrupa impresiones compatibles y crea la tabla editable de publicación masiva."""
    groups: dict[str, dict] = {}
    for idx, row in df_result.iterrows():
        data = row.to_dict()
        key = _inventory_publication_key(data)
        qty_raw = pd.to_numeric(pd.Series([data.get("Cantidad", 1)]), errors="coerce").iloc[0]
        qty = max(1, int(qty_raw)) if pd.notna(qty_raw) else 1
        price_raw = pd.to_numeric(pd.Series([data.get("Precio CLP Sugerido", 0)]), errors="coerce").iloc[0]
        price = max(0, int(price_raw)) if pd.notna(price_raw) else 0
        revisar = str(data.get("Revisar", "No")) == "Sí"
        card_id = str(data.get("Card ID", "") or "").strip()
        valid_match = card_id.lower() not in {"", "-", "nan", "none"} and str(data.get("Set", "")) != "No encontrado"
        already_published = str(data.get("Publicado", "No")) == "Sí"

        if key not in groups:
            status = "Lista para publicar"
            if revisar:
                status = "Revisar coincidencia"
            elif not valid_match:
                status = "Sin coincidencia válida"
            elif price <= 0:
                status = "Completar precio"
            elif already_published:
                status = "Ya publicada"
            groups[key] = {
                "Publicar": bool(valid_match and not revisar and price > 0 and not already_published),
                "Imagen": str(data.get("URL Imagen", "") or ""),
                "Carta": str(data.get("Nombre Original", data.get("Nombre EN", "Carta"))),
                "Set": str(data.get("Set", "-")),
                "Número": str(data.get("Número Carta", "-")),
                "Estado": str(data.get("Estado", "NM")),
                "Disponible": qty,
                "Cantidad": qty,
                "Precio CLP": price,
                "Confianza": str(data.get("Confianza", "ninguna")),
                "Estado publicación": status,
                "_inventory_key": key,
                "_source_indices": [int(idx)],
            }
        else:
            groups[key]["Disponible"] += qty
            groups[key]["Cantidad"] += qty
            groups[key]["_source_indices"].append(int(idx))
            # Si alguna fila del grupo exige revisión, el grupo completo queda bloqueado.
            if revisar:
                groups[key]["Publicar"] = False
                groups[key]["Estado publicación"] = "Revisar coincidencia"
            elif already_published and groups[key]["Estado publicación"] == "Lista para publicar":
                groups[key]["Publicar"] = False
                groups[key]["Estado publicación"] = "Ya publicada"

    records = list(groups.values())
    if not records:
        return pd.DataFrame()
    table = pd.DataFrame(records)
    table["Valor lote"] = table["Cantidad"] * table["Precio CLP"]
    return table


def _init_bulk_publication_state(df_result: pd.DataFrame) -> pd.DataFrame:
    signature = _inventory_signature(df_result)
    if st.session_state.get("bulk_publish_signature") != signature:
        st.session_state["bulk_publish_signature"] = signature
        st.session_state["bulk_publish_table"] = _prepare_bulk_publication_table(df_result)
        st.session_state["bulk_publish_editor_version"] = st.session_state.get("bulk_publish_editor_version", 0) + 1
    table = st.session_state.get("bulk_publish_table")
    return table.copy() if isinstance(table, pd.DataFrame) else pd.DataFrame()


def _set_bulk_selection(mode: str) -> None:
    table = st.session_state.get("bulk_publish_table")
    if not isinstance(table, pd.DataFrame) or table.empty:
        return
    eligible = table["Estado publicación"].isin(["Lista para publicar", "Completar precio"])
    if mode == "all":
        table.loc[:, "Publicar"] = eligible & (pd.to_numeric(table["Precio CLP"], errors="coerce").fillna(0) > 0)
    elif mode == "none":
        table.loc[:, "Publicar"] = False
    elif mode == "priced":
        table.loc[:, "Publicar"] = eligible & (pd.to_numeric(table["Precio CLP"], errors="coerce").fillna(0) > 0)
    elif mode == "nm_lp":
        table.loc[:, "Publicar"] = eligible & table["Estado"].isin(["NM", "LP"]) & (pd.to_numeric(table["Precio CLP"], errors="coerce").fillna(0) > 0)
    st.session_state["bulk_publish_table"] = table
    st.session_state["bulk_publish_editor_version"] = st.session_state.get("bulk_publish_editor_version", 0) + 1


def _apply_bulk_price_rule(percent: float, round_to: int, minimum: int) -> None:
    table = st.session_state.get("bulk_publish_table")
    if not isinstance(table, pd.DataFrame) or table.empty:
        return
    prices = pd.to_numeric(table["Precio CLP"], errors="coerce").fillna(0)
    adjusted = prices * (1 + percent / 100.0)
    if round_to > 0:
        adjusted = (adjusted / round_to).round() * round_to
    adjusted = adjusted.where(prices <= 0, adjusted.clip(lower=minimum))
    table["Precio CLP"] = adjusted.round().astype(int)
    table["Valor lote"] = pd.to_numeric(table["Cantidad"], errors="coerce").fillna(0).astype(int) * table["Precio CLP"]
    st.session_state["bulk_publish_table"] = table
    st.session_state["bulk_publish_editor_version"] = st.session_state.get("bulk_publish_editor_version", 0) + 1


def _publish_inventory_batch(df_result: pd.DataFrame, edited: pd.DataFrame, location: str,
                             shipping: str, negotiable: bool, description_template: str) -> tuple[int, int, list[str]]:
    """Crea o actualiza publicaciones y devuelve (creadas, actualizadas, omitidas)."""
    created = 0
    updated = 0
    skipped: list[str] = []
    marketplace = st.session_state.setdefault("marketplace_db", [])
    selected = edited[edited["Publicar"] == True].copy()  # noqa: E712
    batch_id = _new_id("batch")

    for _, item in selected.iterrows():
        price = int(pd.to_numeric(pd.Series([item.get("Precio CLP", 0)]), errors="coerce").fillna(0).iloc[0])
        available = int(pd.to_numeric(pd.Series([item.get("Disponible", 1)]), errors="coerce").fillna(1).iloc[0])
        quantity = int(pd.to_numeric(pd.Series([item.get("Cantidad", available)]), errors="coerce").fillna(available).iloc[0])
        quantity = max(1, min(quantity, available))
        status = str(item.get("Estado publicación", ""))
        if status in {"Revisar coincidencia", "Sin coincidencia válida"}:
            skipped.append(f"{item.get('Carta', 'Carta')}: {status.lower()}")
            continue
        if price <= 0:
            skipped.append(f"{item.get('Carta', 'Carta')}: falta precio")
            continue

        source_indices = item.get("_source_indices", [])
        if isinstance(source_indices, str):
            source_indices = [int(x) for x in source_indices.split(",") if x.strip().isdigit()]
        source_indices = list(source_indices) if isinstance(source_indices, (list, tuple)) else []
        if not source_indices:
            skipped.append(f"{item.get('Carta', 'Carta')}: origen no identificado")
            continue
        source_idx = int(source_indices[0])
        fila = df_result.loc[source_idx].to_dict()
        inventory_key = str(item.get("_inventory_key", _inventory_publication_key(fila)))
        description = description_template.format(
            carta=str(item.get("Carta", "Carta")),
            set=str(item.get("Set", "-")),
            numero=str(item.get("Número", "-")),
            estado=str(item.get("Estado", "NM")),
            cantidad=quantity,
        )

        existing = next((x for x in marketplace if x.get("owner") and x.get("inventory_key") == inventory_key), None)
        if existing:
            existing.update({
                "price": price,
                "stock": quantity,
                "condition": str(item.get("Estado", fila.get("Estado", "NM"))),
                "location": location,
                "shipping": shipping,
                "negotiable": bool(negotiable),
                "description": description,
                "image": str(item.get("Imagen", fila.get("URL Imagen", ""))),
                "active": True,
                "batch_id": batch_id,
            })
            listing_id = existing["id"]
            updated += 1
        else:
            listing = _listing_from_inventory(
                fila=fila,
                price=price,
                quantity=quantity,
                condition=str(item.get("Estado", fila.get("Estado", "NM"))),
                location=location,
                shipping=shipping,
                negotiable=negotiable,
                description=description,
                track=False,
            )
            listing.update({
                "inventory_key": inventory_key,
                "batch_id": batch_id,
                "source_rows": source_indices,
                "tags": list(dict.fromkeys(listing.get("tags", []) + ["Publicación masiva"])),
            })
            marketplace.insert(0, listing)
            listing_id = listing["id"]
            created += 1

        for idx in source_indices:
            if idx in df_result.index:
                df_result.at[idx, "Publicado"] = "Sí"
                df_result.at[idx, "ID Publicación"] = listing_id

    st.session_state["marketplace_db"] = marketplace
    st.session_state["df_result"] = df_result
    _track_event("publicar_lote", batch_id, f"creadas={created}; actualizadas={updated}; omitidas={len(skipped)}")
    return created, updated, skipped


def render_dashboard(df_result: pd.DataFrame, clp_rate: float = 0, comision: float = 0.0):
    """Panel de inventario y puente directo hacia el marketplace de la demo."""
    st.markdown("---")
    st.markdown("<h3 style='color:#0F172A;font-weight:800;margin-bottom:12px;'>📊 Panel inteligente de inventario</h3>", unsafe_allow_html=True)

    total_u = int(pd.to_numeric(df_result.get("Cantidad", 0), errors="coerce").fillna(0).sum())
    alta = int((df_result.get("Confianza", pd.Series(dtype=str)) == "alta").sum())
    por_rev = int((df_result.get("Revisar", pd.Series(dtype=str)) == "Sí").sum())
    ref_total = int(df_result.apply(
        lambda r: (r.get("Precio CLP Sugerido", 0) * r.get("Cantidad", 1))
        if pd.notna(r.get("Precio CLP Sugerido")) else 0,
        axis=1,
    ).sum())

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("📦 Unidades", total_u)
    m2.metric("🟢 Alta confianza", alta)
    m3.metric("⚠️ Por revisar", por_rev)
    m4.metric("💰 Valor referencia", _fmt_clp(ref_total))
    m5.metric("📣 Publicables", int((df_result.get("Revisar", "Sí") == "No").sum()))

    df_export = df_result.drop(columns=["_candidatos"], errors="ignore")
    exp1, exp2, exp3 = st.columns([1, 1, 1])
    exp1.download_button(
        "📥 Exportar Excel",
        data=exportar_excel(df_export),
        file_name="catalogo_nexogeek.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    exp2.download_button(
        "📄 Exportar CSV",
        data=df_export.to_csv(index=False).encode("utf-8-sig"),
        file_name="catalogo_nexogeek.csv",
        mime="text/csv",
        use_container_width=True,
    )
    if exp3.button("🛒 Ver marketplace", use_container_width=True, key="dash_go_market"):
        _go_to("Marketplace")

    df_rev = df_result[df_result.get("Revisar", "No") == "Sí"]
    if not df_rev.empty:
        st.error(f"⚠️ Hay {len(df_rev)} carta(s) que requieren verificación manual antes de publicarlas.")
        with st.expander("Corregir coincidencias ambiguas", expanded=True):
            cols_rev = [c for c in ["Nombre Original", "Número Buscado", "Número Carta", "Número Coincide", "Set", "Método Búsqueda", "Confianza"] if c in df_rev.columns]
            st.dataframe(df_rev[cols_rev], use_container_width=True, hide_index=True)

            candidatos_all = st.session_state.get("candidatos", {})
            corregibles = [i for i in df_rev.index if len(candidatos_all.get(i, [])) > 1]
            for idx in corregibles:
                row = df_result.loc[idx].to_dict()
                cands = candidatos_all.get(idx, [])
                labels = [
                    f"{c['set']['name']} · #{c['number']} · {c.get('rarity', '-')} · marca {c.get('regulationMark') or '-'}"
                    + (" · con precio" if _tiene_precio(c) else " · sin precio")
                    for c in cands
                ]
                default_i = next((j for j, c in enumerate(cands)
                                  if c['set']['name'] == row.get('Set') and str(c['number']) == str(row.get('Número Carta'))), 0)
                ca, cb = st.columns([4, 1])
                with ca:
                    sel_var = st.selectbox(
                        f"Versión de «{row.get('Nombre Original')}»",
                        labels,
                        index=default_i,
                        key=f"variant_{idx}",
                    )
                with cb:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    if st.button("✅ Usar", key=f"usevar_{idx}", use_container_width=True):
                        chosen = cands[labels.index(sel_var)]
                        nuevo = _rebuild_row(row, chosen, clp_rate)
                        for k, v in nuevo.items():
                            df_result.at[idx, k] = v
                        st.session_state["df_result"] = df_result
                        st.rerun()

    subtabs = st.tabs([
        "🃏 Cuadrícula", "📋 Tabla", "🔍 Ficha técnica",
        "💲 Simulador de venta", "📣 Publicar al marketplace",
    ])

    with subtabs[0]:
        n_cols = st.select_slider("Columnas visibles", options=[2, 3, 4, 5], value=4, key="cols_slider")
        filas_grid = [df_result.iloc[i:i+n_cols] for i in range(0, len(df_result), n_cols)]
        for fila_df in filas_grid:
            cols_st = st.columns(n_cols)
            for col_st, (_, row) in zip(cols_st, fila_df.iterrows()):
                with col_st:
                    st.markdown(_tarjeta_html(row.to_dict()), unsafe_allow_html=True)
                    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    with subtabs[1]:
        st.dataframe(df_export, use_container_width=True, height=430, hide_index=True)

    with subtabs[2]:
        opciones = [f"{i+1}. {r['Nombre Original']} (#{r['Número Carta']})" for i, (_, r) in enumerate(df_result.iterrows())]
        sel = st.selectbox("Carta a inspeccionar", opciones, key="inspect_card")
        idx_sel = opciones.index(sel)
        row_sel = df_result.iloc[idx_sel].to_dict()

        il, ir = st.columns([1, 2])
        with il:
            if _safe_image_url(row_sel.get("URL Imagen")):
                st.image(row_sel.get("URL Imagen"), width=250)
            else:
                st.info("Sin imagen disponible")
        with ir:
            st.markdown(f"### {_safe_text(row_sel.get('Nombre Original', '-'))}", unsafe_allow_html=True)
            st.markdown(
                f"**Set:** {_safe_text(row_sel.get('Set', '-'))}  ·  "
                f"**Rareza:** {_safe_text(row_sel.get('Rareza', '-'))}  ·  "
                f"**Estado:** {_safe_text(row_sel.get('Estado', '-'))}",
                unsafe_allow_html=True,
            )
            st.metric("Precio sugerido", _fmt_clp(row_sel.get("Precio CLP Sugerido")))
            t_pub, d_pub = texto_publicacion(row_sel)
            st.text_area("Texto listo para publicar", f"{t_pub}\n\n{d_pub}", height=150, key="copy_listing_text")

    with subtabs[3]:
        cols_ro = [c for c in ["Nombre Original", "Set", "Número Carta", "Estado", "Cantidad", "Precio CLP Sugerido"] if c in df_result.columns]
        edf = df_result[cols_ro].copy()
        edf["Mi Precio CLP"] = pd.to_numeric(df_result["Precio CLP Sugerido"], errors="coerce").fillna(0).astype(int)
        edited = st.data_editor(
            edf,
            hide_index=True,
            use_container_width=True,
            disabled=cols_ro,
            key="editor_precios",
        )
        mip = pd.to_numeric(edited["Mi Precio CLP"], errors="coerce").fillna(0)
        cant = pd.to_numeric(edited.get("Cantidad", 1), errors="coerce").fillna(1)
        neto_unit = (mip * (1 - comision / 100)).round().astype(int)
        c1, c2, c3 = st.columns(3)
        c1.metric("💵 Venta bruta", _fmt_clp(int((mip * cant).sum())))
        c2.metric("💰 Neto estimado", _fmt_clp(int((neto_unit * cant).sum())))
        c3.metric("🧾 Comisión estimada", _fmt_clp(int(((mip - neto_unit) * cant).sum())))
        st.caption("Simulación para validar el modelo comercial. No procesa pagos reales.")

    with subtabs[4]:
        st.markdown("#### Publicación masiva desde el inventario")
        st.caption(
            "Las cartas aptas quedan seleccionadas automáticamente. Desmarca únicamente las que no quieras vender, "
            "ajusta cantidades o precios y confirma el lote. Las filas dudosas se bloquean para proteger la publicación."
        )

        bulk_table = _init_bulk_publication_state(df_result)
        if bulk_table.empty:
            st.warning("No hay cartas disponibles para preparar publicaciones.")
        else:
            ready = int((bulk_table["Estado publicación"] == "Lista para publicar").sum())
            review = int((bulk_table["Estado publicación"] == "Revisar coincidencia").sum())
            missing_price = int((bulk_table["Estado publicación"] == "Completar precio").sum())
            already = int((bulk_table["Estado publicación"] == "Ya publicada").sum())
            grouped_units = int(pd.to_numeric(bulk_table["Disponible"], errors="coerce").fillna(0).sum())

            b1, b2, b3, b4, b5 = st.columns(5)
            b1.metric("Impresiones agrupadas", len(bulk_table))
            b2.metric("Unidades", grouped_units)
            b3.metric("Listas", ready)
            b4.metric("Por revisar", review)
            b5.metric("Sin precio / publicadas", missing_price + already)

            st.markdown("##### Selección rápida")
            q1, q2, q3, q4 = st.columns(4)
            if q1.button("Seleccionar aptas", use_container_width=True, key="bulk_select_all"):
                _set_bulk_selection("all")
                st.rerun()
            if q2.button("Deseleccionar todas", use_container_width=True, key="bulk_select_none"):
                _set_bulk_selection("none")
                st.rerun()
            if q3.button("Solo con precio", use_container_width=True, key="bulk_select_priced"):
                _set_bulk_selection("priced")
                st.rerun()
            if q4.button("Solo NM y LP", use_container_width=True, key="bulk_select_nm_lp"):
                _set_bulk_selection("nm_lp")
                st.rerun()

            with st.expander("Ajustar precios del lote", expanded=False):
                pr1, pr2, pr3, pr4 = st.columns(4)
                price_percent = pr1.number_input(
                    "Ajuste sobre referencia (%)", min_value=-50.0, max_value=100.0,
                    value=0.0, step=1.0, key="bulk_price_percent",
                )
                round_to = pr2.selectbox("Redondear a", [100, 500, 1000], index=1, key="bulk_round_to")
                minimum = pr3.number_input("Precio mínimo", min_value=0, value=500, step=500, key="bulk_minimum_price")
                pr4.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if pr4.button("Aplicar al lote", type="primary", use_container_width=True, key="bulk_apply_price"):
                    _apply_bulk_price_rule(float(price_percent), int(round_to), int(minimum))
                    st.rerun()

            display_cols = [
                "Publicar", "Imagen", "Carta", "Set", "Número", "Estado",
                "Disponible", "Cantidad", "Precio CLP", "Confianza", "Estado publicación",
                "_inventory_key", "_source_indices",
            ]
            table_for_editor = st.session_state["bulk_publish_table"][display_cols].copy()
            # Las listas internas se serializan para que el editor pueda mantenerlas sin problemas.
            table_for_editor["_source_indices"] = table_for_editor["_source_indices"].apply(
                lambda values: ",".join(str(x) for x in values) if isinstance(values, (list, tuple)) else str(values)
            )
            editor_version = st.session_state.get("bulk_publish_editor_version", 0)
            edited = st.data_editor(
                table_for_editor,
                hide_index=True,
                use_container_width=True,
                height=min(650, 155 + len(table_for_editor) * 35),
                key=f"bulk_publish_editor_{editor_version}",
                disabled=[
                    "Imagen", "Carta", "Set", "Número", "Disponible",
                    "Confianza", "Estado publicación", "_inventory_key", "_source_indices",
                ],
                column_config={
                    "Publicar": st.column_config.CheckboxColumn(
                        "Publicar", help="Desmarca las cartas que deseas conservar.", default=True,
                    ),
                    "Imagen": st.column_config.ImageColumn("Carta", width="small"),
                    "Carta": st.column_config.TextColumn("Nombre", width="medium"),
                    "Set": st.column_config.TextColumn("Set", width="medium"),
                    "Número": st.column_config.TextColumn("N°", width="small"),
                    "Estado": st.column_config.SelectboxColumn("Estado", options=list(ESTADOS.keys()), required=True),
                    "Disponible": st.column_config.NumberColumn("Stock", min_value=1, format="%d"),
                    "Cantidad": st.column_config.NumberColumn("Publicar", min_value=1, step=1, format="%d"),
                    "Precio CLP": st.column_config.NumberColumn("Precio unitario", min_value=0, step=100, format="$%d"),
                    "Confianza": st.column_config.TextColumn("Match", width="small"),
                    "Estado publicación": st.column_config.TextColumn("Validación", width="medium"),
                    "_inventory_key": None,
                    "_source_indices": None,
                },
            )
            # Guardar cambios para que los controles rápidos y reglas de precio trabajen sobre la última edición.
            edited_state = edited.copy()
            edited_state["_source_indices"] = edited_state["_source_indices"].apply(
                lambda value: [int(x) for x in str(value).split(",") if x.strip().isdigit()]
            )
            edited_state["Valor lote"] = (
                pd.to_numeric(edited_state["Cantidad"], errors="coerce").fillna(0)
                * pd.to_numeric(edited_state["Precio CLP"], errors="coerce").fillna(0)
            ).astype(int)
            st.session_state["bulk_publish_table"] = edited_state

            selected = edited_state[edited_state["Publicar"] == True].copy()  # noqa: E712
            selected_value = int(selected["Valor lote"].sum()) if not selected.empty else 0
            selected_units = int(pd.to_numeric(selected.get("Cantidad", 0), errors="coerce").fillna(0).sum()) if not selected.empty else 0
            selected_valid = selected[
                ~selected["Estado publicación"].isin(["Revisar coincidencia", "Sin coincidencia válida"])
                & (pd.to_numeric(selected["Precio CLP"], errors="coerce").fillna(0) > 0)
            ]

            s1, s2, s3, s4 = st.columns(4)
            s1.metric("Publicaciones elegidas", len(selected))
            s2.metric("Aptas para confirmar", len(selected_valid))
            s3.metric("Unidades seleccionadas", selected_units)
            s4.metric("Valor total publicado", _fmt_clp(selected_value))

            st.markdown("##### Condiciones comunes del lote")
            c1, c2, c3 = st.columns(3)
            bulk_location = c1.selectbox("Ubicación", UBICACIONES_DEMO, index=0, key="bulk_location")
            bulk_shipping = c2.selectbox(
                "Entrega", ["Envío y retiro", "Solo envío", "Solo retiro"], key="bulk_shipping"
            )
            bulk_negotiable = c3.checkbox("Precios conversables", value=False, key="bulk_negotiable")
            description_template = st.text_area(
                "Plantilla de descripción",
                value=(
                    "{carta} · {set} #{numero}. Estado {estado}. "
                    "Stock publicado: {cantidad}. Se entrega protegida y embalada con cuidado."
                ),
                help="Puedes usar {carta}, {set}, {numero}, {estado} y {cantidad}.",
                key="bulk_description_template",
            )

            if selected.empty:
                st.info("Selecciona al menos una carta para crear el lote.")
            elif len(selected_valid) < len(selected):
                st.warning(
                    f"{len(selected) - len(selected_valid)} selección(es) serán omitidas porque requieren revisión o no tienen precio."
                )

            confirm = st.checkbox(
                f"Confirmo que revisé las {len(selected_valid)} publicaciones aptas del lote.",
                value=False,
                key="bulk_confirm_checkbox",
            )
            if st.button(
                f"Publicar {len(selected_valid)} productos en el marketplace",
                type="primary",
                use_container_width=True,
                disabled=(not confirm or selected_valid.empty),
                key="bulk_publish_confirm",
            ):
                created, updated, skipped = _publish_inventory_batch(
                    df_result=df_result,
                    edited=edited_state,
                    location=bulk_location,
                    shipping=bulk_shipping,
                    negotiable=bulk_negotiable,
                    description_template=description_template,
                )
                st.session_state.pop("bulk_publish_signature", None)
                st.session_state["bulk_last_result"] = {
                    "created": created, "updated": updated, "skipped": skipped,
                }
                st.rerun()

            last_result = st.session_state.pop("bulk_last_result", None)
            if last_result:
                st.success(
                    f"✅ Lote procesado: {last_result['created']} publicaciones creadas y "
                    f"{last_result['updated']} actualizadas sin duplicarlas."
                )
                if last_result["skipped"]:
                    with st.expander(f"Ver {len(last_result['skipped'])} elementos omitidos"):
                        for reason in last_result["skipped"][:50]:
                            st.write(f"• {reason}")
                if st.button("Abrir mis publicaciones", use_container_width=True, key="bulk_open_market"):
                    _go_to("Vender")

        with st.expander("Publicar solamente una carta", expanded=False):
            publicables = df_result.copy()
            if "Revisar" in publicables.columns:
                publicables = publicables[publicables["Revisar"] != "Sí"]
            if publicables.empty:
                st.warning("Primero corrige las cartas marcadas para revisión.")
            else:
                opciones_pub = [
                    f"{idx} · {row.get('Nombre Original', '-')} · {row.get('Set', '-')} #{row.get('Número Carta', '-')}"
                    for idx, row in publicables.iterrows()
                ]
                seleccion = st.selectbox("Carta a publicar", opciones_pub, key="publish_inventory_select")
                idx_real = int(seleccion.split(" · ", 1)[0])
                fila = df_result.loc[idx_real].to_dict()
                p1, p2, p3 = st.columns(3)
                precio_base = int(fila.get("Precio CLP Sugerido") or 0)
                precio_publicacion = p1.number_input(
                    "Precio", min_value=0, value=precio_base, step=500, key="publish_inventory_price"
                )
                cantidad_pub = p2.number_input(
                    "Cantidad", min_value=1, max_value=max(1, int(fila.get("Cantidad", 1))),
                    value=1, key="publish_inventory_qty",
                )
                estado_pub = p3.selectbox(
                    "Estado", list(ESTADOS.keys()),
                    index=list(ESTADOS.keys()).index(fila.get("Estado", "NM")),
                    key="publish_inventory_condition",
                )
                p4, p5, p6 = st.columns(3)
                ubicacion = p4.selectbox("Ubicación", UBICACIONES_DEMO, index=0, key="publish_inventory_location")
                envio = p5.selectbox(
                    "Entrega", ["Envío y retiro", "Solo envío", "Solo retiro"], key="publish_inventory_shipping"
                )
                negociable = p6.checkbox("Precio conversable", value=False, key="publish_inventory_negotiable")
                descripcion = st.text_area(
                    "Descripción",
                    value=f"Carta {fila.get('Nombre Original', '')} en estado {estado_pub}. Se entrega protegida.",
                    key="publish_inventory_description",
                )
                if st.button("Publicar una carta", use_container_width=True, key="publish_inventory_btn"):
                    listing = _listing_from_inventory(
                        fila=fila, price=int(precio_publicacion), quantity=int(cantidad_pub),
                        condition=estado_pub, location=ubicacion, shipping=envio,
                        negotiable=negociable, description=descripcion,
                    )
                    listing["inventory_key"] = _inventory_publication_key(fila)
                    st.session_state["marketplace_db"].insert(0, listing)
                    df_result.at[idx_real, "Publicado"] = "Sí"
                    df_result.at[idx_real, "ID Publicación"] = listing["id"]
                    st.session_state["df_result"] = df_result
                    st.session_state.pop("bulk_publish_signature", None)
                    st.success("✅ La carta ya aparece en el marketplace.")


def _agregar_resultado_al_dashboard(res: dict):
    cand = res.pop("_candidatos", []) if isinstance(res, dict) else []
    df_new = pd.DataFrame([res])
    if "df_result" in st.session_state and not st.session_state["df_result"].empty:
        df_comb = pd.concat([st.session_state["df_result"], df_new], ignore_index=True)
    else:
        df_comb = df_new
    st.session_state["df_result"] = df_comb
    idx = len(df_comb) - 1
    st.session_state.setdefault("candidatos", {})[idx] = cand


# ══════════════════════════════════════════════════════════════════════════════
# EXPERIENCIA MARKETPLACE PARA EL PILOTO
# ══════════════════════════════════════════════════════════════════════════════
EXTENDED_FEEDBACK_FILE = os.getenv("NEXOGEEK_EXTENDED_FEEDBACK_FILE", "feedback_nexogeek_extendido.csv")
INTERACTIONS_FILE = os.getenv("NEXOGEEK_INTERACTIONS_FILE", "interacciones_nexogeek.csv")
UBICACIONES_DEMO = ["Santiago", "Ñuñoa", "Providencia", "Maipú", "Puente Alto", "Viña del Mar", "Concepción", "Envío nacional"]

EXTRA_CSS = """
<style>
.block-container { padding-top: 2.6rem; padding-bottom: 3rem; max-width: 1450px; }
header[data-testid="stHeader"] { background:rgba(250,250,250,.96); border-bottom:1px solid #E2E8F0; }
button[kind="primary"] { background:#2563EB !important; border-color:#2563EB !important; color:#FFFFFF !important; }
button[kind="primary"] p { color:#FFFFFF !important; }
div[data-testid="stButton"] > button { min-height:2.55rem; white-space:nowrap; }
@media (max-width: 1050px) {
    .block-container { padding-top: 3rem; }
    div[data-testid="stButton"] > button { font-size:.78rem; padding-left:.35rem; padding-right:.35rem; }
}
[data-testid="stMetric"] { background:#FFFFFF; border:1px solid #E2E8F0; border-radius:14px; padding:12px 14px; box-shadow:0 4px 12px rgba(15,23,42,.035); }
[data-testid="stMetricLabel"] { color:#64748B; }
[data-testid="stMetricValue"] { color:#0F172A; }
.pilot-pill { display:inline-flex; align-items:center; gap:7px; background:#ECFDF5; color:#047857; border:1px solid #A7F3D0; padding:5px 11px; border-radius:999px; font-size:.74rem; font-weight:800; }
.live-dot { width:8px; height:8px; border-radius:999px; background:#10B981; display:inline-block; box-shadow:0 0 0 4px rgba(16,185,129,.12); }
.section-kicker { color:#2563EB; font-size:.75rem; font-weight:800; text-transform:uppercase; letter-spacing:.09em; margin-bottom:6px; }
.section-title { color:#0F172A; font-weight:900; font-size:1.65rem; letter-spacing:-.03em; margin:0 0 6px; }
.section-copy { color:#64748B; font-size:.96rem; margin:0 0 18px; }
.soft-card { background:#FFFFFF; border:1px solid #E2E8F0; border-radius:18px; padding:20px; box-shadow:0 6px 18px rgba(15,23,42,.045); height:100%; }
.soft-card h4 { color:#0F172A; margin:0 0 8px; font-size:1.05rem; }
.soft-card p { color:#64748B; margin:0; font-size:.88rem; line-height:1.55; }
.role-card { background:linear-gradient(160deg,#FFFFFF 0%,#F8FAFC 100%); border:1px solid #E2E8F0; border-radius:20px; padding:22px; min-height:185px; }
.role-icon { font-size:2rem; margin-bottom:10px; }
.role-card h3 { color:#0F172A; margin:0 0 8px; font-weight:850; }
.role-card p { color:#64748B; font-size:.9rem; min-height:46px; }
.badge-demo { display:inline-block; background:#EEF2FF; color:#4338CA; border:1px solid #C7D2FE; border-radius:999px; padding:4px 9px; font-size:.68rem; font-weight:800; }
.badge-verified { display:inline-block; background:#ECFDF5; color:#047857; border:1px solid #A7F3D0; border-radius:999px; padding:3px 8px; font-size:.67rem; font-weight:800; }
.badge-stock { display:inline-block; background:#FFF7ED; color:#C2410C; border:1px solid #FED7AA; border-radius:999px; padding:3px 8px; font-size:.67rem; font-weight:800; }
.listing-title { color:#0F172A; font-weight:850; font-size:1rem; line-height:1.25; min-height:2.5em; margin:8px 0 5px; }
.listing-meta { color:#64748B; font-size:.76rem; min-height:2.2em; }
.listing-price { color:#1E3A8A; font-size:1.3rem; font-weight:900; margin:8px 0 2px; }
.seller-line { color:#475569; font-size:.78rem; margin-bottom:10px; }
.empty-visual { height:210px; background:linear-gradient(135deg,#EEF2FF,#F8FAFC); border:1px dashed #CBD5E1; border-radius:14px; display:flex; align-items:center; justify-content:center; text-align:center; padding:16px; }
.empty-visual-icon { font-size:3rem; line-height:1; margin-bottom:9px; }
.empty-visual-label { color:#475569; font-size:.74rem; font-weight:800; letter-spacing:.03em; text-transform:uppercase; }
.trust-strip { background:#0F172A; color:#E2E8F0; border-radius:20px; padding:22px; margin:20px 0; }
.trust-strip strong { color:#FFFFFF; }
.step-number { width:34px; height:34px; display:flex; align-items:center; justify-content:center; background:#2563EB; color:white; border-radius:10px; font-weight:900; margin-bottom:10px; }
.compact-note { background:#F8FAFC; border-left:4px solid #2563EB; border-radius:10px; padding:11px 13px; color:#475569; font-size:.84rem; }
.hero-actions { margin-top:16px; }
div[data-testid="stButton"] > button { border-radius:10px; font-weight:700; }
button[kind="primary"] { box-shadow:0 6px 16px rgba(37,99,235,.18); }
[data-testid="stSidebar"] [data-testid="stMetric"] { background:rgba(255,255,255,.06); border-color:rgba(255,255,255,.1); box-shadow:none; }
[data-testid="stSidebar"] [data-testid="stMetricLabel"], [data-testid="stSidebar"] [data-testid="stMetricValue"] { color:#F8FAFC !important; }
</style>
"""


# ══════════════════════════════════════════════════════════════════════════════
# IDENTIDAD VISUAL NEXOGEEK — PORTAL GEEK / COMIC MODERNO
# Paleta propia: marfil, berenjena, coral, amarillo y turquesa.
# ══════════════════════════════════════════════════════════════════════════════
ORIGINAL_IDENTITY_CSS = """
<style>
:root {
    --ng-ink: #28163A;
    --ng-plum: #5B2A86;
    --ng-violet: #8A4FFF;
    --ng-coral: #FF6B6B;
    --ng-sun: #FFC857;
    --ng-mint: #2EC4B6;
    --ng-cream: #FFF8ED;
    --ng-paper: #FFFDF8;
    --ng-soft: #F3E9FF;
    --ng-line: #D9CCE7;
    --ng-muted: #73667E;
}

/* Fondo general cálido y reconocible */
html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"], .stApp {
    background:
      radial-gradient(circle at 8% 8%, rgba(255,200,87,.20), transparent 24rem),
      radial-gradient(circle at 92% 18%, rgba(46,196,182,.12), transparent 26rem),
      linear-gradient(180deg, #FFFDF8 0%, #FFF8ED 100%) !important;
    color: var(--ng-ink) !important;
    font-family: "Trebuchet MS", "Segoe UI", sans-serif !important;
}
.block-container { max-width: 1480px !important; padding-top: 2.2rem !important; }
header[data-testid="stHeader"] {
    background: rgba(255,253,248,.92) !important;
    border-bottom: 2px solid rgba(40,22,58,.10) !important;
    backdrop-filter: blur(14px);
}

/* Sidebar: cabina del portal, no panel corporativo azul */
[data-testid="stSidebar"] {
    background:
      radial-gradient(circle at 18% 7%, rgba(255,107,107,.28), transparent 13rem),
      radial-gradient(circle at 90% 40%, rgba(138,79,255,.22), transparent 16rem),
      linear-gradient(180deg, #2A153A 0%, #3B1D4F 100%) !important;
    border-right: 3px solid var(--ng-sun) !important;
}
[data-testid="stSidebar"]::before {
    content:""; display:block; height:8px;
    background:linear-gradient(90deg,var(--ng-coral) 0 33%,var(--ng-sun) 33% 66%,var(--ng-mint) 66%);
}
.side-brand { display:flex; align-items:center; gap:12px; margin:5px 0 13px; }
.side-logo {
    width:45px; height:45px; display:flex; align-items:center; justify-content:center;
    background:var(--ng-sun); color:var(--ng-ink); border:2px solid #FFF8ED;
    border-radius:13px 13px 13px 4px; font-weight:950; letter-spacing:-.05em;
    box-shadow:4px 4px 0 rgba(255,107,107,.78); transform:rotate(-2deg);
}
.side-brand strong { color:#FFF8ED; display:block; font-size:1.25rem; letter-spacing:-.04em; }
.side-brand small { color:#DCCBEA; display:block; font-size:.68rem; line-height:1.2; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4 { color:#FFF8ED !important; }
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p { color:#D9CBE4 !important; }
[data-testid="stSidebar"] [data-testid="stExpander"] {
    background:#FFF9EF !important; border:2px solid #D8C5E8 !important;
    border-radius:16px 16px 16px 6px !important; box-shadow:4px 4px 0 rgba(16,8,24,.22);
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary,
[data-testid="stSidebar"] [data-testid="stExpander"] label,
[data-testid="stSidebar"] [data-testid="stExpander"] p,
[data-testid="stSidebar"] [data-testid="stExpander"] input,
[data-testid="stSidebar"] [data-testid="stExpander"] textarea,
[data-testid="stSidebar"] [data-testid="stExpander"] [data-baseweb="select"] div {
    color:var(--ng-ink) !important;
}
[data-testid="stSidebar"] [data-testid="stMetric"] {
    background:rgba(255,248,237,.10) !important; border:1px solid rgba(255,248,237,.24) !important;
    border-radius:14px 14px 14px 5px !important;
}

/* Marca superior */
.nexo-brand { display:flex; align-items:center; gap:11px; min-height:46px; }
.nexo-brand-mark {
    width:40px; height:40px; display:flex; align-items:center; justify-content:center;
    border:2px solid var(--ng-ink); background:var(--ng-coral); border-radius:12px 12px 4px 12px;
    box-shadow:4px 4px 0 var(--ng-sun); font-size:1.15rem; transform:rotate(-3deg);
}
.nexo-brand-name { color:var(--ng-ink); font-size:1.55rem; font-weight:950; letter-spacing:-.055em; line-height:1; }
.nexo-brand-tag { color:var(--ng-muted); font-size:.65rem; margin-top:4px; text-transform:uppercase; letter-spacing:.08em; }

/* Botones tipo sticker / arcade */
div[data-testid="stButton"] > button {
    min-height:2.65rem !important; border-radius:13px 13px 5px 13px !important;
    border:2px solid var(--ng-ink) !important; background:#FFFDF8 !important;
    color:var(--ng-ink) !important; font-weight:850 !important;
    box-shadow:3px 3px 0 rgba(40,22,58,.90) !important;
    transition:transform .13s ease, box-shadow .13s ease, background .13s ease !important;
}
div[data-testid="stButton"] > button:hover {
    background:#F3E9FF !important; transform:translate(2px,2px) !important;
    box-shadow:1px 1px 0 rgba(40,22,58,.90) !important;
}
button[kind="primary"], div[data-testid="stButton"] > button[kind="primary"] {
    background:var(--ng-coral) !important; border-color:var(--ng-ink) !important;
    color:#241231 !important; box-shadow:4px 4px 0 var(--ng-sun) !important;
}
button[kind="primary"] p { color:#241231 !important; }
button[kind="primary"]:hover { background:#FF8585 !important; }

/* Hero distintivo: portal, órbitas y fichas flotantes */
.hero-banner, .nexo-hero {
    position:relative; overflow:hidden;
    background:
      radial-gradient(circle at 74% 42%, rgba(255,200,87,.92) 0 8%, transparent 8.5%),
      radial-gradient(circle at 75% 42%, transparent 0 18%, rgba(46,196,182,.85) 18.5% 20.5%, transparent 21%),
      radial-gradient(circle at 75% 42%, transparent 0 29%, rgba(255,107,107,.88) 29.5% 32%, transparent 32.5%),
      linear-gradient(125deg,#321745 0%,#5B2A86 57%,#8A4FFF 100%) !important;
    border:3px solid var(--ng-ink) !important; border-left:3px solid var(--ng-ink) !important;
    border-radius:28px 28px 10px 28px !important; padding:3.35rem 3.2rem !important;
    box-shadow:10px 10px 0 var(--ng-sun) !important; margin:10px 6px 32px 0 !important;
}
.nexo-hero::after {
    content:"✦"; position:absolute; right:7%; top:13%; color:#FFF8ED; font-size:2.3rem;
    text-shadow:-75px 85px 0 var(--ng-coral), 55px 135px 0 var(--ng-mint), -10px 190px 0 var(--ng-sun);
    transform:rotate(12deg);
}
.nexo-hero-copy { position:relative; width:min(650px,68%); z-index:2; }
.nexo-eyebrow {
    display:inline-block; background:var(--ng-sun); color:var(--ng-ink); border:2px solid var(--ng-ink);
    padding:7px 13px; border-radius:10px 10px 3px 10px; font-size:.72rem; font-weight:950;
    letter-spacing:.08em; transform:rotate(-1deg); box-shadow:3px 3px 0 var(--ng-coral);
}
.hero-title, .nexo-hero h1 {
    color:#FFF8ED !important; font-family:"Arial Black","Trebuchet MS",sans-serif !important;
    font-size:clamp(2rem,4vw,3.65rem) !important; line-height:1.01 !important;
    letter-spacing:-.06em !important; max-width:700px; margin:21px 0 15px !important;
}
.nexo-hero h1 em { color:var(--ng-sun); font-style:normal; text-decoration:underline; text-decoration-color:var(--ng-coral); text-underline-offset:7px; }
.hero-subtitle, .nexo-hero p { color:#F1E7F7 !important; max-width:640px; font-size:1.02rem !important; line-height:1.62; }
.nexo-chip-row { display:flex; flex-wrap:wrap; gap:8px; margin-top:20px; }
.nexo-chip-row span {
    background:rgba(255,253,248,.10); color:#FFF8ED; border:1px solid rgba(255,253,248,.38);
    padding:6px 10px; border-radius:999px; font-size:.72rem; font-weight:800;
}

/* Cinta de universos */
.universe-ribbon { display:flex; flex-wrap:wrap; gap:10px; margin:-4px 0 24px; }
.universe-ribbon span {
    border:2px solid var(--ng-ink); background:#FFFDF8; color:var(--ng-ink);
    border-radius:999px; padding:8px 13px; font-weight:850; font-size:.78rem;
    box-shadow:3px 3px 0 var(--ng-mint);
}
.universe-ribbon span:nth-child(2n) { box-shadow:3px 3px 0 var(--ng-coral); }
.universe-ribbon span:nth-child(3n) { box-shadow:3px 3px 0 var(--ng-sun); }

/* Encabezados editoriales */
.section-kicker {
    display:inline-block; color:var(--ng-ink) !important; background:var(--ng-mint);
    border:2px solid var(--ng-ink); border-radius:9px 9px 3px 9px; padding:5px 9px;
    font-size:.68rem !important; letter-spacing:.10em !important; margin-bottom:10px !important;
    box-shadow:3px 3px 0 rgba(40,22,58,.85);
}
.section-title { color:var(--ng-ink) !important; font-family:"Arial Black","Trebuchet MS",sans-serif; font-size:1.85rem !important; letter-spacing:-.055em !important; }
.section-copy { color:var(--ng-muted) !important; max-width:850px; }

/* Tarjetas, filtros y bloques con lenguaje gráfico */
div[data-testid="stVerticalBlockBorderWrapper"] {
    background:rgba(255,253,248,.94) !important; border:2px solid var(--ng-ink) !important;
    border-radius:20px 20px 8px 20px !important; box-shadow:5px 5px 0 rgba(40,22,58,.16) !important;
}
.soft-card, .role-card, .product-card {
    background:#FFFDF8 !important; border:2px solid var(--ng-ink) !important;
    border-radius:22px 22px 7px 22px !important; box-shadow:6px 6px 0 rgba(91,42,134,.18) !important;
}
.role-card { background:linear-gradient(145deg,#FFFDF8 0%,#F6EDFF 100%) !important; }
.role-icon {
    width:50px; height:50px; display:flex; align-items:center; justify-content:center;
    background:var(--ng-sun); border:2px solid var(--ng-ink); border-radius:15px 15px 5px 15px;
    box-shadow:4px 4px 0 var(--ng-coral); transform:rotate(-3deg);
}
.soft-card:nth-child(2n), .role-card:nth-child(2n) { box-shadow:6px 6px 0 rgba(46,196,182,.28) !important; }
.step-number {
    background:var(--ng-violet) !important; border:2px solid var(--ng-ink); color:#FFF8ED !important;
    border-radius:12px 12px 4px 12px !important; box-shadow:3px 3px 0 var(--ng-sun);
}

/* Tarjetas de productos */
.listing-title { color:var(--ng-ink) !important; font-family:"Trebuchet MS",sans-serif; font-size:1.04rem !important; }
.listing-meta, .seller-line { color:var(--ng-muted) !important; }
.listing-price { color:var(--ng-plum) !important; font-family:"Arial Black","Trebuchet MS",sans-serif; letter-spacing:-.04em; }
.badge-demo {
    background:#F2E7FF !important; color:#512176 !important; border:1.5px solid #A875D2 !important;
    border-radius:8px 8px 3px 8px !important;
}
.badge-verified { background:#DDF8F3 !important; color:#116B61 !important; border:1.5px solid #61CFC3 !important; border-radius:8px 8px 3px 8px !important; }
.badge-stock { background:#FFF0CC !important; color:#7C4B00 !important; border:1.5px solid #F4C15A !important; border-radius:8px 8px 3px 8px !important; }
.empty-visual {
    background:
      linear-gradient(135deg,rgba(138,79,255,.12),rgba(46,196,182,.13)),
      repeating-linear-gradient(45deg,transparent 0 12px,rgba(40,22,58,.035) 12px 13px) !important;
    border:2px dashed var(--ng-plum) !important; border-radius:18px 18px 6px 18px !important;
}
.empty-visual-label { color:var(--ng-plum) !important; }

/* Métricas como fichas de tablero */
[data-testid="stMetric"] {
    background:#FFFDF8 !important; border:2px solid var(--ng-ink) !important;
    border-radius:16px 16px 5px 16px !important; box-shadow:4px 4px 0 var(--ng-sun) !important;
}
[data-testid="stMetricLabel"] { color:var(--ng-muted) !important; }
[data-testid="stMetricValue"] { color:var(--ng-ink) !important; font-family:"Arial Black","Trebuchet MS",sans-serif; }

/* Formularios */
input, textarea, [data-baseweb="select"] > div, [data-baseweb="input"] > div {
    background:#FFFDF8 !important; color:var(--ng-ink) !important;
    border-color:#CDBDDA !important; border-radius:11px !important;
}
[data-baseweb="slider"] div[role="slider"] { background:var(--ng-coral) !important; border-color:var(--ng-ink) !important; }
[data-baseweb="slider"] div[data-testid="stTickBar"] { color:var(--ng-muted) !important; }
.stTabs [data-baseweb="tab-list"] { gap:8px; border-bottom:2px solid #D9CCE7; }
button[data-baseweb="tab"] {
    color:var(--ng-muted) !important; border-radius:10px 10px 0 0 !important; padding:11px 15px !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color:var(--ng-ink) !important; background:#F3E9FF !important;
    border-bottom:4px solid var(--ng-coral) !important;
}

/* Bandas informativas */
.trust-strip {
    background:linear-gradient(110deg,#2A153A,#5B2A86) !important; color:#F6EDFF !important;
    border:2px solid var(--ng-ink); border-radius:22px 22px 7px 22px !important;
    box-shadow:7px 7px 0 var(--ng-mint); padding:25px !important;
}
.trust-strip strong { color:var(--ng-sun) !important; }
.compact-note {
    background:#FFF0D0 !important; border:2px solid var(--ng-ink) !important; border-left:2px solid var(--ng-ink) !important;
    border-radius:13px 13px 4px 13px !important; color:var(--ng-ink) !important; box-shadow:3px 3px 0 var(--ng-coral);
}
.pilot-pill {
    background:#DDF8F3 !important; color:#185D56 !important; border:2px solid #76D8CC !important;
    border-radius:10px 10px 3px 10px !important; box-shadow:3px 3px 0 rgba(255,200,87,.88);
}
.live-dot { background:var(--ng-mint) !important; }

/* Alertas */
[data-testid="stAlert"] { border:2px solid var(--ng-ink) !important; border-radius:14px 14px 5px 14px !important; }

@media (max-width: 900px) {
    .nexo-hero { padding:2.4rem 1.5rem !important; }
    .nexo-hero-copy { width:100%; }
    .nexo-hero::after { opacity:.35; }
    .nexo-brand-tag { display:none; }
}
</style>
"""


def _demo_marketplace_seed() -> list[dict]:
    return [
        {
            "id": "p-001", "title": "Charizard ex · Special Illustration Rare #234", "game": "Pokémon TCG",
            "product_type": "Carta individual", "condition": "NM", "price": 129990, "location": "Ñuñoa",
            "shipping": "Envío y retiro", "seller": "Gabo_Cards", "verified": True, "rating": 4.9,
            "sales": 86, "image": "https://images.pokemontcg.io/sv3/234.png", "stock": 1,
            "negotiable": False, "description": "Carta en excelente estado, almacenada en sleeve y top loader desde apertura.",
            "tags": ["SAR", "Destacado"], "views": 184, "likes": 27, "active": True, "owner": False,
        },
        {
            "id": "p-002", "title": "Pikachu ex · Full Art #219", "game": "Pokémon TCG",
            "product_type": "Carta individual", "condition": "NM", "price": 45990, "location": "Providencia",
            "shipping": "Envío y retiro", "seller": "PokeNorte", "verified": True, "rating": 4.8,
            "sales": 143, "image": "https://images.pokemontcg.io/sv8/219.png", "stock": 2,
            "negotiable": True, "description": "Pikachu ex full art. Fotos adicionales disponibles por chat.",
            "tags": ["Popular", "Conversable"], "views": 121, "likes": 18, "active": True, "owner": False,
        },
        {
            "id": "p-003", "title": "Mew ex · Special Illustration Rare #232", "game": "Pokémon TCG",
            "product_type": "Carta individual", "condition": "LP", "price": 31990, "location": "Santiago",
            "shipping": "Solo envío", "seller": "CardLab_CL", "verified": True, "rating": 4.7,
            "sales": 64, "image": "https://images.pokemontcg.io/sv4pt5/232.png", "stock": 1,
            "negotiable": True, "description": "Leve detalle superficial visible con luz directa. Precio ajustado al estado.",
            "tags": ["Oferta", "LP"], "views": 79, "likes": 11, "active": True, "owner": False,
        },
        {
            "id": "g-001", "title": "Booster Box sellada · expansión japonesa", "game": "Pokémon TCG",
            "product_type": "Producto sellado", "condition": "Sellado", "price": 74990, "location": "Maipú",
            "shipping": "Envío y retiro", "seller": "NekoImports", "verified": True, "rating": 4.9,
            "sales": 211, "image": "", "stock": 5, "negotiable": False,
            "description": "Caja original sellada. Entrega coordinada o envío a regiones.",
            "tags": ["Sellado", "Importado"], "views": 240, "likes": 31, "active": True, "owner": False,
        },
        {
            "id": "o-001", "title": "Monkey D. Luffy · Alternate Art", "game": "One Piece Card Game",
            "product_type": "Carta individual", "condition": "NM", "price": 89990, "location": "Santiago",
            "shipping": "Envío y retiro", "seller": "GrandLineTCG", "verified": True, "rating": 4.8,
            "sales": 51, "image": "", "stock": 1, "negotiable": True,
            "description": "Carta alternativa de colección. Protección rígida incluida.",
            "tags": ["Alt Art", "Conversable"], "views": 93, "likes": 15, "active": True, "owner": False,
        },
        {
            "id": "m-001", "title": "Commander Deck listo para jugar", "game": "Magic: The Gathering",
            "product_type": "Mazo armado", "condition": "Usado", "price": 54990, "location": "Viña del Mar",
            "shipping": "Solo envío", "seller": "ManaVault", "verified": False, "rating": 4.6,
            "sales": 18, "image": "", "stock": 1, "negotiable": True,
            "description": "Mazo Commander funcional, lista completa disponible. Incluye deck box.",
            "tags": ["Commander", "Listo para jugar"], "views": 62, "likes": 8, "active": True, "owner": False,
        },
        {
            "id": "a-001", "title": "Deck box premium para 100+ cartas", "game": "Accesorios",
            "product_type": "Accesorio", "condition": "Nuevo", "price": 24990, "location": "Providencia",
            "shipping": "Envío y retiro", "seller": "CodaGeek", "verified": True, "rating": 5.0,
            "sales": 34, "image": "", "stock": 8, "negotiable": False,
            "description": "Deck box rígido con cierre magnético y espacio para dados.",
            "tags": ["Nuevo", "Premium"], "views": 105, "likes": 20, "active": True, "owner": False,
        },
        {
            "id": "f-001", "title": "Figura coleccionable edición limitada", "game": "Figuras y animé",
            "product_type": "Figura", "condition": "Nuevo", "price": 39990, "location": "Concepción",
            "shipping": "Solo envío", "seller": "OtakuShelf", "verified": False, "rating": 4.5,
            "sales": 12, "image": "", "stock": 3, "negotiable": False,
            "description": "Figura nueva en caja. Embalaje reforzado para despacho.",
            "tags": ["Edición limitada"], "views": 48, "likes": 6, "active": True, "owner": False,
        },
        {
            "id": "manga-001", "title": "Colección manga · tomos 1 al 10", "game": "Mangas y cómics",
            "product_type": "Manga", "condition": "Usado", "price": 59990, "location": "Puente Alto",
            "shipping": "Envío y retiro", "seller": "PanelNueve", "verified": True, "rating": 4.7,
            "sales": 29, "image": "", "stock": 1, "negotiable": True,
            "description": "Colección completa del arco inicial, sin páginas sueltas ni rayados.",
            "tags": ["Colección", "Pack"], "views": 55, "likes": 9, "active": True, "owner": False,
        },
    ]


def _demo_auctions_seed() -> list[dict]:
    return [
        {
            "id": "a-101", "name": "Charizard ex · Special Illustration Rare #234", "game": "Pokémon TCG",
            "current_bid": 120000, "bids": 14, "ends": "01h 45m", "image": "https://images.pokemontcg.io/sv3/234.png",
            "seller": "Gabo_Cards", "verified": True, "increment": 2000, "watchers": 22,
            "history": [118000, 120000],
        },
        {
            "id": "a-102", "name": "Pikachu ex · Full Art", "game": "Pokémon TCG",
            "current_bid": 45000, "bids": 6, "ends": "04h 12m", "image": "https://images.pokemontcg.io/sv8/219.png",
            "seller": "UTEM_Collector", "verified": True, "increment": 1000, "watchers": 11,
            "history": [43000, 45000],
        },
        {
            "id": "a-103", "name": "Lote sorpresa One Piece · 25 cartas", "game": "One Piece Card Game",
            "current_bid": 26000, "bids": 9, "ends": "11h 05m", "image": "",
            "seller": "GrandLineTCG", "verified": True, "increment": 1000, "watchers": 8,
            "history": [25000, 26000],
        },
    ]


def _demo_services_seed() -> list[dict]:
    return [
        {
            "id": "s-001", "type": "Pregrading", "title": "Revisión pregrading PSA / CGC",
            "description": "Evaluación visual, centrado, superficie y bordes con informe fotográfico.",
            "price": 5000, "provider": "Marco_Nuñez_TCG", "rating": 4.9, "location": "Santiago",
            "badge": "🏆 Recomendado", "delivery": "24-48 horas",
        },
        {
            "id": "s-002", "type": "Encargos", "title": "Importación desde Pokémon Center Japón",
            "description": "Gestión de encargos, consolidación y entrega local con seguimiento.",
            "price": 12000, "provider": "NekoImports", "rating": 4.8, "location": "Envío nacional",
            "badge": "✈️ Próximo viaje", "delivery": "3-5 semanas",
        },
        {
            "id": "s-003", "type": "Diseño e impresión", "title": "Deck boxes y exhibidores impresos en 3D",
            "description": "Diseños personalizados para cartas, mazos y vitrinas pequeñas.",
            "price": 18990, "provider": "CodaLab3D", "rating": 5.0, "location": "Providencia",
            "badge": "🧩 Personalizable", "delivery": "4-7 días",
        },
        {
            "id": "s-004", "type": "Organización", "title": "Organización de torneos y ligas locales",
            "description": "Apoyo en brackets, difusión, inscripción y material para comunidades.",
            "price": 45000, "provider": "LigaCentral", "rating": 4.7, "location": "Santiago",
            "badge": "🎮 Comunidad", "delivery": "A coordinar",
        },
    ]


def _init_demo_state() -> None:
    defaults = {
        "page": "Inicio",
        "marketplace_db": _demo_marketplace_seed(),
        "favorites": [],
        "cart": [],
        "compare": [],
        "notifications": [],
        "selected_listing": None,
        "auction_watchlist": [],
        "subastas_db": _demo_auctions_seed(),
        "servicios_db": _demo_services_seed(),
        "feature_votes": {},
        "pilot_alias": "Usuario_Piloto",
        "pilot_location": "Santiago",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _new_id(prefix: str) -> str:
    return f"{prefix}-{time.time_ns()}"


def _notify(message: str, kind: str = "info") -> None:
    st.session_state.setdefault("notifications", []).insert(0, {
        "message": str(message), "kind": kind, "time": datetime.now().strftime("%H:%M"),
    })
    st.session_state["notifications"] = st.session_state["notifications"][:20]


def _save_interaction(feature: str, vote: str, comment: str = "") -> None:
    row = pd.DataFrame([{
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "feature": feature,
        "vote": vote,
        "comment": comment.strip(),
        "alias": st.session_state.get("pilot_alias", ""),
    }])
    with _FEEDBACK_LOCK:
        exists = os.path.exists(INTERACTIONS_FILE)
        row.to_csv(INTERACTIONS_FILE, mode="a", header=not exists, index=False, encoding="utf-8-sig")


def _save_extended_feedback(data: dict) -> None:
    row = pd.DataFrame([{**{"fecha": datetime.now().isoformat(timespec="seconds")}, **data}])
    with _FEEDBACK_LOCK:
        exists = os.path.exists(EXTENDED_FEEDBACK_FILE)
        row.to_csv(EXTENDED_FEEDBACK_FILE, mode="a", header=not exists, index=False, encoding="utf-8-sig")


def _read_csv_safe(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame()


def _listing_by_id(listing_id: str) -> dict | None:
    return next((x for x in st.session_state.get("marketplace_db", []) if x.get("id") == listing_id), None)


def _toggle_item(state_key: str, item_id: str, max_items: int | None = None) -> bool:
    items = list(st.session_state.get(state_key, []))
    added = item_id not in items
    if added:
        if max_items and len(items) >= max_items:
            return False
        items.append(item_id)
    else:
        items.remove(item_id)
    st.session_state[state_key] = items
    return added


def _listing_from_inventory(fila: dict, price: int, quantity: int, condition: str,
                            location: str, shipping: str, negotiable: bool,
                            description: str) -> dict:
    return {
        "id": _new_id("user"),
        "title": f"{fila.get('Nombre Original', fila.get('Nombre EN', 'Carta'))} · {fila.get('Set', '-')} #{fila.get('Número Carta', '-')}",
        "game": "Pokémon TCG",
        "product_type": "Carta individual",
        "condition": condition,
        "price": int(price),
        "location": location,
        "shipping": shipping,
        "seller": st.session_state.get("pilot_alias", "Usuario_Piloto"),
        "verified": False,
        "rating": 5.0,
        "sales": 0,
        "image": fila.get("URL Imagen", ""),
        "stock": int(quantity),
        "negotiable": bool(negotiable),
        "description": description,
        "tags": ["Recién publicado", condition],
        "views": 0,
        "likes": 0,
        "active": True,
        "owner": True,
    }


def _placeholder_for_listing(listing: dict) -> tuple[str, str]:
    """Devuelve un visual coherente cuando la publicación demo no tiene fotografía."""
    product_type = str(listing.get("product_type", "")).lower()
    game = str(listing.get("game", "")).lower()
    if "sellado" in product_type:
        return "📦", "Producto sellado"
    if "manga" in product_type or "manga" in game or "cómic" in game or "comic" in game:
        return "📚", "Manga o cómic"
    if "figura" in product_type or "figura" in game or "animé" in game or "anime" in game:
        return "🗿", "Figura coleccionable"
    if "accesorio" in product_type or "accesorio" in game:
        return "🧰", "Accesorio TCG"
    if "mazo" in product_type:
        return "🗂️", "Mazo armado"
    if "one piece" in game:
        return "🏴‍☠️", "Carta One Piece"
    if "magic" in game:
        return "🧙", "Carta Magic"
    return "🃏", "Carta coleccionable"


def _render_image_or_placeholder(image: str, emoji: str = "🃏", width: int = 220,
                                 label: str = "Imagen de demostración") -> None:
    if _safe_image_url(image):
        st.image(image, width=width)
    else:
        st.markdown(
            f'<div class="empty-visual" style="height:{max(150, int(width*1.05))}px;">'
            f'<div><div class="empty-visual-icon">{emoji}</div>'
            f'<div class="empty-visual-label">{_safe_text(label)}</div></div></div>',
            unsafe_allow_html=True,
        )


def _render_listing_card(listing: dict, prefix: str) -> None:
    lid = listing["id"]
    favorites = st.session_state.get("favorites", [])
    compare = st.session_state.get("compare", [])
    with st.container(border=True):
        placeholder_emoji, placeholder_label = _placeholder_for_listing(listing)
        _render_image_or_placeholder(
            listing.get("image", ""), placeholder_emoji, width=210, label=placeholder_label
        )
        tags = listing.get("tags", [])[:2]
        badge_html = " ".join(f'<span class="badge-demo">{_safe_text(tag)}</span>' for tag in tags)
        st.markdown(badge_html, unsafe_allow_html=True)
        st.markdown(f'<div class="listing-title">{_safe_text(listing.get("title", ""))}</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="listing-meta">{_safe_text(listing.get("game", ""))} · {_safe_text(listing.get("condition", ""))}<br>'
            f'📍 {_safe_text(listing.get("location", ""))} · {_safe_text(listing.get("shipping", ""))}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(f'<div class="listing-price">{_fmt_clp(listing.get("price"))}</div>', unsafe_allow_html=True)
        verified = ' <span class="badge-verified">✓ verificado</span>' if listing.get("verified") else ""
        st.markdown(
            f'<div class="seller-line">👤 {_safe_text(listing.get("seller", ""))}{verified} · ⭐ {listing.get("rating", 0):.1f}</div>',
            unsafe_allow_html=True,
        )
        b1, b2, b3 = st.columns(3)
        fav_label = "♥" if lid in favorites else "♡"
        if b1.button(fav_label, key=f"{prefix}_fav_{lid}", use_container_width=True, help="Guardar en favoritos"):
            added = _toggle_item("favorites", lid)
            listing["likes"] = max(0, int(listing.get("likes", 0)) + (1 if added else -1))
            _notify(("Guardaste" if added else "Quitaste") + f" {listing['title']} de favoritos", "success")
            st.rerun()
        if b2.button("Ver", key=f"{prefix}_view_{lid}", use_container_width=True):
            listing["views"] = int(listing.get("views", 0)) + 1
            st.session_state["selected_listing"] = lid
            st.rerun()
        if b3.button("＋", key=f"{prefix}_cart_{lid}", use_container_width=True, help="Agregar a reserva"):
            if lid not in st.session_state.get("cart", []):
                st.session_state["cart"].append(lid)
                _notify(f"Agregaste {listing['title']} a tu reserva", "success")
            else:
                _notify("Este producto ya estaba en tu reserva", "info")
            st.rerun()
        c1, c2 = st.columns([1, 1])
        compare_label = "✓ Comparando" if lid in compare else "Comparar"
        if c1.button(compare_label, key=f"{prefix}_cmp_{lid}", use_container_width=True):
            if lid in compare:
                _toggle_item("compare", lid)
            elif len(compare) >= 3:
                _notify("Puedes comparar hasta 3 publicaciones", "warning")
            else:
                _toggle_item("compare", lid, max_items=3)
            st.rerun()
        c2.caption(f"👁 {listing.get('views', 0)} · ♥ {listing.get('likes', 0)}")


def _render_listing_detail(listing: dict, prefix: str = "detail") -> None:
    st.markdown("---")
    d1, d2 = st.columns([1, 1.5])
    with d1:
        placeholder_emoji, placeholder_label = _placeholder_for_listing(listing)
        _render_image_or_placeholder(
            listing.get("image", ""), placeholder_emoji, width=320, label=placeholder_label
        )
    with d2:
        st.markdown(f"## {_safe_text(listing.get('title', ''))}", unsafe_allow_html=True)
        tag_html = " ".join(f'<span class="badge-demo">{_safe_text(t)}</span>' for t in listing.get("tags", []))
        st.markdown(tag_html, unsafe_allow_html=True)
        st.markdown(f"### {_fmt_clp(listing.get('price'))}")
        if listing.get("negotiable"):
            st.caption("💬 Precio conversable")
        info1, info2, info3 = st.columns(3)
        info1.metric("Estado", listing.get("condition", "-"))
        info2.metric("Stock", listing.get("stock", 0))
        info3.metric("Visitas", listing.get("views", 0))
        st.write(listing.get("description", ""))
        st.markdown(
            f"**Entrega:** {listing.get('shipping', '-')}  ·  **Ubicación:** {listing.get('location', '-')}  ·  "
            f"**Categoría:** {listing.get('product_type', '-')}")
        st.markdown("#### Vendedor")
        verified = "✅ Perfil verificado" if listing.get("verified") else "🟡 Perfil nuevo"
        st.write(f"**{listing.get('seller')}** · ⭐ {listing.get('rating', 0):.1f} · {listing.get('sales', 0)} ventas · {verified}")
        a1, a2, a3 = st.columns(3)
        if a1.button("💬 Contactar", key=f"{prefix}_contact_{listing['id']}", type="primary", use_container_width=True):
            _notify(f"Chat simulado abierto con {listing['seller']}", "success")
            st.session_state["open_chat_seller"] = listing["seller"]
            st.success("Chat simulado abierto. En producción aquí existiría mensajería y protección de datos.")
        if a2.button("🛒 Reservar", key=f"{prefix}_reserve_{listing['id']}", use_container_width=True):
            if listing["id"] not in st.session_state["cart"]:
                st.session_state["cart"].append(listing["id"])
            _notify(f"Reserva simulada creada para {listing['title']}", "success")
            st.success("Reserva agregada. No se realizó ningún cobro.")
        if a3.button("⚑ Reportar", key=f"{prefix}_report_{listing['id']}", use_container_width=True):
            _notify(f"Reporte simulado enviado sobre {listing['title']}", "warning")
            st.info("Gracias. En una versión real esto llegaría a moderación.")


def _render_feature_vote(feature: str, question: str = "¿Te resultaría útil esta sección?") -> None:
    with st.expander("🧪 Evaluación rápida de esta función", expanded=False):
        st.write(question)
        v1, v2, v3 = st.columns(3)
        vote = None
        if v1.button("👍 Sí", key=f"vote_yes_{feature}", use_container_width=True): vote = "Sí"
        if v2.button("🤔 Tal vez", key=f"vote_maybe_{feature}", use_container_width=True): vote = "Tal vez"
        if v3.button("👎 No", key=f"vote_no_{feature}", use_container_width=True): vote = "No"
        comment = st.text_input("Comentario breve (opcional)", key=f"vote_comment_{feature}")
        if vote:
            st.session_state.setdefault("feature_votes", {})[feature] = vote
            try:
                _save_interaction(feature, vote, comment)
            except Exception:
                pass
            st.success("Respuesta registrada para el piloto.")


def _section_header(kicker: str, title: str, copy: str) -> None:
    st.markdown(
        f'<div class="section-kicker">{_safe_text(kicker)}</div>'
        f'<div class="section-title">{_safe_text(title)}</div>'
        f'<div class="section-copy">{_safe_text(copy)}</div>',
        unsafe_allow_html=True,
    )


def _render_sidebar(api_key_default: str = "") -> tuple[str | None, float, float]:
    with st.sidebar:
        st.markdown(
            "<div class='side-brand'><div class='side-logo'>NG</div>"
            "<div><strong>NexoGeek</strong><small>Tu portal de colección y juego</small></div></div>",
            unsafe_allow_html=True,
        )
        st.markdown('<span class="pilot-pill"><span class="live-dot"></span>Laboratorio piloto</span>', unsafe_allow_html=True)
        st.caption("Explora, colecciona, juega y conecta con la comunidad.")

        with st.expander("👤 Mi perfil de prueba", expanded=True):
            alias = st.text_input("Alias", value=st.session_state.get("pilot_alias", "Usuario_Piloto"), key="profile_alias_input")
            location = st.selectbox(
                "Ubicación",
                UBICACIONES_DEMO,
                index=UBICACIONES_DEMO.index(st.session_state.get("pilot_location", "Santiago"))
                if st.session_state.get("pilot_location", "Santiago") in UBICACIONES_DEMO else 0,
                key="profile_location_input",
            )
            st.session_state["pilot_alias"] = alias.strip() or "Usuario_Piloto"
            st.session_state["pilot_location"] = location
            st.markdown("<span class='badge-demo'>Perfil demo</span>", unsafe_allow_html=True)

        f1, f2, f3 = st.columns(3)
        f1.metric("♥", len(st.session_state.get("favorites", [])))
        f2.metric("🛒", len(st.session_state.get("cart", [])))
        f3.metric("🔔", len(st.session_state.get("notifications", [])))

        with st.expander("🛒 Reserva simulada", expanded=False):
            cart_ids = list(st.session_state.get("cart", []))
            if not cart_ids:
                st.caption("Aún no agregas productos.")
            total = 0
            for lid in cart_ids:
                item = _listing_by_id(lid)
                if not item:
                    continue
                total += int(item.get("price", 0))
                c1, c2 = st.columns([4, 1])
                c1.caption(f"{item.get('title')} · {_fmt_clp(item.get('price'))}")
                if c2.button("×", key=f"sidebar_remove_{lid}"):
                    st.session_state["cart"].remove(lid)
                    st.rerun()
            if cart_ids:
                st.markdown(f"**Total referencial: {_fmt_clp(total)}**")
                if st.button("Simular checkout", use_container_width=True, type="primary", key="sidebar_checkout"):
                    _notify("Checkout simulado completado sin cobro", "success")
                    st.success("Flujo validado. No se realizó ningún pago.")

        with st.expander("🔔 Actividad reciente", expanded=False):
            notifications = st.session_state.get("notifications", [])
            if not notifications:
                st.caption("Sin actividad todavía.")
            for note in notifications[:6]:
                st.caption(f"{note['time']} · {note['message']}")

        with st.expander("⚙️ Motor de tasación", expanded=False):
            api_key_input = st.text_input(
                "PokémonTCG API Key (opcional)",
                type="password",
                value=st.session_state.get("api_key_guardada", api_key_default),
                key="api_key_sidebar",
            )
            st.session_state["api_key_guardada"] = api_key_input
            clp_rate = st.number_input("USD → CLP", min_value=0, value=950, step=10, key="clp_rate_sidebar")
            comision = st.number_input("Comisión simulada %", min_value=0.0, max_value=50.0, value=5.0, step=0.5, key="commission_sidebar")
            cfg_workers = st.slider("Hilos", min_value=1, max_value=8, value=4, key="workers_sidebar")
            cfg_throttle = st.slider("Delay anti rate-limit", min_value=0.0, max_value=2.0, value=0.2, step=0.1, key="throttle_sidebar")
            _CFG["max_workers"] = cfg_workers
            _CFG["throttle"] = cfg_throttle

            if not _DB_LOADED:
                for cp in ("card_data", os.path.join("pokemon-tcg-data", "cards", "en")):
                    if os.path.isdir(cp) and cargar_base_local(cp):
                        break
            if _DB_LOADED:
                st.success(f"DB local: {len(_DB_CARDS):,} cartas")
            else:
                st.info("Fuente actual: API cloud")

        with st.expander("📤 Datos del piloto", expanded=False):
            fb1 = leer_feedback()
            fb2 = _read_csv_safe(EXTENDED_FEEDBACK_FILE)
            interactions = _read_csv_safe(INTERACTIONS_FILE)
            if not fb1.empty:
                st.download_button("Descargar feedback básico", fb1.to_csv(index=False).encode("utf-8-sig"), "feedback_nexogeek.csv", "text/csv", use_container_width=True)
            if not fb2.empty:
                st.download_button("Descargar feedback completo", fb2.to_csv(index=False).encode("utf-8-sig"), "feedback_nexogeek_extendido.csv", "text/csv", use_container_width=True)
            if not interactions.empty:
                st.download_button("Descargar reacciones", interactions.to_csv(index=False).encode("utf-8-sig"), "interacciones_nexogeek.csv", "text/csv", use_container_width=True)
            if fb1.empty and fb2.empty and interactions.empty:
                st.caption("Todavía no hay respuestas guardadas.")

        st.markdown("---")
        if st.button("♻️ Reiniciar datos de demostración", use_container_width=True, key="reset_demo"):
            for key in ["marketplace_db", "favorites", "cart", "compare", "notifications", "selected_listing", "deck_cart", "meta_selected_card", "meta_print_choices",
                        "auction_watchlist", "subastas_db", "servicios_db", "feature_votes", "df_result", "candidatos"]:
                st.session_state.pop(key, None)
            _init_demo_state()
            st.rerun()

    return (api_key_input or None), float(clp_rate), float(comision)


def _sync_nav_from_widget() -> None:
    st.session_state["page"] = st.session_state.get("nav_widget", "Inicio")


def _go_to(page: str) -> None:
    """Navega desde un botón sin modificar directamente el widget ya dibujado."""
    st.session_state["page"] = page
    st.session_state["pending_nav"] = page
    st.rerun()


def _render_top_navigation() -> str:
    """Barra superior visible con botones, sin los círculos de ``st.radio``."""
    options = [
        ("Inicio", "🏠"),
        ("Marketplace", "🛍️"),
        ("Vender", "📦"),
        ("Tasador", "🧠"),
        ("Subastas", "🔨"),
        ("Servicios", "🤝"),
        ("Feedback", "💬"),
    ]

    pending = st.session_state.pop("pending_nav", None)
    current = pending or st.session_state.get("page", "Inicio")
    valid_pages = {page for page, _ in options}
    if current not in valid_pages:
        current = "Inicio"

    nav_cols = st.columns([2.25, 1, 1.28, 1, 1, 1.08, 1.08, 1.08], gap="small")
    with nav_cols[0]:
        st.markdown(
            "<div class='nexo-brand'>"
            "<div class='nexo-brand-mark'>✦</div>"
            "<div><div class='nexo-brand-name'>NexoGeek</div>"
            "<div class='nexo-brand-tag'>colección · juego · comunidad</div></div></div>",
            unsafe_allow_html=True,
        )

    for column, (page, icon) in zip(nav_cols[1:], options):
        with column:
            clicked = st.button(
                f"{icon} {page}",
                key=f"top_nav_{page.lower()}",
                use_container_width=True,
                type="primary" if current == page else "secondary",
            )
            if clicked and page != current:
                st.session_state["page"] = page
                st.rerun()

    st.session_state["page"] = current
    st.markdown(
        "<hr style='margin:9px 0 20px;border:none;border-top:1px solid #E2E8F0;'>",
        unsafe_allow_html=True,
    )
    return current

def render_home() -> None:
    st.markdown("""
    <div class="nexo-hero">
      <div class="nexo-hero-copy">
        <span class="nexo-eyebrow">✦ TU PORTAL GEEK, HECHO EN COMUNIDAD</span>
        <h1>Donde tu próxima <em>obsesión</em> encuentra lugar.</h1>
        <p>Descubre cartas, juegos, mangas, figuras y creaciones únicas. Compra, publica, tasa y conecta en una experiencia pensada para coleccionistas y jugadores.</p>
        <div class="nexo-chip-row">
          <span>Cartas TCG</span><span>Productos sellados</span><span>Mangas</span>
          <span>Figuras</span><span>Accesorios</span><span>Servicios creativos</span>
        </div>
      </div>
    </div>
    <div class="universe-ribbon">
      <span>⚡ Pokémon</span><span>🏴‍☠️ One Piece</span><span>🧙 Magic</span>
      <span>🎲 Juegos de mesa</span><span>📚 Manga</span><span>🗿 Figuras</span>
    </div>
    """, unsafe_allow_html=True)

    cta1, cta2, cta3 = st.columns([1, 1, 2])
    if cta1.button("🛒 Explorar productos", type="primary", use_container_width=True, key="home_buy"):
        _go_to("Marketplace")
    if cta2.button("📣 Publicar algo", use_container_width=True, key="home_sell"):
        _go_to("Vender")
    cta3.markdown('<div class="compact-note">🧪 La compra, reserva, chat y subastas son simulaciones. El tasador puede consultar datos reales.</div>', unsafe_allow_html=True)

    listings = [x for x in st.session_state["marketplace_db"] if x.get("active", True)]
    hm1, hm2, hm3, hm4 = st.columns(4)
    hm1.metric("Publicaciones activas", len(listings))
    hm2.metric("Categorías", len(set(x["product_type"] for x in listings)))
    hm3.metric("Vendedores demo", len(set(x["seller"] for x in listings)))
    hm4.metric("Favoritos guardados", len(st.session_state.get("favorites", [])))

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header("Elige tu misión", "Tres formas de entrar al Nexo", "Recorre la experiencia desde el rol que más se parece a ti.")
    r1, r2, r3 = st.columns(3)
    with r1:
        st.markdown('<div class="role-card"><div class="role-icon">🛒</div><h3>Quiero comprar</h3><p>Busca, filtra, compara, guarda favoritos y simula una reserva con un vendedor.</p></div>', unsafe_allow_html=True)
        if st.button("Comenzar como comprador", use_container_width=True, key="role_buyer"):
            _go_to("Marketplace")
    with r2:
        st.markdown('<div class="role-card"><div class="role-icon">📦</div><h3>Quiero vender</h3><p>Crea una publicación desde cero o transforma una carta tasada en un anuncio.</p></div>', unsafe_allow_html=True)
        if st.button("Comenzar como vendedor", use_container_width=True, key="role_seller"):
            _go_to("Vender")
    with r3:
        st.markdown('<div class="role-card"><div class="role-icon">⚡</div><h3>Quiero tasar</h3><p>Identifica cartas, revisa versiones, calcula precios y prepara inventario.</p></div>', unsafe_allow_html=True)
        if st.button("Abrir tasador", use_container_width=True, key="role_valuer"):
            _go_to("Tasador")

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header("Radar de hallazgos", "Piezas que están llamando la atención", "Una selección dinámica para explorar el catálogo y descubrir nuevas obsesiones.")
    featured = sorted(listings, key=lambda x: (x.get("likes", 0), x.get("views", 0)), reverse=True)[:4]
    cols = st.columns(4)
    for col, item in zip(cols, featured):
        with col:
            _render_listing_card(item, "home")

    st.markdown("""
    <div class="trust-strip">
      <strong>Una experiencia pensada para generar confianza</strong><br>
      Perfiles y reputación · Estado visible · Reportes y moderación · Entrega coordinada · Precio de referencia · Historial de actividad
    </div>
    """, unsafe_allow_html=True)

    _section_header("Cómo funcionaría", "De descubrir a cerrar una operación", "El piloto muestra la experiencia. La versión final incorporaría cuentas, pagos protegidos y persistencia real.")
    s1, s2, s3, s4 = st.columns(4)
    steps = [
        ("1", "Descubre", "Explora cartas, productos y servicios con filtros claros."),
        ("2", "Compara", "Revisa precio, estado, reputación y modalidad de entrega."),
        ("3", "Conversa", "Contacta al vendedor dentro de un chat protegido."),
        ("4", "Concreta", "Reserva, paga y coordina envío cuando exista la versión productiva."),
    ]
    for col, (n, title, copy) in zip([s1, s2, s3, s4], steps):
        with col:
            st.markdown(f'<div class="soft-card"><div class="step-number">{n}</div><h4>{title}</h4><p>{copy}</p></div>', unsafe_allow_html=True)
    _render_feature_vote("inicio", "¿La propuesta se entiende claramente desde el inicio?")


def render_marketplace() -> None:
    _section_header("Explora el nexo", "Hallazgos de la comunidad", "Filtra por universo, tipo de pieza, precio, estado y ubicación para encontrar algo que realmente te represente.")
    listings = [x for x in st.session_state["marketplace_db"] if x.get("active", True)]

    with st.container(border=True):
        f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
        search = f1.text_input("Buscar", placeholder="Charizard, booster box, manga...", key="market_search")
        games = ["Todos"] + sorted(set(x["game"] for x in listings))
        game = f2.selectbox("Juego / universo", games, key="market_game")
        types = ["Todos"] + sorted(set(x["product_type"] for x in listings))
        product_type = f3.selectbox("Tipo", types, key="market_type")
        sort_by = f4.selectbox("Ordenar", ["Más relevantes", "Precio menor", "Precio mayor", "Más vistos", "Más guardados"], key="market_sort")

        max_price = max([int(x.get("price", 0)) for x in listings] + [1000])
        f5, f6, f7, f8 = st.columns(4)
        price_range = f5.slider("Rango de precio", 0, max_price, (0, max_price), step=max(500, max_price // 100), key="market_price")
        condition_options = sorted(set(x["condition"] for x in listings))
        conditions = f6.multiselect("Estado", condition_options, key="market_conditions")
        location = f7.selectbox("Ubicación", ["Todas"] + sorted(set(x["location"] for x in listings)), key="market_location")
        delivery = f8.selectbox("Entrega", ["Todas", "Envío y retiro", "Solo envío", "Solo retiro"], key="market_delivery")
        q1, q2, q3 = st.columns(3)
        verified_only = q1.checkbox("Solo vendedores verificados", key="market_verified")
        favorites_only = q2.checkbox("Solo mis favoritos", key="market_favorites_only")
        negotiable_only = q3.checkbox("Solo precios conversables", key="market_negotiable")

    query = search.strip().lower()
    filtered = []
    for item in listings:
        haystack = " ".join([item.get("title", ""), item.get("game", ""), item.get("product_type", ""), " ".join(item.get("tags", []))]).lower()
        if query and query not in haystack: continue
        if game != "Todos" and item["game"] != game: continue
        if product_type != "Todos" and item["product_type"] != product_type: continue
        if not (price_range[0] <= int(item.get("price", 0)) <= price_range[1]): continue
        if conditions and item.get("condition") not in conditions: continue
        if location != "Todas" and item.get("location") != location: continue
        if delivery != "Todas" and item.get("shipping") != delivery: continue
        if verified_only and not item.get("verified"): continue
        if favorites_only and item["id"] not in st.session_state.get("favorites", []): continue
        if negotiable_only and not item.get("negotiable"): continue
        filtered.append(item)

    sorters = {
        "Precio menor": lambda x: x.get("price", 0),
        "Precio mayor": lambda x: -x.get("price", 0),
        "Más vistos": lambda x: -x.get("views", 0),
        "Más guardados": lambda x: -x.get("likes", 0),
        "Más relevantes": lambda x: -(x.get("likes", 0) * 3 + x.get("views", 0)),
    }
    filtered = sorted(filtered, key=sorters[sort_by])
    st.caption(f"{len(filtered)} resultado(s) · datos de demostración")

    compare_ids = st.session_state.get("compare", [])
    if compare_ids:
        with st.expander(f"⚖️ Comparador ({len(compare_ids)}/3)", expanded=True):
            comp = [x for x in listings if x["id"] in compare_ids]
            if comp:
                df_cmp = pd.DataFrame([{
                    "Producto": x["title"], "Precio": x["price"], "Estado": x["condition"],
                    "Vendedor": x["seller"], "Rating": x["rating"], "Entrega": x["shipping"],
                } for x in comp])
                st.dataframe(df_cmp, use_container_width=True, hide_index=True)
                if st.button("Limpiar comparación", key="clear_compare"):
                    st.session_state["compare"] = []; st.rerun()

    if not filtered:
        st.warning("No hay publicaciones que coincidan. Prueba ampliando los filtros.")
    else:
        for start in range(0, len(filtered), 4):
            cols = st.columns(4)
            for col, item in zip(cols, filtered[start:start+4]):
                with col:
                    _render_listing_card(item, "market")

    selected_id = st.session_state.get("selected_listing")
    selected = _listing_by_id(selected_id) if selected_id else None
    if selected:
        _render_listing_detail(selected, "market_detail")

    if st.session_state.get("favorites"):
        st.markdown("---")
        _section_header("Para ti", "También podría interesarte", "Recomendaciones simples basadas en lo que guardaste.")
        fav_games = {x["game"] for x in listings if x["id"] in st.session_state["favorites"]}
        recs = [x for x in listings if x["id"] not in st.session_state["favorites"] and x["game"] in fav_games][:4]
        if recs:
            cols = st.columns(len(recs))
            for col, item in zip(cols, recs):
                with col: _render_listing_card(item, "reco")
    _render_feature_vote("marketplace", "¿Usarías este catálogo para buscar productos geek o TCG?")


def render_sell() -> None:
    _section_header("Vender", "Publica en pocos pasos", "Prueba el recorrido de un vendedor: crear, revisar y gestionar publicaciones.")
    own = [x for x in st.session_state["marketplace_db"] if x.get("owner")]
    sm1, sm2, sm3, sm4 = st.columns(4)
    sm1.metric("Publicaciones", len(own))
    sm2.metric("Vistas", sum(x.get("views", 0) for x in own))
    sm3.metric("Favoritos", sum(x.get("likes", 0) for x in own))
    sm4.metric("Valor publicado", _fmt_clp(sum(x.get("price", 0) * x.get("stock", 1) for x in own)))

    source_tab, manual_tab, manage_tab = st.tabs(["⚡ Desde el tasador", "✍️ Publicación manual", "📊 Mis publicaciones"])
    with source_tab:
        if "df_result" not in st.session_state or st.session_state["df_result"].empty:
            st.info("Todavía no tienes inventario tasado.")
            if st.button("Ir al tasador", type="primary", key="sell_go_tasador"):
                _go_to("Tasador")
        else:
            st.success("Tu inventario tasado está disponible para publicar.")
            render_dashboard(st.session_state["df_result"], st.session_state.get("clp_rate_sidebar", 950), st.session_state.get("commission_sidebar", 5.0))

    with manual_tab:
        with st.form("manual_listing_form", clear_on_submit=True):
            a1, a2 = st.columns([2, 1])
            title = a1.text_input("Título de la publicación *", placeholder="Ej. Booster Box sellada edición japonesa")
            game = a2.selectbox("Juego / universo", ["Pokémon TCG", "One Piece Card Game", "Magic: The Gathering", "Accesorios", "Figuras y animé", "Mangas y cómics", "Otro"])
            b1, b2, b3 = st.columns(3)
            product_type = b1.selectbox("Tipo de producto", ["Carta individual", "Producto sellado", "Mazo armado", "Accesorio", "Figura", "Manga", "Otro"])
            condition = b2.selectbox("Estado", ["Nuevo", "Sellado", "NM", "LP", "MP", "HP", "Usado"])
            stock = b3.number_input("Cantidad", min_value=1, max_value=100, value=1)
            c1, c2, c3 = st.columns(3)
            price = c1.number_input("Precio CLP *", min_value=0, value=19990, step=500)
            location = c2.selectbox("Ubicación", UBICACIONES_DEMO)
            shipping = c3.selectbox("Entrega", ["Envío y retiro", "Solo envío", "Solo retiro"])
            image = st.text_input("URL de imagen (opcional)", placeholder="https://...")
            description = st.text_area("Descripción", placeholder="Incluye detalles, estado, accesorios y condiciones de entrega.")
            negotiable = st.checkbox("Precio conversable")
            submit = st.form_submit_button("🚀 Publicar en la demo", type="primary", use_container_width=True)
        if submit:
            if not title.strip() or price <= 0:
                st.error("Completa título y precio.")
            else:
                listing = {
                    "id": _new_id("manual"), "title": title.strip(), "game": game, "product_type": product_type,
                    "condition": condition, "price": int(price), "location": location, "shipping": shipping,
                    "seller": st.session_state.get("pilot_alias", "Usuario_Piloto"), "verified": False,
                    "rating": 5.0, "sales": 0, "image": image.strip(), "stock": int(stock),
                    "negotiable": negotiable, "description": description.strip() or "Publicación creada durante la prueba piloto.",
                    "tags": ["Recién publicado", product_type], "views": 0, "likes": 0, "active": True, "owner": True,
                }
                st.session_state["marketplace_db"].insert(0, listing)
                st.session_state["selected_listing"] = listing["id"]
                _notify(f"Publicaste {listing['title']}", "success")
                st.success("✅ Publicación creada. Ya puede verse en Marketplace.")

    with manage_tab:
        own = [x for x in st.session_state["marketplace_db"] if x.get("owner")]
        if not own:
            st.info("Aún no has creado publicaciones durante esta sesión.")
        for item in own:
            with st.container(border=True):
                c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
                c1.markdown(f"**{item['title']}**  \n{_fmt_clp(item['price'])} · stock {item['stock']} · {'Activa' if item.get('active') else 'Pausada'}")
                if c2.button("Ver", key=f"manage_view_{item['id']}", use_container_width=True):
                    st.session_state["selected_listing"] = item["id"]
                    _go_to("Marketplace")
                if c3.button("Pausar" if item.get("active") else "Activar", key=f"manage_toggle_{item['id']}", use_container_width=True):
                    item["active"] = not item.get("active", True); st.rerun()
                if c4.button("Eliminar", key=f"manage_delete_{item['id']}", use_container_width=True):
                    st.session_state["marketplace_db"] = [x for x in st.session_state["marketplace_db"] if x["id"] != item["id"]]
                    st.session_state["favorites"] = [x for x in st.session_state["favorites"] if x != item["id"]]
                    st.session_state["cart"] = [x for x in st.session_state["cart"] if x != item["id"]]
                    st.rerun()
    _render_feature_vote("vender", "¿El flujo para publicar te parece suficientemente simple?")


def render_catalogador(api_key: str | None, clp_rate: float, comision: float) -> None:
    _section_header("Tasación", "Catálogo e importador inteligente", "Carga un Excel o agrega cartas manualmente para identificar versiones y preparar publicaciones.")
    ul, ur = st.columns([1, 1], gap="medium")
    with ul:
        st.markdown("#### 1. Carga tu inventario")
        archivo = st.file_uploader("Excel o CSV", type=["xlsx", "xls", "csv"], label_visibility="collapsed", key="catalog_upload")
        df_ej = pd.DataFrame([
            {"nombre": "Charizard ex", "tipo": "Pokémon", "regulation_mark": "G", "numero": "234", "estado": "NM", "cantidad": 1},
            {"nombre": "Iono", "tipo": "Trainer", "regulation_mark": "G", "numero": "254", "estado": "LP", "cantidad": 2},
        ])
        buf = io.BytesIO(); df_ej.to_excel(buf, index=False); buf.seek(0)
        st.download_button("📥 Descargar plantilla", buf, "plantilla_nexogeek.xlsx", use_container_width=True)
    with ur:
        st.markdown("#### 2. Vista previa")
        df_input = None
        if archivo is not None:
            try:
                if archivo.name.lower().endswith(".csv"):
                    try: df_raw = pd.read_csv(archivo, dtype=str)
                    except UnicodeDecodeError:
                        archivo.seek(0); df_raw = pd.read_csv(archivo, dtype=str, encoding="latin-1")
                else:
                    df_raw = pd.read_excel(archivo, dtype=str)
                df_raw.columns = normalizar_columnas(df_raw.columns)
                if "numero" in df_raw.columns:
                    df_raw["numero"] = df_raw["numero"].apply(normalizar_numero)
                df_input = df_raw
                errs, warns = validar_entrada(df_input)
                for msg in warns: st.warning(msg)
                for msg in errs: st.error(msg)
                st.dataframe(df_input, use_container_width=True, height=190, hide_index=True)
            except Exception as exc:
                st.error(f"No se pudo leer el archivo: {exc}")
        else:
            st.info("Puedes usar el modo demostración para probar el flujo.")

    with st.expander("➕ Agregar una carta manualmente", expanded=False):
        with st.form("manual_card_form", clear_on_submit=True):
            m1, m2 = st.columns([2, 1])
            name = m1.text_input("Nombre oficial *", placeholder="Pikachu ex")
            card_type = m2.selectbox("Tipo", ["Pokémon", "Trainer", "Energy"])
            m3, m4, m5, m6 = st.columns(4)
            number = m3.text_input("Número", placeholder="151")
            mark = m4.text_input("Regulation mark", placeholder="H")
            condition = m5.selectbox("Estado", list(ESTADOS.keys()))
            quantity = m6.number_input("Cantidad", min_value=1, value=1)
            add_card = st.form_submit_button("⚡ Identificar y agregar", type="primary", use_container_width=True)
        if add_card:
            if not name.strip():
                st.error("El nombre es obligatorio.")
            else:
                with st.spinner("Buscando la mejor coincidencia..."):
                    result = procesar_carta({
                        "nombre": name, "tipo": card_type, "regulation_mark": mark,
                        "numero": normalizar_numero(number), "estado": condition,
                        "cantidad": quantity, "es_de_liga": "No", "set_forzado": "",
                    }, api_key, clp_rate)
                _agregar_resultado_al_dashboard(result)
                st.success(f"✅ {name} fue agregada al inventario.")

    st.markdown("#### 3. Procesa el lote")
    demo = st.checkbox("Usar inventario de demostración", value=(archivo is None), key="catalog_demo")
    if demo and archivo is None:
        df_input = pd.DataFrame([
            {"nombre": "Charizard ex", "tipo": "Pokémon", "regulation_mark": "G", "numero": "234", "estado": "NM", "cantidad": 1},
            {"nombre": "Iono", "tipo": "Trainer", "regulation_mark": "G", "numero": "269", "estado": "LP", "cantidad": 1},
            {"nombre": "Pikachu ex", "tipo": "Pokémon", "regulation_mark": "H", "numero": "219", "estado": "NM", "cantidad": 1},
            {"nombre": "Mew ex", "tipo": "Pokémon", "regulation_mark": "G", "numero": "232", "estado": "NM", "cantidad": 1},
        ])

    run = st.button("🚀 Lanzar catalogación", type="primary", use_container_width=True, disabled=df_input is None, key="catalog_run")
    if "df_result" in st.session_state:
        if st.button("🗑️ Limpiar inventario", use_container_width=True, key="catalog_reset"):
            for key in ["df_result", "candidatos", "editor_precios"]: st.session_state.pop(key, None)
            st.rerun()

    if run and df_input is not None:
        errors, warnings = validar_entrada(df_input)
        for msg in warnings: st.warning(msg)
        if errors:
            for msg in errors: st.error(msg)
        else:
            total = len(df_input)
            bar = st.progress(0); status = st.empty(); res_list = [None] * total; _reset_stats()
            filas = []
            for r in df_input.itertuples(index=False):
                filas.append({
                    "nombre": getattr(r, "nombre", ""), "tipo": getattr(r, "tipo", ""),
                    "regulation_mark": getattr(r, "regulation_mark", ""),
                    "numero": normalizar_numero(getattr(r, "numero", "")),
                    "estado": getattr(r, "estado", ""), "cantidad": getattr(r, "cantidad", 1),
                    "es_de_liga": getattr(r, "es_de_liga", "No"), "set_forzado": getattr(r, "set_forzado", ""),
                })
            completed = 0
            with ThreadPoolExecutor(max_workers=_CFG["max_workers"]) as executor:
                futures = {executor.submit(procesar_carta, row, api_key, clp_rate): idx for idx, row in enumerate(filas)}
                for future in as_completed(futures):
                    idx = futures[future]
                    try:
                        res_list[idx] = future.result()
                    except Exception:
                        fila = filas[idx]
                        res_list[idx] = {
                            "Cantidad": 1, "Estado": normalizar_estado(fila.get("estado")),
                            "Nombre Original": fila.get("nombre", ""), "Nombre EN": traducir_nombre(str(fila.get("nombre", ""))),
                            "Tipo": fila.get("tipo", ""), "Número Buscado": fila.get("numero") or "-",
                            "Regulation Mark": fila.get("regulation_mark") or "-", "Es de Liga": "No",
                            "Método Búsqueda": "error de proceso", "Card ID": "-", "Set": "Error",
                            "Número Carta": "-", "Número Coincide": "-", "Rareza": "-", "Confianza": "ninguna",
                            "Revisar": "Sí", "Precio USD Mercado": None, "Precio USD Ajustado": None,
                            "Precio CLP Sugerido": None, "Variante Precio": "-", "Fecha Precio": "-",
                            "URL Imagen": "", "_candidatos": [],
                        }
                    completed += 1; bar.progress(completed / total); status.caption(f"Procesadas {completed}/{total}")
            candidates = {}; clean = []
            for idx, result in enumerate(res_list):
                result = dict(result or {}); candidates[idx] = result.pop("_candidatos", []); clean.append(result)
            st.session_state["candidatos"] = candidates
            df_resultado = pd.DataFrame(clean)

            # Flujo híbrido: card_data entrega identificación e imágenes y la API
            # completa únicamente los precios de cada Card ID único.
            if api_key:
                status.info("✅ Cartas identificadas. Sincronizando precios de mercado...")
                price_bar = st.progress(0)
                df_resultado = enriquecer_precios_en_lote(
                    df_resultado, api_key, clp_rate,
                    progress_bar=price_bar, status_box=status,
                )
                st.session_state["ultima_sincronizacion_precios"] = datetime.now().isoformat(timespec="seconds")
            else:
                status.warning(
                    "Inventario identificado con la base local. Configura POKEMONTCG_API_KEY "
                    "en Streamlit Secrets para añadir precios automáticamente."
                )

            st.session_state["df_result"] = df_resultado
            st.rerun()

    if "df_result" in st.session_state:
        render_dashboard(st.session_state["df_result"], clp_rate, comision)
        df_cur = st.session_state["df_result"]
        if "Card ID" in df_cur.columns and "Precio USD Mercado" in df_cur.columns:
            mask = df_cur["Precio USD Mercado"].isna() & df_cur["Card ID"].notna() & (df_cur["Card ID"] != "-") & (df_cur["Card ID"] != "")
            n_sin = int(mask.sum())
        else: n_sin = 0
        if n_sin:
            st.markdown("---")
            st.info(f"Hay {n_sin} carta(s) identificadas sin precio de referencia.")
            sync1, sync2 = st.columns([2, 1])
            with sync1:
                if st.button(
                    f"💰 Sincronizar precios ({n_sin})",
                    type="primary", use_container_width=True, key="sync_prices",
                    disabled=not bool(api_key),
                ):
                    _reset_stats(); progress = st.progress(0); status2 = st.empty()
                    updated = enriquecer_precios_en_lote(
                        df_cur, api_key, clp_rate,
                        progress_bar=progress, status_box=status2,
                    )
                    st.session_state["df_result"] = updated
                    st.session_state["ultima_sincronizacion_precios"] = datetime.now().isoformat(timespec="seconds")
                    st.rerun()
            with sync2:
                if st.button(
                    "🔄 Reintentar desde cero",
                    use_container_width=True, key="force_sync_prices",
                    disabled=not bool(api_key),
                ):
                    limpiar_cache_precios(); _reset_stats()
                    progress = st.progress(0); status2 = st.empty()
                    updated = enriquecer_precios_en_lote(
                        df_cur, api_key, clp_rate,
                        progress_bar=progress, status_box=status2, force_refresh=True,
                    )
                    st.session_state["df_result"] = updated
                    st.session_state["ultima_sincronizacion_precios"] = datetime.now().isoformat(timespec="seconds")
                    st.rerun()
            if not api_key:
                st.warning(
                    "La base local carga cartas e imágenes, pero los precios requieren "
                    "POKEMONTCG_API_KEY en Streamlit Secrets."
                )
    _render_feature_vote("tasador", "¿El tasador sería una razón importante para usar NexoGeek?")


def render_auctions() -> None:
    _section_header("Subastas", "Puja por piezas especiales", "Flujo simulado con seguimiento, historial y creación de subastas.")
    st.info("🧪 Las pujas existen solo durante esta sesión y no generan obligaciones de pago.")
    with st.expander("➕ Crear una subasta de prueba", expanded=False):
        with st.form("create_auction", clear_on_submit=True):
            a1, a2 = st.columns([2, 1])
            name = a1.text_input("Producto *")
            game = a2.selectbox("Juego", ["Pokémon TCG", "One Piece Card Game", "Magic: The Gathering", "Otro"])
            b1, b2, b3 = st.columns(3)
            start_bid = b1.number_input("Puja inicial", min_value=1000, value=10000, step=1000)
            increment = b2.number_input("Incremento mínimo", min_value=500, value=1000, step=500)
            duration = b3.selectbox("Duración", ["2 horas", "6 horas", "12 horas", "24 horas"])
            image = st.text_input("URL imagen (opcional)")
            submit = st.form_submit_button("Publicar subasta", type="primary", use_container_width=True)
        if submit and name.strip():
            st.session_state["subastas_db"].insert(0, {
                "id": _new_id("auction"), "name": name.strip(), "game": game, "current_bid": int(start_bid),
                "bids": 0, "ends": duration, "image": image.strip(), "seller": st.session_state["pilot_alias"],
                "verified": False, "increment": int(increment), "watchers": 0, "history": [int(start_bid)],
            })
            _notify(f"Subasta creada: {name}", "success"); st.success("Subasta creada para la demo.")

    f1, f2 = st.columns(2)
    game_filter = f1.selectbox("Filtrar por juego", ["Todos"] + sorted(set(x["game"] for x in st.session_state["subastas_db"])), key="auction_filter")
    watch_only = f2.checkbox("Solo las que sigo", key="auction_watch_only")
    auctions = [x for x in st.session_state["subastas_db"] if game_filter == "Todos" or x["game"] == game_filter]
    if watch_only:
        auctions = [x for x in auctions if x["id"] in st.session_state.get("auction_watchlist", [])]

    for start in range(0, len(auctions), 3):
        cols = st.columns(3)
        for col, auction in zip(cols, auctions[start:start+3]):
            with col:
                with st.container(border=True):
                    _render_image_or_placeholder(auction.get("image", ""), "🔨", width=210)
                    st.markdown(f"<span class='badge-demo'>{_safe_text(auction['game'])}</span>", unsafe_allow_html=True)
                    st.markdown(f"### {_safe_text(auction['name'])}", unsafe_allow_html=True)
                    st.metric("Oferta actual", _fmt_clp(auction["current_bid"]))
                    st.caption(f"⏱ {auction['ends']} · {auction['bids']} pujas · {auction.get('watchers',0)} seguidores")
                    verified = "✅" if auction.get("verified") else "🟡"
                    st.caption(f"{verified} {auction['seller']}")
                    watch_label = "★ Siguiendo" if auction["id"] in st.session_state.get("auction_watchlist", []) else "☆ Seguir"
                    if st.button(watch_label, key=f"watch_{auction['id']}", use_container_width=True):
                        added = _toggle_item("auction_watchlist", auction["id"])
                        auction["watchers"] = max(0, auction.get("watchers", 0) + (1 if added else -1)); st.rerun()
                    with st.form(f"bid_form_{auction['id']}", clear_on_submit=True):
                        minimum = int(auction["current_bid"] + auction["increment"])
                        amount = st.number_input("Tu puja", min_value=minimum, value=minimum, step=int(auction["increment"]), key=f"bid_amount_{auction['id']}")
                        bid = st.form_submit_button("🔨 Confirmar puja", type="primary", use_container_width=True)
                    if bid:
                        auction["current_bid"] = int(amount); auction["bids"] += 1; auction.setdefault("history", []).append(int(amount))
                        _notify(f"Lideras la subasta de {auction['name']} con {_fmt_clp(amount)}", "success")
                        st.success("Puja simulada registrada."); st.rerun()
                    with st.expander("Ver historial"):
                        for value in reversed(auction.get("history", [])[-5:]): st.caption(_fmt_clp(value))
    _render_feature_vote("subastas", "¿Participarías en subastas dentro de una plataforma así?")


def render_services() -> None:
    _section_header("Servicios", "Talento y apoyo para la comunidad", "Pregrading, encargos, impresión 3D, organización y otros servicios especializados.")
    with st.expander("📢 Ofrecer un servicio", expanded=False):
        with st.form("service_form", clear_on_submit=True):
            s1, s2 = st.columns([1, 2])
            service_type = s1.selectbox("Categoría", ["Pregrading", "Encargos", "Diseño e impresión", "Organización", "Otros"])
            title = s2.text_input("Título *")
            description = st.text_area("Descripción")
            a1, a2, a3 = st.columns(3)
            price = a1.number_input("Precio desde", min_value=0, value=10000, step=500)
            location = a2.selectbox("Ubicación", UBICACIONES_DEMO)
            delivery = a3.text_input("Plazo", value="A coordinar")
            submit = st.form_submit_button("Publicar servicio", type="primary", use_container_width=True)
        if submit and title.strip():
            st.session_state["servicios_db"].insert(0, {
                "id": _new_id("service"), "type": service_type, "title": title.strip(),
                "description": description.strip() or "Servicio publicado durante el piloto.", "price": int(price),
                "provider": st.session_state["pilot_alias"], "rating": 5.0, "location": location,
                "badge": "🆕 Nuevo", "delivery": delivery.strip() or "A coordinar",
            })
            _notify(f"Servicio publicado: {title}", "success"); st.success("Servicio publicado.")

    s1, s2, s3 = st.columns(3)
    types = ["Todos"] + sorted(set(x["type"] for x in st.session_state["servicios_db"]))
    type_filter = s1.selectbox("Categoría", types, key="services_type")
    location_filter = s2.selectbox("Ubicación", ["Todas"] + sorted(set(x["location"] for x in st.session_state["servicios_db"])), key="services_location")
    sort = s3.selectbox("Ordenar", ["Recomendados", "Menor precio", "Mejor evaluados"], key="services_sort")
    services = [x for x in st.session_state["servicios_db"] if (type_filter == "Todos" or x["type"] == type_filter) and (location_filter == "Todas" or x["location"] == location_filter)]
    if sort == "Menor precio": services.sort(key=lambda x: x["price"])
    elif sort == "Mejor evaluados": services.sort(key=lambda x: -x["rating"])
    else: services.sort(key=lambda x: (-x["rating"], x["price"]))

    cols = st.columns(2)
    for idx, service in enumerate(services):
        with cols[idx % 2]:
            with st.container(border=True):
                st.markdown(f"<span class='badge-demo'>{_safe_text(service['type'])}</span> <span class='badge-verified'>{_safe_text(service['badge'])}</span>", unsafe_allow_html=True)
                st.markdown(f"### {_safe_text(service['title'])}", unsafe_allow_html=True)
                st.write(service["description"])
                a1, a2, a3 = st.columns(3)
                a1.metric("Desde", _fmt_clp(service["price"]))
                a2.metric("Rating", f"⭐ {service['rating']:.1f}")
                a3.metric("Plazo", service["delivery"])
                st.caption(f"👤 {service['provider']} · 📍 {service['location']}")
                if st.button("💬 Solicitar información", key=f"service_contact_{service['id']}", type="primary", use_container_width=True):
                    _notify(f"Solicitud simulada enviada a {service['provider']}", "success")
                    st.success("Contacto simulado iniciado.")
    _render_feature_vote("servicios", "¿Buscarías u ofrecerías servicios especializados aquí?")


def render_feedback() -> None:
    _section_header("Validación", "Tu opinión define la siguiente versión", "Queremos entender qué genera valor, qué causa desconfianza y qué debería priorizarse.")
    with st.form("extended_feedback_form", clear_on_submit=True):
        f1, f2 = st.columns(2)
        name = f1.text_input("Nombre o alias (opcional)")
        profile = f2.selectbox("Perfil", ["Coleccionista", "Jugador competitivo", "Vendedor o tienda", "Organizador", "Nuevo en el hobby", "Otro"])
        n1, n2 = st.columns(2)
        score = n1.slider("Nota general", 1, 10, 8)
        nps = n2.slider("¿Qué tan probable es que la recomendaras?", 0, 10, 8)
        favorite_module = st.selectbox("Función más valiosa", ["Marketplace", "Tasador", "Publicación de productos", "Subastas", "Servicios", "Perfiles y reputación", "Otra"])
        priorities = st.multiselect("¿Qué deberíamos construir primero?", ["Pagos protegidos", "Sistema de envíos", "Chat interno", "Verificación de vendedores", "App móvil", "Alertas y favoritos", "Más juegos TCG", "Integración con tiendas", "Subastas reales", "Servicios comunitarios"], max_selections=3)
        trust = st.selectbox("¿Qué te daría más confianza?", ["Pago retenido hasta confirmar entrega", "Vendedores verificados", "Reputación y comentarios", "Fotos obligatorias", "Moderación y reportes", "Punto de encuentro seguro", "Otro"])
        willingness = st.selectbox("Comisión aceptable para una venta protegida", ["No pagaría comisión", "1% a 3%", "4% a 6%", "7% a 10%", "Depende del servicio"])
        use_intent = st.radio("¿La usarías?", ["Sí", "Tal vez", "No"], horizontal=True)
        useful = st.text_area("¿Qué fue lo mejor?")
        improve = st.text_area("¿Qué cambiarías o agregarías?")
        contact = st.text_input("Contacto para futuras pruebas (opcional)")
        submit = st.form_submit_button("Enviar feedback", type="primary", use_container_width=True)
    if submit:
        if not useful.strip() and not improve.strip():
            st.warning("Escribe al menos una observación.")
        else:
            data = {
                "nombre": name.strip(), "perfil": profile, "nota": score, "nps": nps,
                "modulo_favorito": favorite_module, "prioridades": " | ".join(priorities),
                "factor_confianza": trust, "comision_aceptable": willingness, "usaria": use_intent,
                "lo_mejor": useful.strip(), "mejoraria": improve.strip(), "contacto": contact.strip(),
            }
            try:
                _save_extended_feedback(data)
                st.success("✅ Gracias. Tu respuesta quedó guardada.")
            except Exception as exc:
                st.error(f"No se pudo guardar: {exc}")

    feedback = _read_csv_safe(EXTENDED_FEEDBACK_FILE)
    if not feedback.empty:
        st.markdown("---")
        _section_header("Resultados del piloto", "Resumen de respuestas", "Indicadores locales obtenidos durante las pruebas.")
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Respuestas", len(feedback))
        r2.metric("Nota promedio", f"{pd.to_numeric(feedback['nota'], errors='coerce').mean():.1f}/10")
        r3.metric("NPS promedio", f"{pd.to_numeric(feedback['nps'], errors='coerce').mean():.1f}/10")
        r4.metric("La usaría", f"{(feedback['usaria'].astype(str).str.lower() == 'sí').mean()*100:.0f}%")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Funciones más valoradas")
            counts = feedback["modulo_favorito"].value_counts()
            st.bar_chart(counts)
        with c2:
            st.markdown("#### Perfiles participantes")
            counts = feedback["perfil"].value_counts()
            st.bar_chart(counts)
        st.dataframe(feedback.tail(20), use_container_width=True, hide_index=True)



# ══════════════════════════════════════════════════════════════════════════════
# NEXOGEEK PILOTO V3 — EXPERIENCIA PÚBLICA, ONBOARDING Y ANALÍTICA
# Estas definiciones reemplazan las versiones anteriores manteniendo el motor
# de tasación y los módulos ya construidos.
# ══════════════════════════════════════════════════════════════════════════════

JOURNEY_FILE = os.getenv("NEXOGEEK_JOURNEY_FILE", "recorrido_nexogeek.csv")
ADMIN_PIN = os.getenv("NEXOGEEK_ADMIN_PIN", "2026")

NEXO_COLORS = {
    "ink": "#28163A", "plum": "#5B2A86", "violet": "#8A4FFF",
    "coral": "#FF6B6B", "sun": "#FFC857", "mint": "#2EC4B6",
    "cream": "#FFF8ED", "paper": "#FFFDF8", "soft": "#F3E9FF",
    "muted": "#73667E",
}

# Unificamos también los elementos internos del tasador con la marca.
TIPO_COLORES.update({
    "Pokémon": {"bg": "#FFFDF8", "border": "#5B2A86", "badge": "#F3E9FF", "text": "#5B2A86"},
    "pokemon": {"bg": "#FFFDF8", "border": "#5B2A86", "badge": "#F3E9FF", "text": "#5B2A86"},
    "Trainer": {"bg": "#FFFDF8", "border": "#2EC4B6", "badge": "#DDF8F3", "text": "#185D56"},
    "trainer": {"bg": "#FFFDF8", "border": "#2EC4B6", "badge": "#DDF8F3", "text": "#185D56"},
    "Energy": {"bg": "#FFFDF8", "border": "#FFC857", "badge": "#FFF0CC", "text": "#7C4B00"},
    "energy": {"bg": "#FFFDF8", "border": "#FFC857", "badge": "#FFF0CC", "text": "#7C4B00"},
})
ESTADO_COLOR.update({
    "NM": ("#DDF8F3", "#185D56"), "LP": ("#F3E9FF", "#512176"),
    "MP": ("#FFF0CC", "#7C4B00"), "HP": ("#FFE1D7", "#8B341E"),
    "DMG": ("#FFE0E0", "#8E2430"),
})

PILOT_V3_CSS = """
<style>
/* Oculta cromados que distraen durante una presentación, sin impedir Deploy. */
#MainMenu {visibility:hidden;}
footer {visibility:hidden;}

/* Componentes de experiencia */
.welcome-deck {
    background:linear-gradient(135deg,#FFF0CC 0%,#FFF8ED 48%,#DDF8F3 100%);
    border:3px solid var(--ng-ink); border-radius:24px 24px 8px 24px;
    padding:24px; margin:0 4px 25px 0; box-shadow:7px 7px 0 var(--ng-coral);
}
.welcome-deck h2 {margin:6px 0 8px;color:var(--ng-ink);font-family:"Arial Black","Trebuchet MS",sans-serif;letter-spacing:-.045em;}
.welcome-deck p {margin:0;color:var(--ng-muted);max-width:850px;line-height:1.55;}
.tour-badge {display:inline-flex;align-items:center;gap:7px;padding:5px 9px;background:#FFFDF8;border:2px solid var(--ng-ink);border-radius:8px 8px 3px 8px;font-size:.68rem;font-weight:900;}
.journey-card {background:rgba(255,253,248,.12);border:1px solid rgba(255,248,237,.25);border-radius:15px;padding:12px;margin:12px 0;color:#FFF8ED;}
.journey-card strong {display:block;color:#FFF8ED;margin-bottom:6px;}
.journey-card small {color:#DCCBEA;}

.nexo-empty {
    border:2px dashed var(--ng-plum);border-radius:20px 20px 7px 20px;padding:34px 24px;text-align:center;
    background:linear-gradient(145deg,rgba(243,233,255,.7),rgba(255,248,237,.86));
}
.nexo-empty .icon {font-size:2.4rem;margin-bottom:8px;}.nexo-empty h3{margin:0 0 5px;color:var(--ng-ink)}.nexo-empty p{margin:0;color:var(--ng-muted)}

/* Arte editorial cuando una publicación demo no tiene fotografía. */
.demo-art {
    position:relative;height:235px;overflow:hidden;border:2px solid var(--ng-ink);border-radius:20px 20px 7px 20px;
    display:flex;align-items:flex-end;padding:18px;box-shadow:4px 4px 0 rgba(40,22,58,.18);
}
.demo-art::before,.demo-art::after{content:"";position:absolute;border-radius:50%;border:3px solid rgba(255,255,255,.58)}
.demo-art::before{width:150px;height:150px;right:-38px;top:-35px}.demo-art::after{width:74px;height:74px;left:18px;top:20px}
.demo-art-purple{background:linear-gradient(145deg,#3B1D4F,#8A4FFF)}
.demo-art-coral{background:linear-gradient(145deg,#8F3048,#FF6B6B)}
.demo-art-mint{background:linear-gradient(145deg,#12665D,#2EC4B6)}
.demo-art-sun{background:linear-gradient(145deg,#B96A00,#FFC857)}
.demo-art-ink{background:linear-gradient(145deg,#161020,#5B2A86)}
.demo-art-copy{position:relative;z-index:2;color:#FFFDF8;text-shadow:0 2px 8px rgba(0,0,0,.24)}
.demo-art-icon{font-size:3.25rem;line-height:1;margin-bottom:12px}.demo-art-label{font-family:"Arial Black","Trebuchet MS";font-size:1.05rem;line-height:1.05;max-width:190px}.demo-art-kicker{font-size:.62rem;font-weight:900;letter-spacing:.12em;margin-top:7px;opacity:.85}
.remote-art {height:235px;border:2px solid var(--ng-ink);border-radius:20px 20px 7px 20px;background:#F3E9FF;display:flex;align-items:center;justify-content:center;overflow:hidden;box-shadow:4px 4px 0 rgba(40,22,58,.18)}
.remote-art img{width:100%;height:100%;object-fit:contain;padding:8px;box-sizing:border-box}
.remote-art.detail-art{height:520px}.remote-art.detail-art img{padding:18px}
.demo-art.detail-art{height:520px}.demo-art.detail-art .demo-art-icon{font-size:5rem}.demo-art.detail-art .demo-art-label{font-size:1.65rem;max-width:330px}

.product-breadcrumb {font-size:.78rem;color:var(--ng-muted);margin:3px 0 15px}.product-breadcrumb b{color:var(--ng-plum)}
.product-headline {font-family:"Arial Black","Trebuchet MS",sans-serif;color:var(--ng-ink);font-size:clamp(1.65rem,3vw,2.65rem);letter-spacing:-.055em;line-height:1.04;margin:10px 0 12px}
.product-detail-price {font-family:"Arial Black","Trebuchet MS";color:var(--ng-plum);font-size:2.15rem;letter-spacing:-.05em;margin:12px 0 4px}
.detail-fact {background:#FFFDF8;border:2px solid var(--ng-ink);border-radius:13px 13px 4px 13px;padding:11px;min-height:68px;box-shadow:3px 3px 0 rgba(46,196,182,.45)}
.detail-fact small{display:block;color:var(--ng-muted);font-size:.68rem}.detail-fact strong{display:block;color:var(--ng-ink);margin-top:3px}
.seller-profile {background:linear-gradient(145deg,#F3E9FF,#FFFDF8);border:2px solid var(--ng-ink);border-radius:20px 20px 7px 20px;padding:20px;box-shadow:5px 5px 0 var(--ng-sun)}
.seller-avatar {width:55px;height:55px;display:flex;align-items:center;justify-content:center;border:2px solid var(--ng-ink);background:var(--ng-mint);border-radius:16px 16px 5px 16px;font-family:"Arial Black";font-size:1.05rem;box-shadow:3px 3px 0 var(--ng-coral)}
.safety-grid {display:grid;grid-template-columns:repeat(3,1fr);gap:12px}.safety-item{background:#FFFDF8;border:2px solid var(--ng-ink);border-radius:15px 15px 5px 15px;padding:14px}.safety-item strong{display:block;color:var(--ng-ink);font-size:.88rem}.safety-item span{color:var(--ng-muted);font-size:.76rem;line-height:1.35}
.photo-note {color:var(--ng-muted);font-size:.72rem;text-align:center;margin-top:7px}

/* Tarjeta más limpia y menos cargada de controles. */
.listing-card-actions [data-testid="stButton"] button{font-size:.82rem!important}
.card-trust-line{display:flex;align-items:center;gap:6px;flex-wrap:wrap;color:var(--ng-muted);font-size:.74rem;margin:7px 0 10px}
.card-stock{font-size:.68rem;color:#7C4B00;background:#FFF0CC;border:1px solid #F4C15A;border-radius:999px;padding:2px 7px}
.active-filter{display:inline-block;background:#F3E9FF;color:#512176;border:1px solid #A875D2;padding:4px 8px;border-radius:999px;font-size:.68rem;font-weight:800;margin:0 5px 5px 0}

/* Panel anfitrión visualmente separado de la experiencia pública. */
.admin-lock {background:rgba(255,248,237,.08);border:1px dashed rgba(255,248,237,.28);padding:10px;border-radius:12px;color:#DCCBEA;font-size:.75rem}
.host-badge{display:inline-block;background:#FFC857;color:#28163A;border:2px solid #28163A;border-radius:8px 8px 3px 8px;padding:4px 8px;font-size:.68rem;font-weight:950;box-shadow:2px 2px 0 #FF6B6B}

/* Barra de progreso propia */
.ng-progress {height:10px;border:1px solid rgba(255,248,237,.28);background:rgba(255,255,255,.08);border-radius:999px;overflow:hidden;margin:7px 0}
.ng-progress > span{display:block;height:100%;background:linear-gradient(90deg,var(--ng-coral),var(--ng-sun),var(--ng-mint));border-radius:999px}

/* Móvil */
@media (max-width: 780px) {
    .block-container{padding:1rem .78rem 2.5rem!important;max-width:100%!important}
    .nexo-hero{padding:2rem 1.15rem!important;border-radius:20px 20px 7px 20px!important;box-shadow:6px 6px 0 var(--ng-sun)!important}
    .nexo-hero-copy{width:100%!important}.nexo-hero h1{font-size:2.2rem!important}.nexo-hero::after{display:none}
    .nexo-brand-tag{display:none}.nexo-brand-name{font-size:1.25rem}.nexo-brand-mark{width:35px;height:35px}
    div[data-testid="stButton"] > button{min-height:2.8rem!important;font-size:.82rem!important;white-space:normal!important}
    .demo-art,.remote-art{height:205px}.demo-art.detail-art,.remote-art.detail-art{height:390px}
    .safety-grid{grid-template-columns:1fr}.section-title{font-size:1.55rem!important}.welcome-deck{padding:18px}
    [data-testid="stSidebar"]{min-width:min(88vw,340px)!important}
    .product-headline{font-size:1.8rem}.product-detail-price{font-size:1.8rem}
}
</style>
"""

# Conservamos referencias para envolver módulos ya probados.
_legacy_render_catalogador_v3 = render_catalogador
_legacy_render_auctions_v3 = render_auctions
_legacy_render_services_v3 = render_services
_legacy_demo_marketplace_seed_v3 = _demo_marketplace_seed
_legacy_listing_from_inventory_v3 = _listing_from_inventory


def _demo_marketplace_seed() -> list[dict]:
    items = _legacy_demo_marketplace_seed_v3()
    defaults = {
        "response_time": "Responde en menos de 2 horas", "member_since": "Miembro desde 2024",
        "photo_count": 3, "protected": True, "authenticity": "Fotos y estado declarados",
    }
    for item in items:
        for key, value in defaults.items():
            item.setdefault(key, value)

    extras = [
        {
            "id":"p-004","title":"Gardevoir ex · Special Illustration Rare #245","game":"Pokémon TCG",
            "product_type":"Carta individual","condition":"NM","price":38990,"location":"Santiago",
            "shipping":"Envío y retiro","seller":"MoonCard_CL","verified":True,"rating":4.9,"sales":77,
            "image":"https://images.pokemontcg.io/sv1/245.png","stock":1,"negotiable":False,
            "description":"Carta protegida desde apertura. Incluye sleeve, perfect fit y top loader.",
            "tags":["SIR","Colección"],"views":98,"likes":16,"active":True,"owner":False,
            "response_time":"Responde en 35 minutos","member_since":"Miembro desde 2023","photo_count":4,"protected":True,
            "authenticity":"Fotos reales disponibles en el chat",
        },
        {
            "id":"acc-002","title":"Playmat ilustrado edición comunidad","game":"Accesorios",
            "product_type":"Accesorio","condition":"Nuevo","price":21990,"location":"Ñuñoa",
            "shipping":"Envío y retiro","seller":"PortalCraft","verified":True,"rating":4.9,"sales":46,
            "image":"","stock":6,"negotiable":False,"description":"Playmat de 60 × 35 cm, superficie suave y base antideslizante.",
            "tags":["Diseño original","Comunidad"],"views":87,"likes":14,"active":True,"owner":False,
            "response_time":"Responde en 1 hora","member_since":"Miembro desde 2025","photo_count":5,"protected":True,
            "authenticity":"Diseño del creador verificado",
        },
        {
            "id":"board-001","title":"Juego de mesa cooperativo · edición en español","game":"Juegos de mesa",
            "product_type":"Juego de mesa","condition":"Usado","price":32990,"location":"Providencia",
            "shipping":"Solo retiro","seller":"MesaCritica","verified":True,"rating":4.8,"sales":35,
            "image":"","stock":1,"negotiable":True,"description":"Componentes completos y ordenados. Caja con desgaste superficial leve.",
            "tags":["Completo","Conversable"],"views":64,"likes":10,"active":True,"owner":False,
            "response_time":"Responde el mismo día","member_since":"Miembro desde 2024","photo_count":6,"protected":True,
            "authenticity":"Inventario de componentes revisado",
        },
    ]
    return items + extras


def _init_demo_state() -> None:
    defaults = {
        "page": "Inicio", "marketplace_db": _demo_marketplace_seed(), "favorites": [], "cart": [],
        "compare": [], "notifications": [], "selected_listing": None, "auction_watchlist": [],
        "subastas_db": _demo_auctions_seed(), "servicios_db": _demo_services_seed(), "feature_votes": {},
        "pilot_alias": "Usuario_Piloto", "pilot_location": "Santiago", "pilot_role": "",
        "onboarding_complete": False, "admin_unlocked": False, "_journey_seen": set(),
        "_journey_actions": set(), "session_id": _new_id("session"),
        "api_key_guardada": "", "clp_rate_sidebar": 950, "commission_sidebar": 5.0,
        "workers_sidebar": 4, "throttle_sidebar": 0.2,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _track_event(event: str, object_id: str = "", metadata: str = "", once: bool = False) -> None:
    event = str(event).strip()
    object_id = str(object_id or "").strip()
    key = f"{event}|{object_id}"
    seen = st.session_state.setdefault("_journey_seen", set())
    if once and key in seen:
        return
    seen.add(key)
    st.session_state.setdefault("_journey_actions", set()).add(event)
    row = pd.DataFrame([{
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "session_id": st.session_state.get("session_id", ""),
        "alias": st.session_state.get("pilot_alias", ""),
        "rol": st.session_state.get("pilot_role", ""),
        "evento": event, "objeto": object_id, "detalle": str(metadata or "")[:500],
    }])
    try:
        with _FEEDBACK_LOCK:
            exists = os.path.exists(JOURNEY_FILE)
            row.to_csv(JOURNEY_FILE, mode="a", header=not exists, index=False, encoding="utf-8-sig")
    except Exception:
        pass


def _tour_status() -> tuple[int, int, list[tuple[str, bool]]]:
    actions = st.session_state.get("_journey_actions", set())
    milestones = [
        ("Explorar el marketplace", "visita_marketplace" in actions),
        ("Abrir una ficha", "abrir_ficha" in actions),
        ("Guardar un favorito", "favorito" in actions),
        ("Simular una reserva", "reserva" in actions),
        ("Visitar el tasador o publicar", bool({"visita_tasador", "publicar"} & set(actions))),
    ]
    completed = sum(done for _, done in milestones)
    return completed, len(milestones), milestones


def _render_empty_state(icon: str, title: str, copy: str) -> None:
    st.markdown(
        f'<div class="nexo-empty"><div class="icon">{icon}</div><h3>{_safe_text(title)}</h3>'
        f'<p>{_safe_text(copy)}</p></div>', unsafe_allow_html=True,
    )


def _visual_style_for_listing(listing: dict) -> tuple[str, str, str]:
    kind = f"{listing.get('product_type','')} {listing.get('game','')}".lower()
    if "manga" in kind or "cómic" in kind or "comic" in kind: return "📚", "Biblioteca del Nexo", "demo-art-coral"
    if "figura" in kind or "animé" in kind or "anime" in kind: return "🗿", "Pieza de vitrina", "demo-art-purple"
    if "one piece" in kind: return "🏴‍☠️", "Tesoro de Grand Line", "demo-art-sun"
    if "magic" in kind: return "✦", "Reliquia de maná", "demo-art-ink"
    if "sellado" in kind: return "📦", "Caja sellada", "demo-art-purple"
    if "juego de mesa" in kind: return "🎲", "Noche de mesa", "demo-art-mint"
    if "accesorio" in kind: return "🧰", "Equipo para tu colección", "demo-art-mint"
    if "mazo" in kind: return "🗂️", "Listo para jugar", "demo-art-ink"
    return "🃏", "Hallazgo coleccionable", "demo-art-coral"


def _render_image_or_placeholder(image: str, emoji: str = "🃏", width: int = 220,
                                 label: str = "Imagen de demostración", detail: bool = False,
                                 listing: dict | None = None) -> None:
    safe_url = _safe_image_url(image)
    extra = " detail-art" if detail else ""
    if safe_url:
        st.markdown(
            f'<div class="remote-art{extra}"><img src="{safe_url}" alt="{_safe_text(label)}"></div>',
            unsafe_allow_html=True,
        )
        return
    if listing:
        emoji, label, style = _visual_style_for_listing(listing)
    else:
        style = "demo-art-purple"
    st.markdown(
        f'<div class="demo-art {style}{extra}"><div class="demo-art-copy">'
        f'<div class="demo-art-icon">{emoji}</div><div class="demo-art-label">{_safe_text(label)}</div>'
        f'<div class="demo-art-kicker">NEXOGEEK · VISUAL DEMO</div></div></div>',
        unsafe_allow_html=True,
    )


def _listing_from_inventory(fila: dict, price: int, quantity: int, condition: str,
                            location: str, shipping: str, negotiable: bool,
                            description: str, track: bool = True) -> dict:
    listing = _legacy_listing_from_inventory_v3(
        fila, price, quantity, condition, location, shipping, negotiable, description
    )
    listing.update({
        "response_time":"Perfil nuevo · respuesta por confirmar", "member_since":"Se unió durante el piloto",
        "photo_count":1, "protected":True, "authenticity":"Carta vinculada al tasador de NexoGeek",
    })
    if track:
        _track_event("publicar", listing.get("id", ""), "publicación desde tasador")
    return listing


def _badge_metodo(metodo: str) -> str:
    palette = {
        "número exacto": ("#DDF8F3", "#185D56", "Exacta"),
        "nombre + bloque": ("#F3E9FF", "#512176", "Bloque"),
        "nombre + tipo": ("#FFF0CC", "#7C4B00", "Tipo"),
        "solo nombre": ("#FFE1D7", "#8B341E", "Nombre"),
        "set forzado": ("#F3E9FF", "#512176", "Set fijado"),
        "selección manual": ("#DDF8F3", "#185D56", "Verificada"),
        "sin resultados": ("#FFE0E0", "#8E2430", "Sin resultado"),
    }
    bg, fg, text = palette.get(metodo, ("#FFF8ED", "#73667E", metodo or "—"))
    return f'<span style="background:{bg};color:{fg};border:1px solid {fg}55;font-size:.64rem;font-weight:800;padding:3px 7px;border-radius:8px 8px 3px 8px;margin-right:4px;">{_safe_text(text)}</span>'


def _badge_confianza(conf: str) -> str:
    m = {
        "alta": ("#DDF8F3", "#185D56", "Alta"), "media": ("#FFF0CC", "#7C4B00", "Media"),
        "baja": ("#FFE1D7", "#8B341E", "Baja"), "ninguna": ("#FFF8ED", "#73667E", "Sin datos"),
    }
    bg, fg, txt = m.get(conf, m["ninguna"])
    return f'<span style="background:{bg};color:{fg};border:1px solid {fg}55;font-size:.64rem;font-weight:800;padding:3px 7px;border-radius:8px 8px 3px 8px;">{txt}</span>'


def _render_listing_card(listing: dict, prefix: str) -> None:
    lid = listing["id"]
    favorites = st.session_state.get("favorites", [])
    compare = st.session_state.get("compare", [])
    with st.container(border=True):
        placeholder_emoji, placeholder_label = _placeholder_for_listing(listing)
        _render_image_or_placeholder(
            listing.get("image", ""), placeholder_emoji, width=210,
            label=placeholder_label, listing=listing,
        )
        tags = listing.get("tags", [])[:2]
        if tags:
            st.markdown(" ".join(f'<span class="badge-demo">{_safe_text(tag)}</span>' for tag in tags), unsafe_allow_html=True)
        st.markdown(f'<div class="listing-title">{_safe_text(listing.get("title", ""))}</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="listing-meta">{_safe_text(listing.get("game", ""))} · {_safe_text(listing.get("condition", ""))}<br>'
            f'{_safe_text(listing.get("location", ""))} · {_safe_text(listing.get("shipping", ""))}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(f'<div class="listing-price">{_fmt_clp(listing.get("price"))}</div>', unsafe_allow_html=True)
        verified = '<span class="badge-verified">Verificado</span>' if listing.get("verified") else '<span class="badge-stock">Perfil nuevo</span>'
        st.markdown(
            f'<div class="card-trust-line"><span>{_safe_text(listing.get("seller", ""))}</span>{verified}'
            f'<span>★ {float(listing.get("rating",0)):.1f}</span><span class="card-stock">{listing.get("stock",0)} disponible(s)</span></div>',
            unsafe_allow_html=True,
        )
        a1, a2 = st.columns([3, 1])
        if a1.button("Ver ficha", key=f"{prefix}_view_{lid}", type="primary", use_container_width=True):
            listing["views"] = int(listing.get("views", 0)) + 1
            st.session_state["selected_listing"] = lid
            _track_event("abrir_ficha", lid, listing.get("title", ""))
            _go_to("Detalle")
        fav_label = "♥" if lid in favorites else "♡"
        if a2.button(fav_label, key=f"{prefix}_fav_{lid}", use_container_width=True, help="Guardar en favoritos"):
            added = _toggle_item("favorites", lid)
            listing["likes"] = max(0, int(listing.get("likes", 0)) + (1 if added else -1))
            if added:
                _track_event("favorito", lid, listing.get("title", ""))
            _notify(("Guardaste" if added else "Quitaste") + f" {listing['title']} de favoritos", "success")
            st.rerun()
        c1, c2 = st.columns([1.35, 1])
        compare_label = "✓ Comparando" if lid in compare else "Comparar"
        if c1.button(compare_label, key=f"{prefix}_cmp_{lid}", use_container_width=True):
            if lid in compare:
                _toggle_item("compare", lid)
            elif len(compare) >= 3:
                _notify("Puedes comparar hasta 3 publicaciones", "warning")
            else:
                _toggle_item("compare", lid, max_items=3)
                _track_event("comparar", lid, listing.get("title", ""))
            st.rerun()
        c2.caption(f"{listing.get('views',0)} vistas · {listing.get('likes',0)} guardados")


def _render_seller_profile(listing: dict) -> None:
    initials = "".join(
        part[:1] for part in str(listing.get("seller", "NG")).replace("_", " ").split()[:2]
    ).upper() or "NG"
    status = "Vendedor verificado" if listing.get("verified") else "Perfil en etapa piloto"
    html = f"""
    <div class="seller-profile">
      <div style="display:flex;gap:14px;align-items:center;margin-bottom:14px;">
        <div class="seller-avatar">{_safe_text(initials)}</div>
        <div>
          <div style="font-weight:900;color:var(--ng-ink);font-size:1.05rem;">{_safe_text(listing.get('seller',''))}</div>
          <div style="color:var(--ng-muted);font-size:.78rem;">{_safe_text(status)} · ★ {float(listing.get('rating',0)):.1f} · {listing.get('sales',0)} ventas</div>
        </div>
      </div>
      <div style="font-weight:800;color:var(--ng-ink);">{_safe_text(listing.get('response_time','Respuesta por confirmar'))}</div>
      <div style="color:var(--ng-muted);font-size:.8rem;margin-top:4px;">{_safe_text(listing.get('member_since','Miembro del piloto'))}</div>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


def render_product_detail() -> None:
    listing = _listing_by_id(st.session_state.get("selected_listing", ""))
    if not listing:
        _render_empty_state("🧭", "No encontramos esa publicación", "Regresa al marketplace para elegir otro hallazgo.")
        if st.button("Volver al marketplace", type="primary"):
            _go_to("Marketplace")
        return

    _track_event("visita_detalle", listing["id"], once=True)
    if st.button("← Volver al marketplace", key="detail_back"):
        _go_to("Marketplace")
    st.markdown(
        f'<div class="product-breadcrumb">Marketplace / {_safe_text(listing.get("game",""))} / '
        f'<b>{_safe_text(listing.get("product_type",""))}</b></div>', unsafe_allow_html=True,
    )

    left, right = st.columns([1.03, 1.15], gap="large")
    with left:
        emoji, label = _placeholder_for_listing(listing)
        _render_image_or_placeholder(listing.get("image", ""), emoji, 430, label, detail=True, listing=listing)
        st.markdown(
            f'<div class="photo-note">{int(listing.get("photo_count",3))} imágenes declaradas · Visual principal de demostración</div>',
            unsafe_allow_html=True,
        )
        t1, t2, t3 = st.columns(3)
        t1.button("Vista principal", key="photo_main", disabled=True, use_container_width=True)
        t2.button("Estado", key="photo_state", disabled=True, use_container_width=True)
        t3.button("Reverso", key="photo_back", disabled=True, use_container_width=True)

    with right:
        tags = listing.get("tags", [])
        if tags:
            st.markdown(" ".join(f'<span class="badge-demo">{_safe_text(t)}</span>' for t in tags), unsafe_allow_html=True)
        st.markdown(f'<div class="product-headline">{_safe_text(listing.get("title",""))}</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="product-detail-price">{_fmt_clp(listing.get("price"))}</div>', unsafe_allow_html=True)
        if listing.get("negotiable"):
            st.caption("Precio conversable con el vendedor")

        facts = [
            ("Estado", listing.get("condition", "—")), ("Disponibilidad", f"{listing.get('stock',0)} unidad(es)"),
            ("Ubicación", listing.get("location", "—")), ("Entrega", listing.get("shipping", "—")),
        ]
        fc = st.columns(4)
        for col, (label_f, value) in zip(fc, facts):
            col.markdown(f'<div class="detail-fact"><small>{_safe_text(label_f)}</small><strong>{_safe_text(value)}</strong></div>', unsafe_allow_html=True)

        st.markdown("#### Sobre esta pieza")
        st.write(listing.get("description", "Sin descripción disponible."))
        st.caption(listing.get("authenticity", "Estado declarado por el vendedor"))

        a1, a2 = st.columns(2)
        if a1.button("Contactar al vendedor", key=f"detail_contact_{listing['id']}", type="primary", use_container_width=True):
            _track_event("contactar", listing["id"], listing.get("seller", ""))
            _notify(f"Chat simulado abierto con {listing['seller']}", "success")
            st.success("Chat simulado abierto. No se comparte información personal en esta demo.")
        if a2.button("Agregar a reserva", key=f"detail_reserve_{listing['id']}", use_container_width=True):
            if listing["id"] not in st.session_state["cart"]:
                st.session_state["cart"].append(listing["id"])
            _track_event("reserva", listing["id"], listing.get("title", ""))
            _notify(f"Reserva simulada creada para {listing['title']}", "success")
            st.success("Producto agregado a tu reserva. No se realizó ningún cobro.")
        b1, b2 = st.columns(2)
        fav = listing["id"] in st.session_state.get("favorites", [])
        if b1.button("Quitar favorito" if fav else "Guardar favorito", key=f"detail_fav_{listing['id']}", use_container_width=True):
            added = _toggle_item("favorites", listing["id"])
            if added: _track_event("favorito", listing["id"], listing.get("title", ""))
            st.rerun()
        if b2.button("Reportar publicación", key=f"detail_report_{listing['id']}", use_container_width=True):
            _track_event("reportar", listing["id"], listing.get("title", ""))
            st.info("Reporte simulado enviado al equipo de moderación.")

    st.markdown("<br>", unsafe_allow_html=True)
    s1, s2 = st.columns([1, 1.5], gap="large")
    with s1:
        _render_seller_profile(listing)
    with s2:
        st.markdown("#### Compra con tranquilidad")
        st.markdown(
            '<div class="safety-grid">'
            '<div class="safety-item"><strong>Estado transparente</strong><span>Condición, fotografías y detalles visibles antes de contactar.</span></div>'
            '<div class="safety-item"><strong>Perfiles con reputación</strong><span>Ventas, evaluación y verificación para tomar una decisión informada.</span></div>'
            '<div class="safety-item"><strong>Moderación comunitaria</strong><span>Reportes y reglas claras para reducir publicaciones engañosas.</span></div>'
            '</div>', unsafe_allow_html=True,
        )
        st.info("En la versión productiva se evaluará pago protegido y seguimiento de entrega. En este piloto no se procesa dinero.")

    st.markdown("---")
    _section_header("Más del Nexo", "Hallazgos relacionados", "Piezas del mismo universo o categoría que podrían interesarte.")
    all_items = [x for x in st.session_state.get("marketplace_db", []) if x.get("active", True) and x["id"] != listing["id"]]
    related = [x for x in all_items if x.get("game") == listing.get("game") or x.get("product_type") == listing.get("product_type")][:4]
    if related:
        cols = st.columns(len(related))
        for col, item in zip(cols, related):
            with col: _render_listing_card(item, "detail_related")
    else:
        _render_empty_state("✦", "Aún no hay piezas relacionadas", "El catálogo crecerá con las publicaciones de la comunidad.")


def _render_onboarding() -> None:
    if st.session_state.get("onboarding_complete"):
        return
    st.markdown(
        '<div class="welcome-deck"><span class="tour-badge">BIENVENIDA · PASO 1 DE 3</span>'
        '<h2>¿Qué quieres explorar primero?</h2><p>Elige una misión para probar la experiencia. '
        'Puedes recorrer las demás secciones cuando quieras.</p></div>', unsafe_allow_html=True,
    )
    c1, c2, c3 = st.columns(3)
    missions = [
        (c1, "Comprador", "Buscar y comparar", "Explora publicaciones, abre una ficha y simula una reserva.", "Marketplace"),
        (c2, "Vendedor", "Publicar una pieza", "Crea un anuncio y revisa cómo se administraría.", "Vender"),
        (c3, "Coleccionista", "Tasar mis cartas", "Identifica versiones y prepara un inventario para vender.", "Tasador"),
    ]
    for col, role, title, copy, page in missions:
        with col:
            st.markdown(f'<div class="role-card"><div class="role-icon">✦</div><h3>{title}</h3><p>{copy}</p></div>', unsafe_allow_html=True)
            if st.button(f"Entrar como {role.lower()}", key=f"onboard_{role}", use_container_width=True, type="primary" if role == "Comprador" else "secondary"):
                st.session_state["pilot_role"] = role
                st.session_state["onboarding_complete"] = True
                _track_event("onboarding", role, page)
                _go_to(page)
    if st.button("Explorar libremente", key="skip_onboarding"):
        st.session_state["pilot_role"] = "Explorador"
        st.session_state["onboarding_complete"] = True
        _track_event("onboarding", "Explorador", "sin guía")
        st.rerun()


def _render_sidebar(api_key_default: str = "") -> tuple[str | None, float, float]:
    # Valores técnicos disponibles sin ocupar espacio en la experiencia pública.
    api_key_input = st.session_state.get("api_key_guardada", api_key_default) or ""
    clp_rate = float(st.session_state.get("clp_rate_sidebar", 950))
    comision = float(st.session_state.get("commission_sidebar", 5.0))
    _CFG["max_workers"] = int(st.session_state.get("workers_sidebar", 4))
    _CFG["throttle"] = float(st.session_state.get("throttle_sidebar", 0.2))

    if not _DB_LOADED:
        for cp in ("card_data", os.path.join("pokemon-tcg-data", "cards", "en")):
            if os.path.isdir(cp) and cargar_base_local(cp):
                break

    with st.sidebar:
        st.markdown(
            "<div class='side-brand'><div class='side-logo'>NG</div>"
            "<div><strong>NexoGeek</strong><small>Tu portal de colección y juego</small></div></div>",
            unsafe_allow_html=True,
        )
        st.markdown('<span class="pilot-pill"><span class="live-dot"></span>Piloto activo</span>', unsafe_allow_html=True)
        st.caption("Compra, vende y descubre dentro de una experiencia demostrativa.")

        # Navegación accesible en móvil.
        nav_pages = ["Inicio", "Marketplace", "Vender", "Tasador", "Meta Lab", "Subastas", "Servicios", "Feedback"]
        current = st.session_state.get("page", "Inicio")
        current_nav = "Marketplace" if current == "Detalle" else current
        # Sincroniza el selector con botones y enlaces internos sin provocar rebotes.
        if st.session_state.get("_sidebar_nav_page") != current_nav:
            st.session_state["sidebar_public_nav"] = current_nav
            st.session_state["_sidebar_nav_page"] = current_nav
        selected_nav = st.selectbox(
            "Ir a", nav_pages,
            index=nav_pages.index(current_nav) if current_nav in nav_pages else 0,
            key="sidebar_public_nav",
        )
        if selected_nav != current_nav:
            st.session_state["page"] = selected_nav
            st.session_state["pending_nav"] = selected_nav
            st.session_state["_sidebar_nav_page"] = selected_nav
            st.rerun()

        with st.expander("Mi espacio", expanded=True):
            alias = st.text_input("Alias", value=st.session_state.get("pilot_alias", "Usuario_Piloto"), key="profile_alias_input_v3")
            location = st.selectbox(
                "Ubicación", UBICACIONES_DEMO,
                index=UBICACIONES_DEMO.index(st.session_state.get("pilot_location", "Santiago")) if st.session_state.get("pilot_location", "Santiago") in UBICACIONES_DEMO else 0,
                key="profile_location_input_v3",
            )
            st.session_state["pilot_alias"] = alias.strip() or "Usuario_Piloto"
            st.session_state["pilot_location"] = location
            role = st.session_state.get("pilot_role") or "Explorador"
            st.caption(f"Rol del recorrido: {role}")

        f1, f2, f3 = st.columns(3)
        f1.metric("Favoritos", len(st.session_state.get("favorites", [])))
        f2.metric("Reserva", len(st.session_state.get("cart", [])))
        f3.metric("Avisos", len(st.session_state.get("notifications", [])))

        completed, total, milestones = _tour_status()
        pct = int(completed / total * 100) if total else 0
        st.markdown(
            f'<div class="journey-card"><strong>Tu recorrido · {completed}/{total}</strong>'
            f'<div class="ng-progress"><span style="width:{pct}%"></span></div>'
            f'<small>{"Siguiente: " + next((name for name, done in milestones if not done), "Recorrido completado")}</small></div>',
            unsafe_allow_html=True,
        )

        with st.expander("Reserva", expanded=False):
            cart_ids = list(st.session_state.get("cart", []))
            if not cart_ids:
                st.caption("Tu reserva está vacía.")
            total_price = 0
            for lid in cart_ids:
                item = _listing_by_id(lid)
                if not item: continue
                total_price += int(item.get("price", 0))
                c1, c2 = st.columns([4, 1])
                c1.caption(f"{item.get('title')} · {_fmt_clp(item.get('price'))}")
                if c2.button("×", key=f"sidebar_remove_v3_{lid}"):
                    st.session_state["cart"].remove(lid); st.rerun()
            if cart_ids:
                st.markdown(f"**Total referencial: {_fmt_clp(total_price)}**")
                if st.button("Simular checkout", use_container_width=True, type="primary", key="sidebar_checkout_v3"):
                    _track_event("checkout", metadata=str(total_price))
                    st.success("Flujo completado sin cobro.")

        with st.expander("Actividad reciente", expanded=False):
            notifications = st.session_state.get("notifications", [])
            if not notifications:
                st.caption("Aún no hay actividad.")
            for note in notifications[:6]:
                st.caption(f"{note['time']} · {note['message']}")

        # El panel técnico queda bloqueado para que el público no vea opciones de desarrollo.
        with st.expander("Acceso anfitrión", expanded=False):
            if not st.session_state.get("admin_unlocked"):
                st.markdown('<div class="admin-lock">Configuración técnica y descargas del piloto. Solo para quien presenta.</div>', unsafe_allow_html=True)
                pin = st.text_input("PIN", type="password", key="admin_pin_input")
                if st.button("Desbloquear", key="admin_unlock", use_container_width=True):
                    if pin == ADMIN_PIN:
                        st.session_state["admin_unlocked"] = True
                        st.rerun()
                    else:
                        st.error("PIN incorrecto.")
            else:
                st.markdown('<span class="host-badge">MODO ANFITRIÓN</span>', unsafe_allow_html=True)
                api_key_input = st.text_input("PokémonTCG API Key", type="password", value=api_key_input, key="api_key_sidebar_v3")
                clp_rate = st.number_input("USD → CLP", min_value=0, value=int(clp_rate), step=10, key="clp_rate_sidebar_v3")
                comision = st.number_input("Comisión simulada %", min_value=0.0, max_value=50.0, value=float(comision), step=0.5, key="commission_sidebar_v3")
                workers = st.slider("Hilos", 1, 8, int(_CFG.get("max_workers",4)), key="workers_sidebar_v3")
                throttle = st.slider("Delay anti rate-limit", 0.0, 2.0, float(_CFG.get("throttle",.2)), .1, key="throttle_sidebar_v3")
                st.session_state["api_key_guardada"] = api_key_input
                st.session_state["clp_rate_sidebar"] = int(clp_rate)
                st.session_state["commission_sidebar"] = float(comision)
                st.session_state["workers_sidebar"] = int(workers)
                st.session_state["throttle_sidebar"] = float(throttle)
                _CFG["max_workers"] = int(workers); _CFG["throttle"] = float(throttle)
                st.caption(f"Fuente de cartas: {'DB local' if _DB_LOADED else 'API cloud'}")

                fb1 = leer_feedback(); fb2 = _read_csv_safe(EXTENDED_FEEDBACK_FILE)
                interactions = _read_csv_safe(INTERACTIONS_FILE); journey = _read_csv_safe(JOURNEY_FILE)
                if not fb1.empty: st.download_button("Feedback básico", fb1.to_csv(index=False).encode("utf-8-sig"), "feedback_nexogeek.csv", "text/csv", use_container_width=True)
                if not fb2.empty: st.download_button("Feedback completo", fb2.to_csv(index=False).encode("utf-8-sig"), "feedback_nexogeek_extendido.csv", "text/csv", use_container_width=True)
                if not interactions.empty: st.download_button("Reacciones", interactions.to_csv(index=False).encode("utf-8-sig"), "interacciones_nexogeek.csv", "text/csv", use_container_width=True)
                if not journey.empty: st.download_button("Recorrido de usuarios", journey.to_csv(index=False).encode("utf-8-sig"), "recorrido_nexogeek.csv", "text/csv", use_container_width=True)
                if st.button("Cerrar modo anfitrión", key="admin_lock", use_container_width=True):
                    st.session_state["admin_unlocked"] = False; st.rerun()

        if st.session_state.get("admin_unlocked"):
            if st.button("Reiniciar demo", use_container_width=True, key="reset_demo_v3"):
                for key in ["marketplace_db", "favorites", "cart", "compare", "notifications", "selected_listing", "deck_cart", "meta_selected_card", "meta_print_choices",
                            "auction_watchlist", "subastas_db", "servicios_db", "feature_votes", "df_result", "candidatos",
                            "onboarding_complete", "pilot_role", "_journey_seen", "_journey_actions"]:
                    st.session_state.pop(key, None)
                _init_demo_state(); st.rerun()

    return (api_key_input or None), float(clp_rate), float(comision)


def _render_top_navigation() -> str:
    options = [
        ("Inicio", "Inicio"), ("Marketplace", "Marketplace"), ("Vender", "Vender"),
        ("Tasador", "Tasador"), ("Subastas", "Subastas"), ("Servicios", "Servicios"), ("Feedback", "Feedback"),
    ]
    pending = st.session_state.pop("pending_nav", None)
    current = pending or st.session_state.get("page", "Inicio")
    if current not in {x[0] for x in options} | {"Detalle"}: current = "Inicio"
    active = "Marketplace" if current == "Detalle" else current

    nav_cols = st.columns([2.05, 1, 1.2, .92, .92, 1.03, 1.03, 1.03], gap="small")
    with nav_cols[0]:
        st.markdown(
            "<div class='nexo-brand'><div class='nexo-brand-mark'>✦</div>"
            "<div><div class='nexo-brand-name'>NexoGeek</div><div class='nexo-brand-tag'>colección · juego · comunidad</div></div></div>",
            unsafe_allow_html=True,
        )
    for column, (page, label) in zip(nav_cols[1:], options):
        with column:
            if st.button(label, key=f"top_nav_v3_{page.lower()}", use_container_width=True, type="primary" if active == page else "secondary"):
                if page != current:
                    _track_event("navegacion", page)
                    st.session_state["page"] = page; st.rerun()
    st.session_state["page"] = current
    st.markdown("<hr style='margin:9px 0 20px;border:none;border-top:2px solid rgba(40,22,58,.12);'>", unsafe_allow_html=True)
    return current


def render_home() -> None:
    _track_event("visita_inicio", once=True)
    _render_onboarding()
    st.markdown("""
    <div class="nexo-hero">
      <div class="nexo-hero-copy">
        <span class="nexo-eyebrow">TU PORTAL GEEK, HECHO EN COMUNIDAD</span>
        <h1>Donde tu próxima <em>obsesión</em> encuentra lugar.</h1>
        <p>Descubre cartas, juegos, mangas, figuras y creaciones únicas. Compra, publica, tasa y conecta en una experiencia pensada para coleccionistas y jugadores.</p>
        <div class="nexo-chip-row"><span>Cartas TCG</span><span>Productos sellados</span><span>Mangas</span><span>Figuras</span><span>Accesorios</span><span>Servicios creativos</span></div>
      </div>
    </div>
    <div class="universe-ribbon"><span>Pokémon</span><span>One Piece</span><span>Magic</span><span>Juegos de mesa</span><span>Manga</span><span>Figuras</span></div>
    """, unsafe_allow_html=True)

    cta1, cta2, cta3 = st.columns([1, 1, 2])
    if cta1.button("Explorar productos", type="primary", use_container_width=True, key="home_buy_v3"):
        _go_to("Marketplace")
    if cta2.button("Publicar algo", use_container_width=True, key="home_sell_v3"):
        _go_to("Vender")
    cta3.markdown('<div class="compact-note">Piloto funcional: reserva, chat y pagos son simulados; el tasador puede consultar referencias reales.</div>', unsafe_allow_html=True)

    listings = [x for x in st.session_state["marketplace_db"] if x.get("active", True)]
    hm1, hm2, hm3, hm4 = st.columns(4)
    hm1.metric("Piezas activas", len(listings)); hm2.metric("Universos", len(set(x["game"] for x in listings)))
    hm3.metric("Perfiles demo", len(set(x["seller"] for x in listings))); hm4.metric("Guardados", len(st.session_state.get("favorites", [])))

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header("Radar del Nexo", "Hallazgos que están llamando la atención", "Una selección para descubrir piezas, vendedores y categorías.")
    featured = sorted(listings, key=lambda x: (x.get("likes",0),x.get("views",0)), reverse=True)[:4]
    cols = st.columns(4)
    for col, item in zip(cols, featured):
        with col: _render_listing_card(item, "home_v3")

    st.markdown('<div class="trust-strip"><strong>Confianza antes que velocidad</strong><br>Estado visible · perfiles con reputación · moderación comunitaria · entrega coordinada · referencias de precio</div>', unsafe_allow_html=True)
    _section_header("Recorrido", "Una operación comprensible de principio a fin", "El piloto permite validar cada momento antes de construir pagos y logística reales.")
    steps = [("1","Descubre","Busca con filtros útiles."),("2","Evalúa","Revisa estado y reputación."),("3","Conversa","Aclara dudas sin exponer datos."),("4","Concreta","Reserva y coordina en un flujo protegido.")]
    cols = st.columns(4)
    for col,(n,title,copy) in zip(cols,steps):
        with col: st.markdown(f'<div class="soft-card"><div class="step-number">{n}</div><h4>{title}</h4><p>{copy}</p></div>', unsafe_allow_html=True)
    _render_feature_vote("inicio", "¿La propuesta se entiende claramente desde el inicio?")


def render_marketplace() -> None:
    _track_event("visita_marketplace", once=True)
    _section_header("Explora el Nexo", "Hallazgos de la comunidad", "Busca por universo, tipo de pieza, precio, estado y ubicación.")
    listings = [x for x in st.session_state["marketplace_db"] if x.get("active", True)]

    with st.expander("Buscar y filtrar", expanded=True):
        f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
        search = f1.text_input("Buscar", placeholder="Charizard, playmat, manga...", key="market_search_v3")
        games = ["Todos"] + sorted(set(x["game"] for x in listings)); game = f2.selectbox("Universo", games, key="market_game_v3")
        types = ["Todos"] + sorted(set(x["product_type"] for x in listings)); product_type = f3.selectbox("Tipo", types, key="market_type_v3")
        sort_by = f4.selectbox("Ordenar", ["Más relevantes","Precio menor","Precio mayor","Más vistos","Más guardados"], key="market_sort_v3")
        max_price = max([int(x.get("price",0)) for x in listings] + [1000])
        f5,f6,f7,f8 = st.columns(4)
        price_range = f5.slider("Precio",0,max_price,(0,max_price),step=max(500,max_price//100),key="market_price_v3")
        conditions = f6.multiselect("Estado",sorted(set(x["condition"] for x in listings)),key="market_conditions_v3")
        location = f7.selectbox("Ubicación",["Todas"]+sorted(set(x["location"] for x in listings)),key="market_location_v3")
        delivery = f8.selectbox("Entrega",["Todas","Envío y retiro","Solo envío","Solo retiro"],key="market_delivery_v3")
        q1,q2,q3 = st.columns(3)
        verified_only=q1.checkbox("Vendedores verificados",key="market_verified_v3")
        favorites_only=q2.checkbox("Solo favoritos",key="market_favorites_v3")
        negotiable_only=q3.checkbox("Precio conversable",key="market_negotiable_v3")

    query=search.strip().lower(); filtered=[]
    for item in listings:
        haystack=" ".join([item.get("title",""),item.get("game",""),item.get("product_type","")," ".join(item.get("tags",[]))]).lower()
        if query and query not in haystack: continue
        if game!="Todos" and item["game"]!=game: continue
        if product_type!="Todos" and item["product_type"]!=product_type: continue
        if not(price_range[0]<=int(item.get("price",0))<=price_range[1]): continue
        if conditions and item.get("condition") not in conditions: continue
        if location!="Todas" and item.get("location")!=location: continue
        if delivery!="Todas" and item.get("shipping")!=delivery: continue
        if verified_only and not item.get("verified"): continue
        if favorites_only and item["id"] not in st.session_state.get("favorites",[]): continue
        if negotiable_only and not item.get("negotiable"): continue
        filtered.append(item)
    sorters={"Precio menor":lambda x:x.get("price",0),"Precio mayor":lambda x:-x.get("price",0),"Más vistos":lambda x:-x.get("views",0),"Más guardados":lambda x:-x.get("likes",0),"Más relevantes":lambda x:-(x.get("likes",0)*3+x.get("views",0))}
    filtered=sorted(filtered,key=sorters[sort_by])
    if query: _track_event("buscar", query, once=True)

    active=[]
    if game!="Todos": active.append(game)
    if product_type!="Todos": active.append(product_type)
    if conditions: active.extend(conditions)
    if location!="Todas": active.append(location)
    if verified_only: active.append("Verificados")
    if favorites_only: active.append("Favoritos")
    if negotiable_only: active.append("Conversables")
    if active: st.markdown(" ".join(f'<span class="active-filter">{_safe_text(x)}</span>' for x in active), unsafe_allow_html=True)
    st.caption(f"{len(filtered)} resultado(s) · publicaciones demostrativas")

    compare_ids=st.session_state.get("compare",[])
    if compare_ids:
        with st.expander(f"Comparador ({len(compare_ids)}/3)", expanded=True):
            comp=[x for x in listings if x["id"] in compare_ids]
            if comp:
                df_cmp=pd.DataFrame([{"Producto":x["title"],"Precio":x["price"],"Estado":x["condition"],"Vendedor":x["seller"],"Rating":x["rating"],"Entrega":x["shipping"]} for x in comp])
                st.dataframe(df_cmp,use_container_width=True,hide_index=True)
                if st.button("Limpiar comparación",key="clear_compare_v3"): st.session_state["compare"]=[];st.rerun()

    if not filtered:
        _render_empty_state("🔭","No encontramos coincidencias","Prueba eliminando un filtro o ampliando el rango de precio.")
    else:
        for start in range(0,len(filtered),4):
            cols=st.columns(4)
            for col,item in zip(cols,filtered[start:start+4]):
                with col:_render_listing_card(item,"market_v3")

    if st.session_state.get("favorites"):
        st.markdown("---");_section_header("Para ti","Nuevas rutas para explorar","Sugerencias basadas en los universos que guardaste.")
        fav_games={x["game"] for x in listings if x["id"] in st.session_state["favorites"]}
        recs=[x for x in listings if x["id"] not in st.session_state["favorites"] and x["game"] in fav_games][:4]
        if recs:
            cols=st.columns(len(recs))
            for col,item in zip(cols,recs):
                with col:_render_listing_card(item,"reco_v3")
        else:_render_empty_state("✦","Ya revisaste lo más cercano","Pronto aparecerán más recomendaciones para tu colección.")
    _render_feature_vote("marketplace","¿Usarías este catálogo para buscar productos geek o TCG?")


def render_sell() -> None:
    _track_event("visita_vender", once=True)
    _section_header("Vender", "Publica sin perderte en formularios", "Elige una pieza, describe su estado y define cómo quieres entregarla.")
    st.markdown('<div class="compact-note">Recorrido sugerido: 1. Describe la pieza · 2. Define precio y entrega · 3. Revisa cómo se verá en el marketplace.</div>', unsafe_allow_html=True)
    own=[x for x in st.session_state["marketplace_db"] if x.get("owner")]
    sm1,sm2,sm3,sm4=st.columns(4)
    sm1.metric("Publicaciones",len(own));sm2.metric("Vistas",sum(x.get("views",0) for x in own));sm3.metric("Guardados",sum(x.get("likes",0) for x in own));sm4.metric("Valor publicado",_fmt_clp(sum(x.get("price",0)*x.get("stock",1) for x in own)))
    source_tab,manual_tab,manage_tab=st.tabs(["Desde el tasador","Nueva publicación","Mis publicaciones"])
    with source_tab:
        if "df_result" not in st.session_state or st.session_state["df_result"].empty:
            _render_empty_state("⚡","Aún no tienes cartas tasadas","Usa el tasador para identificar cartas y publicarlas con menos pasos.")
            if st.button("Ir al tasador",type="primary",key="sell_go_tasador_v3"):_go_to("Tasador")
        else:
            st.success("Tu inventario tasado está listo para convertirse en publicaciones.")
            render_dashboard(st.session_state["df_result"],st.session_state.get("clp_rate_sidebar",950),st.session_state.get("commission_sidebar",5.0))
    with manual_tab:
        with st.form("manual_listing_form_v3",clear_on_submit=True):
            a1,a2=st.columns([2,1]);title=a1.text_input("Título *",placeholder="Ej. Booster Box sellada edición japonesa");game=a2.selectbox("Universo",["Pokémon TCG","One Piece Card Game","Magic: The Gathering","Juegos de mesa","Accesorios","Figuras y animé","Mangas y cómics","Otro"])
            b1,b2,b3=st.columns(3);product_type=b1.selectbox("Tipo",["Carta individual","Producto sellado","Mazo armado","Accesorio","Figura","Manga","Juego de mesa","Otro"]);condition=b2.selectbox("Estado",["Nuevo","Sellado","NM","LP","MP","HP","Usado"]);stock=b3.number_input("Cantidad",1,100,1)
            c1,c2,c3=st.columns(3);price=c1.number_input("Precio CLP *",min_value=0,value=19990,step=500);location=c2.selectbox("Ubicación",UBICACIONES_DEMO);shipping=c3.selectbox("Entrega",["Envío y retiro","Solo envío","Solo retiro"])
            image=st.text_input("URL de imagen (opcional)",placeholder="https://...")
            description=st.text_area("Descripción",placeholder="Estado, accesorios incluidos y condiciones de entrega.")
            negotiable=st.checkbox("Precio conversable");submit=st.form_submit_button("Publicar en la demo",type="primary",use_container_width=True)
        if submit:
            if not title.strip() or price<=0:st.error("Completa título y precio.")
            else:
                listing={"id":_new_id("manual"),"title":title.strip(),"game":game,"product_type":product_type,"condition":condition,"price":int(price),"location":location,"shipping":shipping,"seller":st.session_state.get("pilot_alias","Usuario_Piloto"),"verified":False,"rating":5.0,"sales":0,"image":image.strip(),"stock":int(stock),"negotiable":negotiable,"description":description.strip() or "Publicación creada durante la prueba piloto.","tags":["Recién publicado",product_type],"views":0,"likes":0,"active":True,"owner":True,"response_time":"Perfil nuevo · respuesta por confirmar","member_since":"Se unió durante el piloto","photo_count":1,"protected":True,"authenticity":"Información declarada por el usuario piloto"}
                st.session_state["marketplace_db"].insert(0,listing);st.session_state["selected_listing"]=listing["id"]
                _track_event("publicar",listing["id"],listing["title"]);_notify(f"Publicaste {listing['title']}","success")
                st.success("Publicación creada. Puedes abrir su ficha desde Mis publicaciones.")
    with manage_tab:
        own=[x for x in st.session_state["marketplace_db"] if x.get("owner")]
        if not own:_render_empty_state("📦","Todavía no publicas nada","Crea tu primera publicación manual o usa una carta tasada.")
        for item in own:
            with st.container(border=True):
                c1,c2,c3,c4=st.columns([3,1,1,1]);c1.markdown(f"**{item['title']}**  \n{_fmt_clp(item['price'])} · stock {item['stock']} · {'Activa' if item.get('active') else 'Pausada'}")
                if c2.button("Ver ficha",key=f"manage_view_v3_{item['id']}",use_container_width=True):st.session_state["selected_listing"]=item["id"];_go_to("Detalle")
                if c3.button("Pausar" if item.get("active") else "Activar",key=f"manage_toggle_v3_{item['id']}",use_container_width=True):item["active"]=not item.get("active",True);st.rerun()
                if c4.button("Eliminar",key=f"manage_delete_v3_{item['id']}",use_container_width=True):
                    st.session_state["marketplace_db"]=[x for x in st.session_state["marketplace_db"] if x["id"]!=item["id"]];st.rerun()
    _render_feature_vote("vender","¿El flujo para publicar te parece suficientemente simple?")


def render_catalogador(api_key: str | None, clp_rate: float, comision: float) -> None:
    _track_event("visita_tasador", once=True)
    st.markdown('<div class="compact-note">El tasador es la parte más funcional del piloto: identifica versiones, calcula referencias y permite convertir resultados en publicaciones.</div>', unsafe_allow_html=True)
    _legacy_render_catalogador_v3(api_key, clp_rate, comision)
    if "df_result" in st.session_state and not st.session_state["df_result"].empty:
        _track_event("tasacion_completada", str(len(st.session_state["df_result"])), once=True)


def render_auctions() -> None:
    _track_event("visita_subastas", once=True)
    _legacy_render_auctions_v3()


def render_services() -> None:
    _track_event("visita_servicios", once=True)
    _legacy_render_services_v3()


def render_feedback() -> None:
    _track_event("visita_feedback", once=True)
    _section_header("Validación", "Tu experiencia define la siguiente versión", "Cuéntanos qué genera valor, qué causa dudas y qué deberíamos construir primero.")
    completed,total,milestones=_tour_status()
    st.markdown(f'<div class="compact-note">Has completado {completed} de {total} momentos del recorrido sugerido. Puedes responder ahora o seguir explorando.</div>', unsafe_allow_html=True)
    with st.form("extended_feedback_form_v3",clear_on_submit=True):
        f1,f2=st.columns(2);name=f1.text_input("Nombre o alias (opcional)");profile=f2.selectbox("Perfil",["Coleccionista","Jugador competitivo","Vendedor o tienda","Organizador","Nuevo en el hobby","Otro"])
        n1,n2=st.columns(2);score=n1.slider("Nota general",1,10,8);nps=n2.slider("Probabilidad de recomendarla",0,10,8)
        favorite_module=st.selectbox("Función más valiosa",["Marketplace","Ficha de producto","Tasador","Meta Lab","Publicación","Subastas","Servicios","Perfiles y reputación","Otra"])
        priorities=st.multiselect("¿Qué construir primero?",["Pagos protegidos","Sistema de envíos","Chat interno","Verificación de vendedores","Meta competitivo y torneos","Aplicación móvil","Alertas y favoritos","Más juegos TCG","Integración con tiendas","Subastas reales","Servicios comunitarios"],max_selections=3)
        trust=st.selectbox("¿Qué te daría más confianza?",["Pago retenido hasta confirmar entrega","Vendedores verificados","Reputación y comentarios","Fotos obligatorias","Moderación y reportes","Punto de encuentro seguro","Otro"])
        mobile=st.selectbox("¿Cómo se sintió en tu dispositivo?",["Muy cómoda","Cómoda","Regular","Difícil de usar","No la probé en móvil"])
        willingness=st.selectbox("Comisión aceptable para una venta protegida",["No pagaría comisión","1% a 3%","4% a 6%","7% a 10%","Depende del servicio"])
        use_intent=st.radio("¿La usarías?",["Sí","Tal vez","No"],horizontal=True)
        useful=st.text_area("¿Qué fue lo mejor?");improve=st.text_area("¿Qué cambiarías o agregarías?");contact=st.text_input("Contacto para futuras pruebas (opcional)")
        submit=st.form_submit_button("Enviar feedback",type="primary",use_container_width=True)
    if submit:
        if not useful.strip() and not improve.strip():st.warning("Escribe al menos una observación.")
        else:
            data={"nombre":name.strip(),"perfil":profile,"nota":score,"nps":nps,"modulo_favorito":favorite_module,"prioridades":" | ".join(priorities),"factor_confianza":trust,"experiencia_dispositivo":mobile,"comision_aceptable":willingness,"usaria":use_intent,"lo_mejor":useful.strip(),"mejoraria":improve.strip(),"contacto":contact.strip(),"recorrido_completado":f"{completed}/{total}"}
            try:_save_extended_feedback(data);_track_event("enviar_feedback",metadata=str(score));st.success("Gracias. Tu respuesta quedó guardada.")
            except Exception as exc:st.error(f"No se pudo guardar: {exc}")

    feedback=_read_csv_safe(EXTENDED_FEEDBACK_FILE);journey=_read_csv_safe(JOURNEY_FILE)
    if not feedback.empty or not journey.empty:
        st.markdown("---");_section_header("Pulso del piloto","Lo que está ocurriendo durante las pruebas","Resultados locales y recorrido agregado de las personas participantes.")
    if not feedback.empty:
        r1,r2,r3,r4=st.columns(4);r1.metric("Respuestas",len(feedback));r2.metric("Nota promedio",f"{pd.to_numeric(feedback['nota'],errors='coerce').mean():.1f}/10");r3.metric("Recomendación",f"{pd.to_numeric(feedback['nps'],errors='coerce').mean():.1f}/10");r4.metric("La usaría",f"{(feedback['usaria'].astype(str).str.lower()=='sí').mean()*100:.0f}%")
        c1,c2=st.columns(2)
        with c1:st.markdown("#### Funciones más valoradas");st.bar_chart(feedback["modulo_favorito"].value_counts())
        with c2:st.markdown("#### Perfiles participantes");st.bar_chart(feedback["perfil"].value_counts())
    if not journey.empty and "evento" in journey.columns:
        st.markdown("#### Recorrido real dentro del piloto")
        funnel_order=["visita_marketplace","abrir_ficha","favorito","comparar","reserva","visita_tasador","publicar","enviar_feedback"]
        labels={"visita_marketplace":"Visitó marketplace","abrir_ficha":"Abrió ficha","favorito":"Guardó favorito","comparar":"Comparó","reserva":"Reservó","visita_tasador":"Visitó tasador","publicar":"Publicó","enviar_feedback":"Envió feedback"}
        counts=[]
        for ev in funnel_order:
            rows=journey[journey["evento"]==ev]
            count=rows["session_id"].nunique() if "session_id" in rows.columns else len(rows)
            counts.append({"Momento":labels[ev],"Personas / sesiones":int(count)})
        st.dataframe(pd.DataFrame(counts),use_container_width=True,hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# META LAB BETA — COMPETITIVO, TORNEOS Y CONEXIÓN CON COLECCIÓN/MARKETPLACE
# ══════════════════════════════════════════════════════════════════════════════
META_LAB_CSS = """
<style>
.meta-hero {
    position:relative; overflow:hidden; border:2px solid var(--ng-ink);
    border-radius:28px 28px 10px 28px; padding:28px;
    background:
      radial-gradient(circle at 86% 18%, rgba(255,200,87,.34), transparent 13rem),
      radial-gradient(circle at 73% 94%, rgba(46,196,182,.28), transparent 15rem),
      linear-gradient(135deg,#341544 0%,#5B2A86 55%,#7847A5 100%);
    color:#FFF8ED; box-shadow:8px 8px 0 var(--ng-sun); margin-bottom:22px;
}
.meta-hero:after { content:""; position:absolute; width:190px; height:190px; right:-35px; top:-70px;
    border:24px solid rgba(255,255,255,.08); border-radius:50%; transform:rotate(18deg); }
.meta-hero h1 { color:#FFF8ED !important; font-size:2.45rem; margin:7px 0 8px; letter-spacing:-.055em; }
.meta-hero p { color:#E9DDF2 !important; max-width:790px; margin:0; line-height:1.55; }
.meta-beta { display:inline-block; padding:5px 10px; border-radius:999px; color:var(--ng-ink);
    background:var(--ng-sun); border:2px solid var(--ng-ink); font-size:.7rem; font-weight:950; letter-spacing:.08em; }
.meta-source { display:inline-flex; align-items:center; gap:7px; background:#FFFDF8; border:2px solid var(--ng-ink);
    border-radius:999px; padding:6px 11px; font-size:.72rem; font-weight:850; color:var(--ng-ink); }
.meta-source i { width:8px; height:8px; border-radius:50%; background:var(--ng-mint); display:inline-block; }
.meta-rank-card { background:#FFFDF8; border:2px solid var(--ng-ink); border-radius:18px 18px 6px 18px;
    padding:16px; box-shadow:4px 4px 0 var(--ng-coral); min-height:180px; }
.meta-rank { width:34px; height:34px; display:flex; align-items:center; justify-content:center;
    border:2px solid var(--ng-ink); border-radius:11px 11px 4px 11px; background:var(--ng-sun); font-weight:950; }
.meta-rank-card h4 { margin:10px 0 5px; color:var(--ng-ink); font-size:1rem; line-height:1.25; }
.meta-rank-card p { margin:0; color:var(--ng-muted); font-size:.78rem; }
.meta-tier { display:inline-block; padding:3px 8px; border-radius:999px; border:1px solid var(--ng-ink);
    font-size:.65rem; font-weight:900; background:#F3E9FF; }
.meta-tier-s { background:#FFE8A8; } .meta-tier-a { background:#CFF7EF; } .meta-tier-b { background:#F2E8FF; }
.meta-stat-line { display:flex; justify-content:space-between; gap:8px; margin-top:11px; padding-top:9px;
    border-top:1px dashed #CDBED8; color:var(--ng-ink); font-size:.72rem; }
.meta-event { border:2px solid var(--ng-ink); border-radius:20px 20px 7px 20px; padding:18px;
    background:linear-gradient(145deg,#FFFDF8,#FFF4DF); box-shadow:5px 5px 0 var(--ng-mint); height:100%; }
.meta-event h4 { margin:6px 0; color:var(--ng-ink); }
.meta-event p { color:var(--ng-muted); font-size:.8rem; margin:3px 0; }
.meta-winner { background:var(--ng-ink); color:#FFF8ED; border-radius:12px 12px 4px 12px; padding:10px 12px; margin-top:10px; }
.meta-winner strong { color:var(--ng-sun); }
.meta-progress { height:13px; border:2px solid var(--ng-ink); background:#FFF8ED; border-radius:999px; overflow:hidden; }
.meta-progress span { display:block; height:100%; background:linear-gradient(90deg,var(--ng-mint),var(--ng-violet)); }
.meta-match-good { color:#067A69; font-weight:900; } .meta-match-bad { color:#C23C52; font-weight:900; }
.meta-deck-head { padding:20px; border:2px solid var(--ng-ink); border-radius:22px 22px 7px 22px;
    background:linear-gradient(135deg,#FFFDF8,#F3E9FF); box-shadow:6px 6px 0 var(--ng-sun); }
.meta-deck-head h2 { margin:0 0 5px; color:var(--ng-ink); }
.meta-card-pill { display:inline-block; margin:3px 4px 3px 0; padding:4px 8px; border:1px solid #A98FBB;
    background:#FFFDF8; border-radius:999px; font-size:.68rem; color:var(--ng-ink); }
.meta-admin { border:2px dashed var(--ng-violet); border-radius:18px; padding:16px; background:rgba(243,233,255,.55); }
@media(max-width:760px){
    .meta-hero{padding:20px;box-shadow:5px 5px 0 var(--ng-sun)} .meta-hero h1{font-size:1.8rem}
    .meta-rank-card{min-height:auto} .meta-event{height:auto}
}
</style>
"""


META_DECK_BUILDER_CSS = """
<style>
.deck-builder-shell {
    border:2px solid var(--ng-ink); border-radius:24px 24px 8px 24px;
    background:linear-gradient(145deg,#FFFDF8,#FFF7E8); padding:18px;
    box-shadow:6px 6px 0 var(--ng-mint); margin:12px 0 18px;
}
.deck-builder-title {display:flex;align-items:flex-start;justify-content:space-between;gap:14px;flex-wrap:wrap}
.deck-builder-title h3 {margin:0;color:var(--ng-ink)}
.deck-builder-title p {margin:4px 0 0;color:var(--ng-muted);font-size:.82rem}
.deck-progress-large {height:18px;border:2px solid var(--ng-ink);background:#FFF;border-radius:999px;overflow:hidden;margin-top:12px}
.deck-progress-large span {display:block;height:100%;background:linear-gradient(90deg,var(--ng-mint),var(--ng-sun),var(--ng-coral));transition:width .25s ease}
.deck-card-art {position:relative;border:2px solid var(--ng-ink);border-radius:17px 17px 6px 17px;background:#F5EDF9;overflow:hidden;min-height:248px;display:flex;align-items:center;justify-content:center}
.deck-card-art img {display:block;width:100%;height:270px;object-fit:contain;background:radial-gradient(circle at 50% 35%,#FFF 0%,#F5EDF9 75%);padding:8px;transition:transform .18s ease}.deck-card-click{display:block;width:100%;text-decoration:none}.deck-card-click:hover img{transform:scale(1.025)}
.deck-card-placeholder {height:270px;width:100%;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:18px;text-align:center;background:linear-gradient(145deg,#F4E8FF,#FFF3D2)}
.deck-card-placeholder b {font-size:.9rem;color:var(--ng-ink);margin-top:8px}.deck-card-placeholder span {font-size:2.2rem}
.deck-qty-badge {position:absolute;right:9px;top:9px;min-width:38px;height:38px;padding:0 9px;border:2px solid var(--ng-ink);border-radius:13px 13px 4px 13px;background:var(--ng-coral);color:#FFF;font-size:1rem;font-weight:950;display:flex;align-items:center;justify-content:center;box-shadow:3px 3px 0 var(--ng-sun)}
.deck-owned-badge {position:absolute;left:9px;top:9px;padding:5px 8px;border:2px solid var(--ng-ink);border-radius:999px;background:#FFFDF8;color:var(--ng-ink);font-size:.64rem;font-weight:950}
.deck-card-meta {padding:10px 2px 2px}.deck-card-meta h4 {margin:0 0 4px;color:var(--ng-ink);font-size:.92rem;line-height:1.25;min-height:2.3em}.deck-card-meta p {margin:0;color:var(--ng-muted);font-size:.72rem}
.deck-status {display:inline-flex;align-items:center;gap:5px;margin-top:7px;padding:4px 8px;border:1px solid var(--ng-ink);border-radius:999px;font-size:.64rem;font-weight:900}
.deck-status.complete {background:#D5F6EA}.deck-status.partial {background:#FFF0B8}.deck-status.missing {background:#FFE0DD}.deck-status.cart {background:#E8D8FF}
.deck-detail-panel {border:2px solid var(--ng-ink);border-radius:22px 22px 7px 22px;background:linear-gradient(135deg,#FFFDF8,#F2E7FF);padding:18px;box-shadow:5px 5px 0 var(--ng-sun);margin:8px 0 18px}
.deck-detail-image {border:2px solid var(--ng-ink);border-radius:16px;background:#FFF;overflow:hidden;padding:8px;text-align:center}.deck-detail-image img {max-height:390px;max-width:100%;object-fit:contain}
.deck-cart-box {border:2px solid var(--ng-ink);border-radius:20px 20px 6px 20px;background:#2A1639;color:#FFF8ED;padding:16px;box-shadow:5px 5px 0 var(--ng-coral)}
.deck-cart-box h4 {color:#FFF8ED;margin:0 0 8px}.deck-cart-box p {color:#E8DDF0;font-size:.76rem;margin:3px 0}.deck-cart-total {font-size:1.45rem;font-weight:950;color:var(--ng-sun)}
.deck-card-meta + div[data-testid="stHorizontalBlock"] {margin-top:4px}
.deck-builder-shell .deck-availability {flex-wrap:wrap;gap:7px 12px}
.deck-cart-line {border-top:1px dashed rgba(255,255,255,.28);padding-top:8px;margin-top:8px}
.deck-mini-note {padding:10px 12px;border:2px dashed #A98FBB;border-radius:14px;background:#FFFDF8;color:var(--ng-muted);font-size:.75rem}
.deck-availability {display:flex;gap:6px;flex-wrap:wrap;margin-top:8px}.deck-availability span {padding:4px 7px;border-radius:999px;border:1px solid var(--ng-ink);background:#FFF;font-size:.64rem;font-weight:850;color:var(--ng-ink)}
@media(max-width:760px){
  .deck-card-art,.deck-card-art img,.deck-card-placeholder{min-height:210px;height:230px}.deck-builder-shell{padding:13px}.deck-detail-panel{padding:13px}.deck-qty-badge{min-width:34px;height:34px}
}
</style>
"""


def _meta_core(cards: list[tuple[int, str, str]], deck_size: int = 60, filler: str = "Otras cartas de soporte") -> list[dict]:
    """Completa una lista representativa hasta el tamaño reglamentario del mazo."""
    rows = [{"cantidad": int(q), "carta": str(name), "categoria": str(cat)} for q, name, cat in cards]
    total = sum(x["cantidad"] for x in rows)
    if total < deck_size:
        rows.append({"cantidad": deck_size - total, "carta": filler, "categoria": "Soporte"})
    return rows


META_DEMO = {
    "Pokémon TCG": {
        "formatos": ["Standard"],
        "actualizado": "Snapshot competitivo · NAIC 2026 · 12–14 junio 2026 · 3.752 jugadores",
        "fuente": "Limitless Labs · datos reales de referencia",
        "source_url": "https://labs.limitlesstcg.com/0070/decks",
        "legalidad": "Standard 2026 · regulación H, I y J · formato TEF–CRI",
        "conversion_label": "Conversión Día 2",
        "tipo_datos": "snapshot_real",
        "mazos": [
            {
                "id":"pkm-dragapult", "nombre":"Dragapult", "tier":"S",
                "uso":20.01, "win_rate":53.49, "top_cut":32.04, "entradas":749,
                "tendencia":0, "costo":41900, "mejor":"3.º NAIC 2026",
                "descripcion":"La referencia central del formato: presión de daño, interrupción y múltiples variantes competitivas.",
                "lista_fuente":"Justin Newdorf · 3.º NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Dreepy","Pokémon"),(4,"Drakloak","Pokémon"),(3,"Dragapult ex","Pokémon"),
                    (2,"Munkidori","Pokémon"),(1,"Dunsparce","Pokémon"),(1,"Dudunsparce","Pokémon"),
                    (1,"Budew","Pokémon"),(1,"Fezandipiti ex","Pokémon"),(1,"Meowth ex","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(3,"Boss's Orders","Entrenador"),
                    (2,"Crispin","Entrenador"),(1,"Rosa's Encouragement","Entrenador"),
                    (4,"Buddy-Buddy Poffin","Objeto"),(4,"Poké Pad","Objeto"),(4,"Ultra Ball","Objeto"),
                    (4,"Crushing Hammer","Objeto"),(3,"Night Stretcher","Objeto"),
                    (1,"Special Red Card","Objeto"),(1,"Unfair Stamp","Objeto"),(2,"Risky Ruins","Estadio"),
                    (4,"Psychic Energy","Energía"),(3,"Fire Energy","Energía"),(2,"Darkness Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-dragapult-dusknoir", "nombre":"Dragapult Dusknoir", "tier":"A",
                "uso":9.99, "win_rate":48.29, "top_cut":15.51, "entradas":374,
                "tendencia":0, "costo":37900, "mejor":"2.º NAIC 2026",
                "descripcion":"Variante que combina el daño distribuido de Dragapult con cierres explosivos mediante Dusknoir.",
                "lista_fuente":"Neddy Kosek · 2.º NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Dreepy","Pokémon"),(4,"Drakloak","Pokémon"),(2,"Dragapult ex","Pokémon"),
                    (2,"Duskull","Pokémon"),(2,"Dusclops","Pokémon"),(1,"Dusknoir","Pokémon"),
                    (1,"Budew","Pokémon"),(1,"Fezandipiti ex","Pokémon"),(1,"Meowth ex","Pokémon"),
                    (1,"Munkidori","Pokémon"),(4,"Lillie's Determination","Entrenador"),
                    (3,"Crispin","Entrenador"),(2,"Boss's Orders","Entrenador"),(1,"Dawn","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Poké Pad","Objeto"),(4,"Buddy-Buddy Poffin","Objeto"),
                    (4,"Crushing Hammer","Objeto"),(2,"Night Stretcher","Objeto"),(1,"Unfair Stamp","Objeto"),
                    (1,"Special Red Card","Objeto"),(1,"Handheld Fan","Herramienta"),
                    (1,"Team Rocket's Watchtower","Estadio"),(1,"Jamming Tower","Estadio"),
                    (3,"Psychic Energy","Energía"),(3,"Fire Energy","Energía"),(2,"Darkness Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-ogerpon-hydrapple", "nombre":"Ogerpon Meganium Hydrapple", "tier":"A",
                "uso":6.20, "win_rate":45.48, "top_cut":13.36, "entradas":232,
                "tendencia":0, "costo":66500, "mejor":"25.º NAIC 2026",
                "descripcion":"Motor de Energía Planta con Ogerpon, Meganium y Hydrapple para sostener atacantes de alto impacto.",
                "lista_fuente":"Arquetipo NAIC 2026 · núcleo representativo legal H/I/J",
                "core":_meta_core([
                    (4,"Teal Mask Ogerpon ex","Pokémon"),(3,"Chikorita","Pokémon"),(2,"Bayleef","Pokémon"),
                    (2,"Meganium","Pokémon"),(3,"Applin","Pokémon"),(2,"Dipplin","Pokémon"),
                    (2,"Hydrapple ex","Pokémon"),(2,"Mega Kangaskhan ex","Pokémon"),(1,"Fezandipiti ex","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(3,"Crispin","Entrenador"),(2,"Boss's Orders","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Poké Pad","Objeto"),(3,"Wondrous Patch","Objeto"),
                    (2,"Night Stretcher","Objeto"),(2,"Energy Retrieval","Objeto"),(3,"Area Zero Underdepths","Estadio"),
                    (12,"Grass Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-alakazam", "nombre":"Alakazam Dudunsparce", "tier":"A",
                "uso":6.14, "win_rate":48.78, "top_cut":20.00, "entradas":230,
                "tendencia":0, "costo":36900, "mejor":"1.º Regional Indianapolis 2026",
                "descripcion":"Mazo de mano amplia y consistencia que aprovecha Powerful Hand junto al motor Dudunsparce.",
                "lista_fuente":"Cerys Jones · campeón Regional Indianapolis 2026 · núcleo representativo",
                "core":_meta_core([
                    (4,"Abra","Pokémon"),(4,"Kadabra","Pokémon"),(3,"Alakazam","Pokémon"),
                    (3,"Dunsparce","Pokémon"),(3,"Dudunsparce","Pokémon"),(1,"Dudunsparce ex","Pokémon"),
                    (1,"Fezandipiti ex","Pokémon"),(1,"Meowth ex","Pokémon"),(4,"Lillie's Determination","Entrenador"),
                    (3,"Boss's Orders","Entrenador"),(2,"Ciphermaniac's Codebreaking","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Poké Pad","Objeto"),(3,"Buddy-Buddy Poffin","Objeto"),
                    (2,"Night Stretcher","Objeto"),(2,"Rare Candy","Objeto"),(1,"Unfair Stamp","Objeto"),
                    (3,"Academy at Night","Estadio"),(8,"Psychic Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-zoroark", "nombre":"N's Zoroark", "tier":"A",
                "uso":5.58, "win_rate":50.30, "top_cut":24.40, "entradas":209,
                "tendencia":0, "costo":36500, "mejor":"10.º NAIC 2026",
                "descripcion":"Caja de atacantes de N con gran flexibilidad, aceleración propia y acceso a respuestas situacionales.",
                "lista_fuente":"Tord Reklev · 10.º NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"N's Zorua","Pokémon"),(4,"N's Zoroark ex","Pokémon"),(2,"N's Zekrom","Pokémon"),
                    (1,"N's Darumaka","Pokémon"),(1,"N's Darmanitan","Pokémon"),(1,"Tatsugiri","Pokémon"),
                    (1,"Budew","Pokémon"),(1,"Yveltal","Pokémon"),(1,"Munkidori","Pokémon"),
                    (1,"Pecharunt ex","Pokémon"),(1,"Fezandipiti ex","Pokémon"),(1,"Meowth ex","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(3,"Boss's Orders","Entrenador"),
                    (2,"Cyrano","Entrenador"),(1,"Black Belt's Training","Entrenador"),(1,"Ruffian","Entrenador"),
                    (4,"Buddy-Buddy Poffin","Objeto"),(4,"Transformation Tome","Objeto"),(3,"Ultra Ball","Objeto"),
                    (3,"N's PP Up","Objeto"),(2,"Night Stretcher","Objeto"),(1,"Poké Pad","Objeto"),
                    (1,"Special Red Card","Objeto"),(1,"Secret Box","Objeto"),(2,"Binding Mochi","Herramienta"),
                    (2,"N's Castle","Estadio"),(7,"Darkness Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-dragapult-blaziken", "nombre":"Dragapult Blaziken", "tier":"A",
                "uso":4.97, "win_rate":46.98, "top_cut":12.37, "entradas":186,
                "tendencia":0, "costo":42400, "mejor":"6.º NAIC 2026",
                "descripcion":"Variante de Dragapult que suma aceleración y presión adicional con Blaziken ex.",
                "lista_fuente":"Jon Webb · 6.º NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Dreepy","Pokémon"),(4,"Drakloak","Pokémon"),(2,"Dragapult ex","Pokémon"),
                    (2,"Torchic","Pokémon"),(1,"Combusken","Pokémon"),(2,"Blaziken ex","Pokémon"),
                    (2,"Munkidori","Pokémon"),(1,"Lillie's Clefairy ex","Pokémon"),(1,"Fezandipiti ex","Pokémon"),
                    (1,"Meowth ex","Pokémon"),(1,"Budew","Pokémon"),(1,"Chi-Yu","Pokémon"),(1,"Shaymin","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(3,"Boss's Orders","Entrenador"),
                    (2,"Crispin","Entrenador"),(1,"Dawn","Entrenador"),(4,"Buddy-Buddy Poffin","Objeto"),
                    (4,"Ultra Ball","Objeto"),(3,"Poké Pad","Objeto"),(2,"Night Stretcher","Objeto"),
                    (2,"Rare Candy","Objeto"),(1,"Special Red Card","Objeto"),(1,"Unfair Stamp","Objeto"),
                    (1,"Area Zero Underdepths","Estadio"),(1,"Team Rocket's Watchtower","Estadio"),
                    (3,"Fire Energy","Energía"),(3,"Psychic Energy","Energía"),(2,"Darkness Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-slowking", "nombre":"Slowking", "tier":"A",
                "uso":4.49, "win_rate":47.91, "top_cut":20.83, "entradas":168,
                "tendencia":0, "costo":48900, "mejor":"4.º NAIC 2026",
                "descripcion":"Motor de Seek Inspiration con atacantes diversos y una caja de recursos muy flexible.",
                "lista_fuente":"Brennan Kamerman · 2.º Special Event Turin 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Slowpoke","Pokémon"),(3,"Slowking","Pokémon"),(2,"Mega Kangaskhan ex","Pokémon"),
                    (2,"Latias ex","Pokémon"),(2,"Kyurem","Pokémon"),(2,"Metagross","Pokémon"),
                    (1,"Lillie's Clefairy ex","Pokémon"),(1,"Meowth ex","Pokémon"),
                    (1,"Fezandipiti ex","Pokémon"),(1,"Smoochum","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(3,"Ciphermaniac's Codebreaking","Entrenador"),
                    (1,"Dawn","Entrenador"),(1,"Lana's Aid","Entrenador"),(1,"Surfer","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Poké Pad","Objeto"),(3,"Wondrous Patch","Objeto"),
                    (2,"Night Stretcher","Objeto"),(1,"Switch","Objeto"),(1,"Secret Box","Objeto"),
                    (1,"Lucky Helmet","Herramienta"),(1,"Brave Bangle","Herramienta"),(4,"Academy at Night","Estadio"),
                    (4,"Telepathic Psychic Energy","Energía"),(3,"Psychic Energy","Energía"),(3,"Boomerang Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-raging-bolt", "nombre":"Raging Bolt Ogerpon", "tier":"B",
                "uso":4.35, "win_rate":42.64, "top_cut":11.66, "entradas":163,
                "tendencia":0, "costo":76900, "mejor":"4.º Special Event Turin 2026",
                "descripcion":"Ataques explosivos y aceleración multicolor con Ogerpon y Mega Kangaskhan.",
                "lista_fuente":"Toby Clark · 4.º Special Event Turin 2026 · lista exacta",
                "core":_meta_core([
                    (3,"Mega Kangaskhan ex","Pokémon"),(3,"Meowth ex","Pokémon"),
                    (3,"Teal Mask Ogerpon ex","Pokémon"),(2,"Raging Bolt ex","Pokémon"),
                    (2,"Latias ex","Pokémon"),(1,"Lillie's Clefairy ex","Pokémon"),
                    (1,"Wellspring Mask Ogerpon ex","Pokémon"),(1,"Iron Leaves ex","Pokémon"),
                    (1,"Fezandipiti ex","Pokémon"),(1,"Passimian","Pokémon"),(1,"Chien-Pao","Pokémon"),
                    (4,"Crispin","Entrenador"),(2,"Boss's Orders","Entrenador"),(2,"Cyrano","Entrenador"),
                    (1,"Ciphermaniac's Codebreaking","Entrenador"),(1,"Lillie's Determination","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Energy Switch","Objeto"),(2,"Night Stretcher","Objeto"),
                    (2,"Glass Trumpet","Objeto"),(1,"Unfair Stamp","Objeto"),(4,"Area Zero Underdepths","Estadio"),
                    (7,"Grass Energy","Energía"),(2,"Lightning Energy","Energía"),(2,"Fighting Energy","Energía"),
                    (2,"Psychic Energy","Energía"),(1,"Water Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-mega-greninja", "nombre":"Mega Greninja", "tier":"B",
                "uso":3.29, "win_rate":37.89, "top_cut":5.69, "entradas":123,
                "tendencia":0, "costo":72900, "mejor":"Top 200 NAIC 2026",
                "descripcion":"Arquetipo de evolución que combina presión de banca y herramientas de control de ritmo.",
                "lista_fuente":"Arquetipo NAIC 2026 · núcleo representativo legal H/I/J",
                "core":_meta_core([
                    (4,"Froakie","Pokémon"),(3,"Frogadier","Pokémon"),(3,"Mega Greninja ex","Pokémon"),
                    (3,"Mega Kangaskhan ex","Pokémon"),(2,"Latias ex","Pokémon"),(1,"Fezandipiti ex","Pokémon"),
                    (1,"Meowth ex","Pokémon"),(4,"Lillie's Determination","Entrenador"),
                    (3,"Crispin","Entrenador"),(3,"Boss's Orders","Entrenador"),(4,"Ultra Ball","Objeto"),
                    (4,"Poké Pad","Objeto"),(4,"Buddy-Buddy Poffin","Objeto"),(3,"Rare Candy","Objeto"),
                    (2,"Night Stretcher","Objeto"),(1,"Unfair Stamp","Objeto"),(2,"Academy at Night","Estadio"),
                    (8,"Water Energy","Energía"),(4,"Psychic Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-crustle", "nombre":"Crustle", "tier":"A",
                "uso":3.23, "win_rate":54.40, "top_cut":28.93, "entradas":121,
                "tendencia":0, "costo":53200, "mejor":"5.º NAIC 2026",
                "descripcion":"Estrategia defensiva y disruptiva que destacó por su excelente win rate y conversión a Día 2.",
                "lista_fuente":"Rahul Reddy · 5.º NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Mega Kangaskhan ex","Pokémon"),(3,"Dwebble","Pokémon"),(3,"Crustle","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(4,"Boss's Orders","Entrenador"),
                    (4,"Team Rocket's Petrel","Entrenador"),(2,"Hilda","Entrenador"),(2,"Eri","Entrenador"),
                    (1,"Xerosic's Machinations","Entrenador"),(1,"Pokémon Center Lady","Entrenador"),
                    (1,"Bianca's Devotion","Entrenador"),(1,"Lisia's Appeal","Entrenador"),
                    (4,"Jumbo Ice Cream","Objeto"),(3,"Pokégear 3.0","Objeto"),(2,"Buddy-Buddy Poffin","Objeto"),
                    (1,"Ultra Ball","Objeto"),(1,"Switch","Objeto"),(1,"Hand Trimmer","Objeto"),
                    (1,"Hero's Cape","Herramienta"),(1,"Handheld Fan","Herramienta"),
                    (1,"Team Rocket's Factory","Estadio"),(1,"Community Center","Estadio"),(1,"Festival Grounds","Estadio"),
                    (4,"Spiky Energy","Energía"),(4,"Growing Grass Energy","Energía"),
                    (4,"Mist Energy","Energía"),(1,"Grass Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-clefairy", "nombre":"Lillie's Clefairy", "tier":"A",
                "uso":1.12, "win_rate":51.48, "top_cut":16.67, "entradas":42,
                "tendencia":0, "costo":55500, "mejor":"1.º NAIC 2026",
                "descripcion":"El mazo campeón del NAIC: caja de atacantes con Clefairy, Kangaskhan, Meowth y energías multicolor.",
                "lista_fuente":"James Kowalski · campeón NAIC 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Mega Kangaskhan ex","Pokémon"),(4,"Meowth ex","Pokémon"),
                    (4,"Lillie's Clefairy ex","Pokémon"),(3,"Latias ex","Pokémon"),
                    (2,"Wellspring Mask Ogerpon ex","Pokémon"),(2,"Fezandipiti ex","Pokémon"),
                    (1,"Moltres","Pokémon"),(1,"Chien-Pao","Pokémon"),(1,"Koraidon ex","Pokémon"),
                    (4,"Crispin","Entrenador"),(3,"Boss's Orders","Entrenador"),
                    (2,"Ciphermaniac's Codebreaking","Entrenador"),(1,"Cyrano","Entrenador"),
                    (4,"Ultra Ball","Objeto"),(4,"Dusk Ball","Objeto"),(3,"Wondrous Patch","Objeto"),
                    (1,"Prime Catcher","Objeto"),(2,"Lillie's Pearl","Herramienta"),
                    (4,"Area Zero Underdepths","Estadio"),(4,"Psychic Energy","Energía"),
                    (2,"Water Energy","Energía"),(2,"Fighting Energy","Energía"),
                    (1,"Telepathic Psychic Energy","Energía"),(1,"Fire Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
            {
                "id":"pkm-trevenant", "nombre":"Hop's Trevenant", "tier":"B",
                "uso":1.44, "win_rate":43.27, "top_cut":9.26, "entradas":54,
                "tendencia":0, "costo":31500, "mejor":"1.º Special Event Turin 2026",
                "descripcion":"Mazo sorpresa ganador de Turín, con atacantes de Hop y un plan disruptivo de recursos.",
                "lista_fuente":"Jose López · campeón Special Event Turin 2026 · lista exacta",
                "core":_meta_core([
                    (4,"Hop's Phantump","Pokémon"),(3,"Hop's Trevenant","Pokémon"),
                    (2,"Hop's Snorlax","Pokémon"),(1,"Hop's Wooloo","Pokémon"),(1,"Hop's Dubwool","Pokémon"),
                    (1,"Hop's Cramorant","Pokémon"),(1,"Lillie's Clefairy ex","Pokémon"),(1,"Shaymin","Pokémon"),
                    (4,"Lillie's Determination","Entrenador"),(4,"Team Rocket's Petrel","Entrenador"),
                    (3,"Boss's Orders","Entrenador"),(2,"Hassel","Entrenador"),(1,"Ruffian","Entrenador"),
                    (4,"Poké Pad","Objeto"),(3,"Pokégear 3.0","Objeto"),(3,"Night Stretcher","Objeto"),
                    (2,"Hop's Bag","Objeto"),(1,"Ultra Ball","Objeto"),(1,"Secret Box","Objeto"),
                    (1,"Switch","Objeto"),(4,"Hop's Choice Band","Herramienta"),(1,"Air Balloon","Herramienta"),
                    (4,"Postwick","Estadio"),(4,"Telepathic Psychic Energy","Energía"),(4,"Mist Energy","Energía")
                ]), "favorables":[], "dificiles":[]
            },
        ],
        "torneos": [
            {
                "id":"pkm-event-naic-2026", "evento":"NAIC 2026 · New Orleans",
                "fecha":"2026-06-12 al 2026-06-14", "region":"Norteamérica", "jugadores":3752,
                "ganador":"James Kowalski", "mazo":"Lillie's Clefairy",
                "top8":["Lillie's Clefairy","Dragapult Dusknoir","Dragapult","Slowking","Crustle","Dragapult Blaziken","Rocket's Mewtwo","Dragapult"]
            },
            {
                "id":"pkm-event-turin-2026", "evento":"Special Event Turin 2026",
                "fecha":"2026-06-06 al 2026-06-07", "region":"Europa", "jugadores":2033,
                "ganador":"Jose López", "mazo":"Hop's Trevenant",
                "top8":["Hop's Trevenant","Slowking","Dragapult Blaziken","Raging Bolt Ogerpon","Lillie's Clefairy","Alakazam Dudunsparce","Dragapult Blaziken","Dragapult"]
            },
            {
                "id":"pkm-event-indianapolis-2026", "evento":"Regional Indianapolis 2026",
                "fecha":"2026-05-30 al 2026-05-31", "region":"Norteamérica", "jugadores":1974,
                "ganador":"Cerys Jones", "mazo":"Alakazam Dudunsparce",
                "top8":["Alakazam Dudunsparce","Crustle","Dragapult Dusknoir","Dragapult","Dragapult","Dragapult","Dragapult","Dragapult"]
            },
        ],
    },
    "One Piece Card Game": {
        "formatos": ["Standard"], "actualizado":"Dataset demostrativo · preparado para validación",
        "mazos": [
            {"id":"op-dofla","nombre":"Blue Doflamingo","tier":"S","uso":14.4,"win_rate":55.1,"top_cut":19.3,"entradas":226,"tendencia":1,"costo":119900,"mejor":"1.º","descripcion":"Tempo, presión de personajes y gran eficiencia de recursos.","core":_meta_core([(1,"Donquixote Doflamingo","Líder"),(4,"Jinbe","Personaje"),(4,"Boa Hancock","Personaje"),(4,"Gecko Moria","Personaje"),(4,"Crocodile","Personaje"),(4,"Gravity Blade Raging Tiger","Evento"),(4,"The Seven Warlords of the Sea","Evento"),(4,"Perfume Femur","Evento")],50,"Otras cartas del mazo"),"favorables":["Black Lucci","Red Shanks","Yellow Enel"],"dificiles":["Purple Luffy","Green Bonney","Blackbeard"]},
            {"id":"op-lucci","nombre":"Black Lucci","tier":"S","uso":12.6,"win_rate":54.0,"top_cut":17.1,"entradas":198,"tendencia":2,"costo":104900,"mejor":"1.º","descripcion":"Control de mesa y reducción de coste para remover amenazas.","core":_meta_core([(1,"Rob Lucci","Líder"),(4,"Rob Lucci","Personaje"),(4,"Gecko Moria","Personaje"),(4,"Rebecca","Personaje"),(4,"Sabo","Personaje"),(4,"Spandine","Personaje"),(4,"Tempest Kick","Evento")],50,"Otras cartas del mazo"),"favorables":["Green Bonney","Yellow Enel","Red Shanks"],"dificiles":["Blue Doflamingo","Purple Luffy","Blackbeard"]},
            {"id":"op-pluffy","nombre":"Purple Luffy","tier":"A","uso":11.9,"win_rate":53.7,"top_cut":15.8,"entradas":187,"tendencia":1,"costo":94900,"mejor":"2.º","descripcion":"Aceleración de DON!! y amenazas de alto impacto antes de lo esperado.","core":_meta_core([(1,"Monkey D. Luffy","Líder"),(4,"Monkey D. Luffy","Personaje"),(4,"Magellan","Personaje"),(4,"Paulie","Personaje"),(4,"Kaido","Personaje"),(4,"Queen","Personaje"),(4,"Onigashima Island","Stage")],50,"Otras cartas del mazo"),"favorables":["Blue Doflamingo","Black Lucci","Red Shanks"],"dificiles":["Yellow Enel","Green Bonney","Blackbeard"]},
            {"id":"op-shanks","nombre":"Red Shanks","tier":"A","uso":10.8,"win_rate":52.8,"top_cut":13.9,"entradas":170,"tendencia":-1,"costo":109900,"mejor":"3.º","descripcion":"Presión lineal, cuerpos grandes y excelente cierre de partida.","core":_meta_core([(1,"Shanks","Líder"),(4,"Shanks","Personaje"),(4,"Benn Beckman","Personaje"),(4,"Lucky Roux","Personaje"),(4,"Yasopp","Personaje"),(4,"Monkey D. Luffy","Personaje"),(4,"Red-Haired Pirates","Evento")],50,"Otras cartas del mazo"),"favorables":["Green Bonney","Yellow Enel","Blackbeard"],"dificiles":["Blue Doflamingo","Black Lucci","Purple Luffy"]},
            {"id":"op-bonney","nombre":"Green Bonney","tier":"A","uso":9.7,"win_rate":52.4,"top_cut":12.8,"entradas":152,"tendencia":0,"costo":88900,"mejor":"4.º","descripcion":"Control por descanso y valor incremental en partidas largas.","core":_meta_core([(1,"Jewelry Bonney","Líder"),(4,"Jewelry Bonney","Personaje"),(4,"Basil Hawkins","Personaje"),(4,"Cavendish","Personaje"),(4,"Donquixote Rosinante","Personaje"),(4,"Eustass Kid","Personaje"),(4,"Because the Side of Justice Will Be Whichever Side Wins!!","Evento")],50,"Otras cartas del mazo"),"favorables":["Purple Luffy","Blackbeard","Yellow Enel"],"dificiles":["Black Lucci","Red Shanks","Blue Doflamingo"]},
            {"id":"op-enel","nombre":"Yellow Enel","tier":"A","uso":8.9,"win_rate":51.9,"top_cut":11.2,"entradas":140,"tendencia":-1,"costo":79900,"mejor":"5.º","descripcion":"Manipulación de vida y enorme capacidad de supervivencia.","core":_meta_core([(1,"Enel","Líder"),(4,"Enel","Personaje"),(4,"Charlotte Katakuri","Personaje"),(4,"Yamato","Personaje"),(4,"Kikunojo","Personaje"),(4,"Gedatsu","Personaje"),(4,"You're the One Who Should Disappear.","Evento")],50,"Otras cartas del mazo"),"favorables":["Purple Luffy","Blackbeard","Red Shanks"],"dificiles":["Blue Doflamingo","Black Lucci","Green Bonney"]},
            {"id":"op-blackbeard","nombre":"Blackbeard","tier":"B","uso":7.6,"win_rate":51.4,"top_cut":9.8,"entradas":119,"tendencia":2,"costo":98900,"mejor":"6.º","descripcion":"Negación de habilidades y amenazas resistentes.","core":_meta_core([(1,"Marshall.D.Teach","Líder"),(4,"Marshall.D.Teach","Personaje"),(4,"Jesus Burgess","Personaje"),(4,"Van Augur","Personaje"),(4,"Doc Q","Personaje"),(4,"Shiryu","Personaje"),(4,"Black Vortex","Evento")],50,"Otras cartas del mazo"),"favorables":["Blue Doflamingo","Black Lucci","Purple Luffy"],"dificiles":["Red Shanks","Green Bonney","Yellow Enel"]},
            {"id":"op-nami","nombre":"Blue Nami","tier":"B","uso":5.6,"win_rate":50.8,"top_cut":7.0,"entradas":88,"tendencia":0,"costo":55900,"mejor":"9.º","descripcion":"Condición de victoria alternativa basada en vaciar el deck.","core":_meta_core([(1,"Nami","Líder"),(4,"Nami","Personaje"),(4,"Kayas","Personaje"),(4,"Mr.1 (Daz.Bonez)","Personaje"),(4,"Love-Love Mellow","Evento"),(4,"Gum-Gum Rain","Evento"),(4,"Pilaf's Treasure","Evento")],50,"Otras cartas del mazo"),"favorables":["Black Lucci","Green Bonney","Yellow Enel"],"dificiles":["Red Shanks","Purple Luffy","Blue Doflamingo"]},
            {"id":"op-law","nombre":"Red/Purple Law","tier":"B","uso":4.8,"win_rate":49.9,"top_cut":6.2,"entradas":75,"tendencia":-2,"costo":84900,"mejor":"11.º","descripcion":"Combo de reducción y reposición de personajes.","core":_meta_core([(1,"Trafalgar Law","Líder"),(4,"Trafalgar Law","Personaje"),(4,"Bepo","Personaje"),(4,"Shachi & Penguin","Personaje"),(4,"Raise Max","Personaje"),(4,"Gordon","Personaje"),(4,"Kid & Killer","Personaje")],50,"Otras cartas del mazo"),"favorables":["Yellow Enel","Green Bonney","Blackbeard"],"dificiles":["Blue Doflamingo","Black Lucci","Red Shanks"]},
            {"id":"op-katakuri","nombre":"Yellow Katakuri","tier":"B","uso":3.7,"win_rate":49.5,"top_cut":5.1,"entradas":58,"tendencia":0,"costo":67900,"mejor":"14.º","descripcion":"Curva sólida y manipulación superior de vidas.","core":_meta_core([(1,"Charlotte Katakuri","Líder"),(4,"Charlotte Katakuri","Personaje"),(4,"Charlotte Linlin","Personaje"),(4,"Charlotte Cracker","Personaje"),(4,"Sanji","Personaje"),(4,"Gedatsu","Personaje"),(4,"Thunder Bolt","Evento")],50,"Otras cartas del mazo"),"favorables":["Purple Luffy","Blackbeard","Red/Purple Law"],"dificiles":["Blue Doflamingo","Black Lucci","Green Bonney"]},
        ],
        "torneos":[
            {"id":"op-event-1","evento":"Championship Finals A · DEMO","fecha":"2026-05-25","region":"Norteamérica","jugadores":768,"ganador":"Player Grand","mazo":"Blue Doflamingo","top8":["Blue Doflamingo","Black Lucci","Purple Luffy","Red Shanks","Blackbeard","Green Bonney","Blue Doflamingo","Yellow Enel"]},
            {"id":"op-event-2","evento":"Regional Oceanía · DEMO","fecha":"2026-05-11","region":"Oceanía","jugadores":442,"ganador":"Player Cipher","mazo":"Black Lucci","top8":["Black Lucci","Purple Luffy","Blue Doflamingo","Green Bonney","Red Shanks","Black Lucci","Yellow Enel","Blackbeard"]},
            {"id":"op-event-3","evento":"Treasure Cup Europa · DEMO","fecha":"2026-04-27","region":"Europa","jugadores":612,"ganador":"Player Violet","mazo":"Purple Luffy","top8":["Purple Luffy","Blue Doflamingo","Red Shanks","Black Lucci","Green Bonney","Yellow Enel","Purple Luffy","Blue Nami"]},
        ],
    },
    "Magic: The Gathering": {
        "formatos": ["Standard", "Modern"], "actualizado":"Dataset demostrativo · preparado para validación",
        "mazos": [
            {"id":"mtg-izzet","nombre":"Izzet Prowess","tier":"S","uso":13.1,"win_rate":54.3,"top_cut":17.6,"entradas":258,"tendencia":2,"costo":164900,"mejor":"1.º","descripcion":"Amenazas eficientes y hechizos baratos para cerrar partidas rápidamente.","core":_meta_core([(4,"Monastery Swiftspear","Criatura"),(4,"Slickshot Show-Off","Criatura"),(4,"Opt","Instantáneo"),(4,"Lightning Strike","Instantáneo"),(4,"Play with Fire","Instantáneo"),(4,"Monstrous Rage","Instantáneo"),(4,"Spirebluff Canal","Tierra"),(4,"Shivan Reef","Tierra")],60,"Otras cartas y tierras"),"favorables":["Domain Ramp","Azorius Control","Temur Analyst"],"dificiles":["Golgari Midrange","Boros Convoke","Mono-Black"]},
            {"id":"mtg-domain","nombre":"Domain Ramp","tier":"S","uso":12.2,"win_rate":53.7,"top_cut":16.4,"entradas":240,"tendencia":0,"costo":249900,"mejor":"1.º","descripcion":"Aceleración, dominio y amenazas de alto impacto en el juego tardío.","core":_meta_core([(4,"Up the Beanstalk","Encantamiento"),(4,"Leyline Binding","Encantamiento"),(4,"Topiary Stomper","Criatura"),(4,"Atraxa, Grand Unifier","Criatura"),(4,"Herd Migration","Conjuro"),(4,"Sunfall","Conjuro"),(4,"Jetmir's Garden","Tierra"),(4,"Spara's Headquarters","Tierra")],60,"Otras cartas y tierras"),"favorables":["Golgari Midrange","Mono-Black","Azorius Control"],"dificiles":["Izzet Prowess","Boros Convoke","Mono-Red Aggro"]},
            {"id":"mtg-golgari","nombre":"Golgari Midrange","tier":"A","uso":11.4,"win_rate":53.2,"top_cut":15.0,"entradas":224,"tendencia":1,"costo":219900,"mejor":"2.º","descripcion":"Interacción, criaturas resilientes y valor incremental.","core":_meta_core([(4,"Deep-Cavern Bat","Criatura"),(4,"Mosswood Dreadknight","Criatura"),(4,"Glissa Sunslayer","Criatura"),(4,"Go for the Throat","Instantáneo"),(4,"Cut Down","Instantáneo"),(4,"Liliana of the Veil","Planeswalker"),(4,"Restless Cottage","Tierra"),(4,"Llanowar Wastes","Tierra")],60,"Otras cartas y tierras"),"favorables":["Izzet Prowess","Mono-Red Aggro","Boros Convoke"],"dificiles":["Domain Ramp","Azorius Control","Temur Analyst"]},
            {"id":"mtg-boros","nombre":"Boros Convoke","tier":"A","uso":10.1,"win_rate":52.8,"top_cut":13.7,"entradas":199,"tendencia":1,"costo":129900,"mejor":"3.º","descripcion":"Despliegue explosivo y presión de mesa desde los primeros turnos.","core":_meta_core([(4,"Novice Inspector","Criatura"),(4,"Voldaren Epicure","Criatura"),(4,"Warden of the Inner Sky","Criatura"),(4,"Knight-Errant of Eos","Criatura"),(4,"Imodane's Recruiter","Criatura"),(4,"Gleeful Demolition","Conjuro"),(4,"Battlefield Forge","Tierra"),(4,"Inspiring Vantage","Tierra")],60,"Otras cartas y tierras"),"favorables":["Domain Ramp","Azorius Control","Temur Analyst"],"dificiles":["Golgari Midrange","Mono-Black","Izzet Prowess"]},
            {"id":"mtg-azorius","nombre":"Azorius Control","tier":"A","uso":9.0,"win_rate":52.1,"top_cut":12.1,"entradas":177,"tendencia":-1,"costo":239900,"mejor":"4.º","descripcion":"Contrahechizos, limpiezas y inevitabilidad.","core":_meta_core([(4,"No More Lies","Instantáneo"),(4,"Three Steps Ahead","Instantáneo"),(4,"Sunfall","Conjuro"),(4,"Temporary Lockdown","Encantamiento"),(4,"The Wandering Emperor","Planeswalker"),(4,"Memory Deluge","Instantáneo"),(4,"Adarkar Wastes","Tierra"),(4,"Restless Anchorage","Tierra")],60,"Otras cartas y tierras"),"favorables":["Golgari Midrange","Mono-Black","Temur Analyst"],"dificiles":["Izzet Prowess","Boros Convoke","Domain Ramp"]},
            {"id":"mtg-red","nombre":"Mono-Red Aggro","tier":"A","uso":8.6,"win_rate":51.8,"top_cut":10.9,"entradas":169,"tendencia":-1,"costo":74900,"mejor":"5.º","descripcion":"Curva baja y daño directo para castigar salidas lentas.","core":_meta_core([(4,"Monastery Swiftspear","Criatura"),(4,"Phoenix Chick","Criatura"),(4,"Squee, Dubious Monarch","Criatura"),(4,"Play with Fire","Instantáneo"),(4,"Lightning Strike","Instantáneo"),(4,"Monstrous Rage","Instantáneo"),(20,"Mountain","Tierra")],60,"Otras cartas"),"favorables":["Domain Ramp","Azorius Control","Temur Analyst"],"dificiles":["Golgari Midrange","Mono-Black","Boros Convoke"]},
            {"id":"mtg-black","nombre":"Mono-Black Midrange","tier":"B","uso":7.4,"win_rate":51.2,"top_cut":9.4,"entradas":146,"tendencia":1,"costo":184900,"mejor":"6.º","descripcion":"Descartes, remoción y amenazas resistentes.","core":_meta_core([(4,"Deep-Cavern Bat","Criatura"),(4,"Preacher of the Schism","Criatura"),(4,"Sheoldred, the Apocalypse","Criatura"),(4,"Cut Down","Instantáneo"),(4,"Go for the Throat","Instantáneo"),(4,"Duress","Conjuro"),(4,"Mishra's Foundry","Tierra"),(20,"Swamp","Tierra")],60,"Otras cartas"),"favorables":["Izzet Prowess","Boros Convoke","Mono-Red Aggro"],"dificiles":["Domain Ramp","Azorius Control","Temur Analyst"]},
            {"id":"mtg-temur","nombre":"Temur Analyst","tier":"B","uso":6.1,"win_rate":50.8,"top_cut":8.0,"entradas":120,"tendencia":0,"costo":199900,"mejor":"8.º","descripcion":"Motor de tierras y turnos explosivos de valor.","core":_meta_core([(4,"Aftermath Analyst","Criatura"),(4,"Nissa, Resurgent Animist","Criatura"),(4,"Worldsoul's Rage","Conjuro"),(4,"Memory Deluge","Instantáneo"),(4,"Ill-Timed Explosion","Conjuro"),(4,"Spelunking","Encantamiento"),(4,"Brokers Hideout","Tierra"),(4,"Cabaretti Courtyard","Tierra")],60,"Otras cartas y tierras"),"favorables":["Golgari Midrange","Mono-Black","Domain Ramp"],"dificiles":["Izzet Prowess","Boros Convoke","Mono-Red Aggro"]},
            {"id":"mtg-dimir","nombre":"Dimir Midrange","tier":"B","uso":5.2,"win_rate":50.4,"top_cut":6.8,"entradas":102,"tendencia":0,"costo":179900,"mejor":"10.º","descripcion":"Amenazas evasivas e interacción eficiente.","core":_meta_core([(4,"Deep-Cavern Bat","Criatura"),(4,"Faerie Mastermind","Criatura"),(4,"Gix, Yawgmoth Praetor","Criatura"),(4,"Go for the Throat","Instantáneo"),(4,"Make Disappear","Instantáneo"),(4,"Cut Down","Instantáneo"),(4,"Underground River","Tierra"),(4,"Darkslick Shores","Tierra")],60,"Otras cartas y tierras"),"favorables":["Izzet Prowess","Azorius Control","Mono-Red Aggro"],"dificiles":["Domain Ramp","Golgari Midrange","Boros Convoke"]},
            {"id":"mtg-selesnya","nombre":"Selesnya Enchantments","tier":"B","uso":4.1,"win_rate":49.9,"top_cut":5.6,"entradas":81,"tendencia":-1,"costo":139900,"mejor":"13.º","descripcion":"Sinergias de encantamientos y crecimiento acumulativo.","core":_meta_core([(4,"Generous Visitor","Criatura"),(4,"Spirited Companion","Criatura"),(4,"Calix, Guided by Fate","Criatura"),(4,"Audacity","Encantamiento"),(4,"Ossification","Encantamiento"),(4,"Michiko's Reign of Truth","Encantamiento"),(4,"Brushland","Tierra"),(4,"Razorverge Thicket","Tierra")],60,"Otras cartas y tierras"),"favorables":["Mono-Black","Golgari Midrange","Azorius Control"],"dificiles":["Izzet Prowess","Boros Convoke","Domain Ramp"]},
        ],
        "torneos":[
            {"id":"mtg-event-1","evento":"Regional Championship A · DEMO","fecha":"2026-05-30","region":"Europa","jugadores":948,"ganador":"Player Spark","mazo":"Izzet Prowess","top8":["Izzet Prowess","Domain Ramp","Golgari Midrange","Boros Convoke","Azorius Control","Izzet Prowess","Mono-Black Midrange","Temur Analyst"]},
            {"id":"mtg-event-2","evento":"Spotlight Series Pacific · DEMO","fecha":"2026-05-16","region":"Asia-Pacífico","jugadores":612,"ganador":"Player Grove","mazo":"Domain Ramp","top8":["Domain Ramp","Golgari Midrange","Izzet Prowess","Azorius Control","Boros Convoke","Domain Ramp","Mono-Red Aggro","Dimir Midrange"]},
            {"id":"mtg-event-3","evento":"Arena Championship Qualifier · DEMO","fecha":"2026-05-02","region":"Online","jugadores":512,"ganador":"Player Bloom","mazo":"Golgari Midrange","top8":["Golgari Midrange","Izzet Prowess","Boros Convoke","Domain Ramp","Mono-Black Midrange","Azorius Control","Temur Analyst","Golgari Midrange"]},
        ],
    },
}

META_DEMO_COLLECTION = {
    "Dreepy":4,"Drakloak":4,"Dragapult ex":2,"Duskull":2,"Dusclops":1,"Dusknoir":1,
    "Munkidori":1,"Meowth ex":1,"Fezandipiti ex":1,"Lillie's Determination":4,
    "Boss's Orders":3,"Crispin":2,"Buddy-Buddy Poffin":4,"Poké Pad":4,"Ultra Ball":4,
    "Night Stretcher":2,"Psychic Energy":4,"Fire Energy":3,"Darkness Energy":2,
    "N's Zorua":3,"N's Zoroark ex":2,"Lillie's Clefairy ex":2,"Mega Kangaskhan ex":2,
}


def _init_meta_state() -> None:
    defaults = {
        "meta_game": "Pokémon TCG", "meta_format": "Standard", "selected_meta_deck": None,
        "meta_uploaded_df": None, "meta_upload_name": "", "meta_use_demo_collection": True,
        "deck_cart": {}, "meta_selected_card": None, "meta_print_choices": {},
        "meta_deck_view": "Galería visual", "meta_only_missing": False,
        "meta_missing_estimates": {},
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _meta_norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _parse_decklist_text(raw: str, deck_size: int = 60) -> list[dict]:
    rows = []
    for part in re.split(r"[;\n|]+", str(raw or "")):
        part = part.strip()
        if not part:
            continue
        match = re.match(r"^\s*(\d+)\s*[xX]?\s+(.+?)\s*$", part)
        if match:
            rows.append({"cantidad": int(match.group(1)), "carta": match.group(2).strip(), "categoria": "Importada"})
        else:
            rows.append({"cantidad": 1, "carta": part, "categoria": "Importada"})
    total = sum(x["cantidad"] for x in rows)
    if rows and total < deck_size:
        rows.append({"cantidad": deck_size-total, "carta":"Otras cartas no informadas", "categoria":"Soporte"})
    return rows


def _uploaded_meta_dataset(game: str, fmt: str) -> dict | None:
    df = st.session_state.get("meta_uploaded_df")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return None
    data = df.copy()
    data.columns = [_meta_norm(c).replace(" ", "_") for c in data.columns]
    aliases = {
        "juego":"game", "formato":"format", "evento":"event", "fecha":"date", "jugador":"player",
        "posicion":"placement", "puesto":"placement", "mazo":"archetype", "arquetipo":"archetype",
        "rival":"opponent_archetype", "mazo_rival":"opponent_archetype", "resultado":"result",
        "participantes":"players", "lista":"decklist", "deck_list":"decklist",
    }
    data = data.rename(columns={c: aliases.get(c, c) for c in data.columns})
    required = {"event", "player", "archetype"}
    if not required.issubset(data.columns):
        return None
    if "game" in data.columns:
        data = data[data["game"].astype(str).map(_meta_norm) == _meta_norm(game)]
    if "format" in data.columns:
        data = data[data["format"].astype(str).map(_meta_norm) == _meta_norm(fmt)]
    if data.empty:
        return None

    entries = data.drop_duplicates(subset=["event", "player", "archetype"])
    total_entries = max(len(entries), 1)
    decks = []
    for archetype, group in entries.groupby("archetype", dropna=True):
        arch = str(archetype).strip()
        if not arch:
            continue
        raw_group = data[data["archetype"].astype(str) == str(archetype)]
        result_scores = []
        if "result" in raw_group.columns:
            for value in raw_group["result"].dropna().astype(str):
                v = _meta_norm(value)
                if v in {"w","win","gano","ganada","victoria","2 0","2 1"}: result_scores.append(1.0)
                elif v in {"l","loss","perdio","perdida","derrota","0 2","1 2"}: result_scores.append(0.0)
                elif v in {"d","draw","empate","1 1"}: result_scores.append(0.5)
        win_rate = (sum(result_scores)/len(result_scores)*100) if result_scores else 50.0
        placements = pd.to_numeric(group.get("placement", pd.Series(dtype=float)), errors="coerce")
        topcut = float((placements <= 8).mean()*100) if len(placements.dropna()) else 0.0
        best = int(placements.min()) if len(placements.dropna()) else None
        decklist = []
        if "decklist" in raw_group.columns:
            values = raw_group["decklist"].dropna().astype(str)
            if not values.empty:
                decklist = _parse_decklist_text(values.iloc[0], 60 if "magic" in _meta_norm(game) or "pokemon" in _meta_norm(game) else 50)
        share = len(group) / total_entries * 100
        score = share * .55 + win_rate * .45
        tier = "S" if score >= 29 else "A" if score >= 23 else "B"
        decks.append({
            "id":f"upload-{_meta_norm(arch).replace(' ','-')}", "nombre":arch, "tier":tier,
            "uso":round(share,1), "win_rate":round(win_rate,1), "top_cut":round(topcut,1),
            "entradas":int(len(group)), "tendencia":0, "costo":0, "mejor":f"{best}.º" if best else "—",
            "descripcion":"Arquetipo calculado desde el archivo cargado por el anfitrión.",
            "core":decklist, "favorables":[], "dificiles":[],
        })
    decks = sorted(decks, key=lambda x:(x["uso"],x["win_rate"]), reverse=True)

    tournaments = []
    for event, group in entries.groupby("event", dropna=True):
        placements = pd.to_numeric(group.get("placement", pd.Series(dtype=float)), errors="coerce")
        ordered = group.assign(_placement=placements).sort_values("_placement", na_position="last")
        winner = ordered.iloc[0] if len(ordered) else None
        date = "—"
        if "date" in group.columns and len(group["date"].dropna()): date = str(group["date"].dropna().iloc[0])
        players = int(pd.to_numeric(group.get("players", pd.Series(dtype=float)), errors="coerce").max()) if "players" in group.columns and pd.to_numeric(group["players"], errors="coerce").notna().any() else int(group["player"].nunique())
        tournaments.append({
            "id":f"upload-event-{_meta_norm(event).replace(' ','-')}", "evento":str(event), "fecha":date,
            "region":"Cargado por anfitrión", "jugadores":players,
            "ganador":str(winner.get("player","—")) if winner is not None else "—",
            "mazo":str(winner.get("archetype","—")) if winner is not None else "—",
            "top8":ordered.head(8)["archetype"].astype(str).tolist(),
        })
    return {"formatos":[fmt], "actualizado":f"Archivo: {st.session_state.get('meta_upload_name','datos cargados')}", "mazos":decks, "torneos":tournaments, "raw":data}


def _meta_dataset(game: str, fmt: str) -> tuple[dict, str]:
    uploaded = _uploaded_meta_dataset(game, fmt)
    if uploaded and uploaded.get("mazos"):
        return uploaded, "Resultados cargados por el anfitrión"
    dataset = META_DEMO.get(game, META_DEMO["Pokémon TCG"])
    return dataset, dataset.get("fuente", "Datos demostrativos")


def _meta_find_deck(dataset: dict, deck_id_or_name: str | None) -> dict | None:
    target = str(deck_id_or_name or "")
    return next((d for d in dataset.get("mazos", []) if d.get("id") == target or d.get("nombre") == target), None)


def _meta_matchup_value(a: dict, b: dict) -> float:
    if a.get("id") == b.get("id"):
        return 50.0
    # Matriz ilustrativa coherente y simétrica, basada en rendimiento y afinidades declaradas.
    base = 50 + (float(a.get("win_rate",50))-float(b.get("win_rate",50))) * 1.35
    if b.get("nombre") in a.get("favorables",[]): base += 5.5
    if b.get("nombre") in a.get("dificiles",[]): base -= 5.5
    return max(37.0, min(63.0, round(base,1)))


def _inventory_from_results() -> dict[str, int]:
    df = st.session_state.get("df_result")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return {}
    inventory: dict[str,int] = {}
    for _, row in df.iterrows():
        name = row.get("Nombre EN") or row.get("Nombre Original") or ""
        if not str(name).strip(): continue
        try: qty = int(row.get("Cantidad",1) or 1)
        except Exception: qty = 1
        inventory[_meta_norm(name)] = inventory.get(_meta_norm(name),0) + max(qty,0)
    return inventory


def _deck_completion(deck: dict, inventory: dict[str,int]) -> tuple[float, list[dict], int, int]:
    relevant = [x for x in deck.get("core",[]) if not str(x.get("carta","")).lower().startswith("otras cartas") and "basic energy" not in str(x.get("carta","")).lower()]
    required = sum(int(x.get("cantidad",0)) for x in relevant)
    owned_total = 0; missing = []
    for row in relevant:
        needed = int(row.get("cantidad",0)); name = str(row.get("carta","")); owned = inventory.get(_meta_norm(name),0)
        used = min(needed, owned); owned_total += used
        if used < needed:
            missing.append({"Carta":name,"Necesitas":needed,"Tienes":owned,"Faltan":needed-used,"Categoría":row.get("categoria","—")})
    pct = (owned_total/required*100) if required else 0.0
    return round(pct,1), missing, owned_total, required


def _meta_decklist_text(deck: dict) -> str:
    lines = [f"{deck.get('nombre','Mazo')} · NexoGeek Meta Lab", ""]
    categories = []
    for row in deck.get("core",[]):
        cat = row.get("categoria","Otros")
        if cat not in categories: categories.append(cat)
    for cat in categories:
        rows = [x for x in deck.get("core",[]) if x.get("categoria","Otros") == cat]
        lines.append(f"# {cat} ({sum(int(x.get('cantidad',0)) for x in rows)})")
        lines.extend(f"{int(x.get('cantidad',0))} {x.get('carta','')}" for x in rows)
        lines.append("")
    lines.append("Lista demostrativa para validar la experiencia. Verificar legalidad y versión antes de jugar un evento.")
    return "\n".join(lines)


def _meta_search_marketplace(card_name: str, deck_name: str = "") -> None:
    st.session_state["market_search_v3"] = str(card_name)
    st.session_state["meta_marketplace_context"] = {"carta":str(card_name),"mazo":str(deck_name)}
    _track_event("buscar_carta_meta", str(card_name), str(deck_name))
    _go_to("Marketplace")


# Caché específica de imágenes/impresiones del constructor visual. También
# guarda búsquedas vacías para no repetir consultas fallidas en cada rerun.
_META_VISUAL_CACHE: dict[tuple, list[dict]] = {}
_META_VISUAL_LOCK = threading.Lock()


def _meta_is_real_card_row(row: dict) -> bool:
    name = str(row.get("carta", "")).strip()
    normalized = _meta_norm(name)
    return bool(name) and not normalized.startswith("otras cartas") and normalized not in {"soporte", "cartas no informadas"}


def _meta_query_name(card_name: str) -> str:
    aliases = {
        "psychic energy": "Basic Psychic Energy", "fire energy": "Basic Fire Energy",
        "darkness energy": "Basic Darkness Energy", "water energy": "Basic Water Energy",
        "lightning energy": "Basic Lightning Energy", "grass energy": "Basic Grass Energy",
        "fighting energy": "Basic Fighting Energy", "metal energy": "Basic Metal Energy",
    }
    return aliases.get(_meta_norm(card_name), card_name)


def _meta_card_candidates(card_name: str, game: str, api_key: str | None, limit: int = 10) -> list[dict]:
    """Obtiene impresiones candidatas para la galería.

    Pokémon usa la DB local o PokémonTCG API. Los demás juegos aceptan una
    image_url incluida en la decklist futura y, mientras tanto, usan placeholder.
    """
    if game != "Pokémon TCG" or not str(card_name).strip():
        return []
    query_name = _meta_query_name(card_name)
    cache_key = (_meta_norm(query_name), bool(api_key), bool(_DB_LOADED), int(limit))
    with _META_VISUAL_LOCK:
        if cache_key in _META_VISUAL_CACHE:
            return [dict(x) for x in _META_VISUAL_CACHE[cache_key]]
    try:
        pool = _pool_nombre(query_name, api_key)
    except Exception:
        pool = []
    target = _meta_norm(query_name)
    exact = [c for c in pool if _meta_norm(c.get("name", "")) == target]
    cards = exact or pool

    def score(card: dict) -> tuple:
        mark = str(card.get("regulationMark", "")).upper()
        legal = 3 if mark in {"H", "I", "J"} else 1 if not mark else 0
        image = 1 if _safe_image_url((card.get("images", {}) or {}).get("small", "")) else 0
        has_price = 1 if _tiene_precio(card) else 0
        release = str((card.get("set", {}) or {}).get("releaseDate", ""))
        return legal, image, has_price, release

    cards = sorted(cards, key=score, reverse=True)
    output = []
    seen = set()
    for card in cards:
        cid = str(card.get("id", ""))
        if not cid or cid in seen:
            continue
        image = _safe_image_url((card.get("images", {}) or {}).get("small", ""))
        if not image:
            continue
        price_usd, variant, updated = extraer_precio(card)
        set_data = card.get("set", {}) or {}
        output.append({
            "card_id": cid, "name": card.get("name", card_name), "image_url": image,
            "set": set_data.get("name", "Set no informado"), "number": str(card.get("number", "—")),
            "regulation": str(card.get("regulationMark", "")) or "—",
            "rarity": str(card.get("rarity", "—")), "price_usd": price_usd,
            "variant": variant or "—", "updated": updated or "—",
        })
        seen.add(cid)
        if len(output) >= limit:
            break
    with _META_VISUAL_LOCK:
        _META_VISUAL_CACHE[cache_key] = [dict(x) for x in output]
    return output


def _meta_prefetch_candidates(card_names: list[str], game: str, api_key: str | None) -> None:
    names = list(dict.fromkeys([str(x).strip() for x in card_names if str(x).strip()]))
    if game != "Pokémon TCG" or not names:
        return
    pending = []
    for name in names:
        key = (_meta_norm(_meta_query_name(name)), bool(api_key), bool(_DB_LOADED), 10)
        with _META_VISUAL_LOCK:
            if key not in _META_VISUAL_CACHE:
                pending.append(name)
    if not pending:
        return
    workers = min(4 if api_key or _DB_LOADED else 2, len(pending))
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        futures = [executor.submit(_meta_card_candidates, name, game, api_key) for name in pending]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception:
                pass


def _meta_print_key(deck: dict, card_name: str) -> str:
    return f"{deck.get('id','deck')}|{_meta_norm(card_name)}"


def _meta_selected_visual(deck: dict, card_name: str, game: str, api_key: str | None) -> dict:
    candidates = _meta_card_candidates(card_name, game, api_key)
    if not candidates:
        return {}
    saved_id = st.session_state.get("meta_print_choices", {}).get(_meta_print_key(deck, card_name))
    return next((x for x in candidates if x.get("card_id") == saved_id), candidates[0])


def _meta_market_matches(card_name: str) -> list[dict]:
    needle = _meta_norm(card_name)
    if not needle:
        return []
    matches = []
    for item in st.session_state.get("marketplace_db", []):
        if not item.get("active", True):
            continue
        hay = _meta_norm(item.get("title") or item.get("name") or "")
        if needle == hay or needle in hay or (hay and hay in needle):
            matches.append(item)
    return matches


def _meta_cart_key(deck: dict, card_name: str) -> str:
    return f"{deck.get('id','deck')}::{_meta_norm(card_name)}"


def _meta_cart_items(deck: dict) -> list[dict]:
    prefix = f"{deck.get('id','deck')}::"
    return [dict(v) for k, v in st.session_state.get("deck_cart", {}).items() if str(k).startswith(prefix)]


def _meta_manual_owned_key(deck: dict, card_name: str) -> str:
    """Clave estable para la cantidad marcada manualmente en un mazo."""
    return f"manual_owned::{deck.get('id','deck')}::{_meta_norm(card_name)}"


def _meta_manual_inventory(deck: dict, rows: list[dict]) -> dict[str, int]:
    """Construye el inventario editable del checklist manual."""
    inventory: dict[str, int] = {}
    for row in rows:
        name = str(row.get("carta", ""))
        needed = max(int(row.get("cantidad", 0) or 0), 0)
        key = _meta_manual_owned_key(deck, name)
        if key not in st.session_state:
            st.session_state[key] = 0
        try:
            owned = int(st.session_state.get(key, 0) or 0)
        except Exception:
            owned = 0
        owned = max(0, min(owned, needed))
        st.session_state[key] = owned
        inventory[_meta_norm(name)] = owned
    return inventory


def _meta_set_manual_owned(widget_key: str, qty: int) -> None:
    st.session_state[widget_key] = max(int(qty or 0), 0)


def _meta_fill_manual_inventory(deck: dict, rows: list[dict], source: dict[str, int] | None = None, mode: str = "source") -> None:
    """Carga cantidades al checklist antes de renderizar sus widgets."""
    source = source or {}
    for row in rows:
        name = str(row.get("carta", ""))
        needed = max(int(row.get("cantidad", 0) or 0), 0)
        if mode == "complete":
            qty = needed
        elif mode == "empty":
            qty = 0
        else:
            qty = min(needed, max(int(source.get(_meta_norm(name), 0) or 0), 0))
        st.session_state[_meta_manual_owned_key(deck, name)] = qty
    st.session_state.setdefault("meta_missing_estimates", {}).pop(str(deck.get("id", "deck")), None)


def _meta_inventory_signature(deck: dict, rows: list[dict], inventory: dict[str, int]) -> str:
    values = [f"{_meta_norm(row.get('carta',''))}:{int(inventory.get(_meta_norm(row.get('carta','')),0) or 0)}" for row in rows]
    return f"{deck.get('id','deck')}|" + "|".join(values)


def _meta_reference_unit_price(row: dict, visual: dict | None, clp_rate: float) -> tuple[int, int]:
    """Devuelve precio unitario estimado y cantidad de publicaciones coincidentes."""
    visual = visual or {}
    matches = _meta_market_matches(row.get("carta", ""))
    listing_prices = [int(x.get("price", 0) or 0) for x in matches if int(x.get("price", 0) or 0) > 0]
    api_price = 0
    if visual.get("price_usd") and clp_rate:
        api_price = int(round(float(visual.get("price_usd") or 0) * float(clp_rate or 0) / 100.0) * 100)
    # Para el piloto, una publicación local tiene prioridad; si no existe usamos referencia API.
    unit_price = min(listing_prices) if listing_prices else api_price
    return unit_price, len(matches)


def _meta_estimate_missing_cost(deck: dict, rows: list[dict], inventory: dict[str, int], game: str,
                                api_key: str | None, clp_rate: float) -> dict:
    missing_rows = []
    for row in rows:
        state = _meta_card_state(row, inventory, deck)
        if state["missing"] > 0:
            missing_rows.append((row, state))
    _meta_prefetch_candidates([row.get("carta", "") for row, _ in missing_rows], game, api_key)
    total = 0
    priced_units = 0
    unknown_units = 0
    details = []
    for row, state in missing_rows:
        visual = _meta_selected_visual(deck, row.get("carta", ""), game, api_key)
        unit_price, matches = _meta_reference_unit_price(row, visual, clp_rate)
        subtotal = unit_price * int(state["missing"])
        if unit_price > 0:
            total += subtotal
            priced_units += int(state["missing"])
        else:
            unknown_units += int(state["missing"])
        details.append({
            "Carta": row.get("carta", ""), "Faltan": int(state["missing"]),
            "Precio unitario": unit_price, "Subtotal": subtotal,
            "Publicaciones": matches, "Versión": visual.get("set", "Por definir"),
        })
    return {
        "signature": _meta_inventory_signature(deck, rows, inventory),
        "total": int(total), "priced_units": priced_units, "unknown_units": unknown_units,
        "missing_units": sum(int(state["missing"]) for _, state in missing_rows),
        "details": details, "updated_at": time.strftime("%Y-%m-%d %H:%M"),
    }


def _meta_add_to_cart(deck: dict, row: dict, qty: int, visual: dict | None, clp_rate: float) -> None:
    qty = max(int(qty or 0), 0)
    if qty <= 0:
        return
    visual = visual or {}
    unit_price, match_count = _meta_reference_unit_price(row, visual, clp_rate)
    matches = _meta_market_matches(row.get("carta", ""))
    key = _meta_cart_key(deck, row.get("carta", ""))
    st.session_state.setdefault("deck_cart", {})[key] = {
        "deck_id": deck.get("id"), "deck_name": deck.get("nombre"),
        "card_name": row.get("carta", ""), "category": row.get("categoria", "—"),
        "qty": qty, "image_url": visual.get("image_url", ""), "card_id": visual.get("card_id", ""),
        "set": visual.get("set", "Versión por definir"), "number": visual.get("number", "—"),
        "regulation": visual.get("regulation", "—"), "unit_price": unit_price,
        "market_matches": match_count,
    }
    _track_event("agregar_carta_carrito_mazo", str(row.get("carta", "")), f"{deck.get('nombre')}|{qty}")


def _meta_remove_from_cart(deck: dict, card_name: str) -> None:
    st.session_state.setdefault("deck_cart", {}).pop(_meta_cart_key(deck, card_name), None)
    _track_event("quitar_carta_carrito_mazo", card_name, deck.get("nombre", ""))


def _meta_card_state(row: dict, inventory: dict[str, int], deck: dict) -> dict:
    needed = max(int(row.get("cantidad", 0) or 0), 0)
    owned = max(int(inventory.get(_meta_norm(row.get("carta", "")), 0) or 0), 0)
    used = min(needed, owned)
    missing = max(needed - used, 0)
    in_cart = st.session_state.get("deck_cart", {}).get(_meta_cart_key(deck, row.get("carta", "")))
    status = "cart" if in_cart else "complete" if missing == 0 else "partial" if owned > 0 else "missing"
    label = "En carrito" if in_cart else "Completa" if missing == 0 else f"Faltan {missing}" if owned == 0 else f"Tienes {used} · faltan {missing}"
    return {"needed": needed, "owned": owned, "used": used, "missing": missing, "in_cart": in_cart, "status": status, "label": label}


def _meta_render_selected_card(deck: dict, row: dict, game: str, api_key: str | None, clp_rate: float, inventory: dict[str, int], manual_mode: bool = False) -> None:
    name = str(row.get("carta", ""))
    state = _meta_card_state(row, inventory, deck)
    candidates = _meta_card_candidates(name, game, api_key)
    st.markdown('<div class="deck-detail-panel">', unsafe_allow_html=True)
    top1, top2 = st.columns([5, 1])
    top1.markdown(f"### {name}")
    if top2.button("Cerrar", key=f"close_meta_card_{_meta_norm(name)}", use_container_width=True):
        st.session_state["meta_selected_card"] = None; st.rerun()
    if candidates:
        saved = st.session_state.get("meta_print_choices", {}).get(_meta_print_key(deck, name))
        labels = [f"{x['set']} · #{x['number']} · {x['regulation']} · {x['rarity']}" for x in candidates]
        default = next((i for i, x in enumerate(candidates) if x.get("card_id") == saved), 0)
        selected_label = st.selectbox("Impresión / versión", labels, index=default, key=f"print_select_{deck.get('id')}_{_meta_norm(name)}")
        visual = candidates[labels.index(selected_label)]
        st.session_state.setdefault("meta_print_choices", {})[_meta_print_key(deck, name)] = visual.get("card_id")
    else:
        visual = {}
    left, right = st.columns([1, 1.55], gap="large")
    with left:
        image = _safe_image_url(visual.get("image_url", ""))
        if image:
            st.markdown(f'<div class="deck-detail-image"><img src="{escape(image)}" alt="{escape(name)}"></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="deck-card-placeholder" style="border:2px solid var(--ng-ink);border-radius:16px"><span>🃏</span><b>{escape(name)}</b><small>Imagen pendiente</small></div>', unsafe_allow_html=True)
    with right:
        a, b, c = st.columns(3)
        a.metric("Necesitas", state["needed"]); b.metric("En colección", state["owned"]); c.metric("Faltan", state["missing"])
        if manual_mode:
            st.markdown("**Checklist manual**")
            manual_key = _meta_manual_owned_key(deck, name)
            q0, q1, q2 = st.columns([1, 2.2, 1.4])
            q0.button("0", key=f"detail_zero_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True,
                      on_click=_meta_set_manual_owned, args=(manual_key, 0))
            q1.number_input("Copias que ya tienes", min_value=0, max_value=max(state["needed"], 0), step=1,
                            key=manual_key, label_visibility="collapsed")
            q2.button(f"✓ {state['needed']}", key=f"detail_all_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True,
                      on_click=_meta_set_manual_owned, args=(manual_key, state["needed"]))
            st.caption("Marca 0, una cantidad intermedia o todas las copias necesarias.")
        if visual:
            price_clp = int(round(float(visual.get("price_usd") or 0) * float(clp_rate or 0) / 100.0) * 100) if visual.get("price_usd") and clp_rate else 0
            st.markdown(f"**Set:** {visual.get('set')} · **N.º:** {visual.get('number')} · **Marca:** {visual.get('regulation')}  ")
            st.caption(f"Rareza: {visual.get('rarity')} · Variante de precio: {visual.get('variant')} · Actualización: {visual.get('updated')}")
            if price_clp:
                st.markdown(f"**Referencia aproximada:** {_fmt_clp(price_clp)} por copia")
        matches = _meta_market_matches(name)
        if matches:
            prices = [int(x.get("price", 0) or 0) for x in matches if int(x.get("price", 0) or 0) > 0]
            st.success(f"{len(matches)} publicación(es) coincidente(s) en la demo" + (f" · desde {_fmt_clp(min(prices))}" if prices else ""))
        else:
            st.info("Sin publicaciones coincidentes en el marketplace demo. Puedes agregarla igualmente como necesidad del mazo.")
        if state["missing"] > 0:
            qty = st.number_input("Cantidad a agregar", min_value=1, max_value=max(state["missing"], 1), value=max(state["missing"], 1), step=1, key=f"detail_qty_{deck.get('id')}_{_meta_norm(name)}")
            x1, x2 = st.columns(2)
            if x1.button("Agregar al carrito del mazo", type="primary", key=f"detail_add_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True):
                _meta_add_to_cart(deck, row, qty, visual, clp_rate); _notify(f"Agregaste {qty} × {name} al carrito del mazo.", "success"); st.rerun()
            if x2.button("Buscar en Marketplace", key=f"detail_market_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True):
                _meta_search_marketplace(name, deck.get("nombre", ""))
        else:
            st.success("Tu colección ya cubre la cantidad necesaria de esta carta.")
    st.markdown('</div>', unsafe_allow_html=True)


def _meta_render_cart_panel(deck: dict) -> None:
    items = _meta_cart_items(deck)
    units = sum(int(x.get("qty", 0)) for x in items)
    priced = [x for x in items if int(x.get("unit_price", 0) or 0) > 0]
    total = sum(int(x.get("unit_price", 0) or 0) * int(x.get("qty", 0)) for x in items)
    matched = sum(1 for x in items if int(x.get("market_matches", 0) or 0) > 0)
    st.markdown(
        f'<div class="deck-cart-box"><h4>Carrito del mazo</h4><div class="deck-cart-total">{units} carta(s)</div>'
        f'<p>{len(items)} nombres únicos · {matched} con publicaciones coincidentes</p><p>Estimación conocida: <b>{_fmt_clp(total)}</b></p>'
        f'<p>{len(items)-len(priced)} línea(s) todavía sin precio.</p></div>', unsafe_allow_html=True)
    if not items:
        st.caption("Agrega faltantes desde la galería o usa “Agregar todas las faltantes”.")
        return
    with st.expander("Revisar contenido", expanded=False):
        for item in items:
            c1, c2, c3 = st.columns([4, 1.3, .8])
            c1.markdown(f"**{item.get('qty')} × {item.get('card_name')}**  \n{item.get('set','Versión por definir')}")
            c2.caption(_fmt_clp(int(item.get("unit_price", 0) or 0) * int(item.get("qty", 0))) if item.get("unit_price") else "Sin precio")
            if c3.button("Quitar", key=f"cart_remove_{deck.get('id')}_{_meta_norm(item.get('card_name'))}", use_container_width=True):
                _meta_remove_from_cart(deck, item.get("card_name", "")); st.rerun()
        if st.button("Vaciar carrito de este mazo", key=f"empty_deck_cart_{deck.get('id')}", use_container_width=True):
            prefix = f"{deck.get('id')}::"
            st.session_state["deck_cart"] = {k:v for k,v in st.session_state.get("deck_cart",{}).items() if not str(k).startswith(prefix)}
            _track_event("vaciar_carrito_mazo", deck.get("id", ""), deck.get("nombre", "")); st.rerun()


def _meta_render_card_tile(deck: dict, row: dict, game: str, api_key: str | None, clp_rate: float, inventory: dict[str, int], prefix: str, manual_mode: bool = False) -> None:
    name = str(row.get("carta", "")); state = _meta_card_state(row, inventory, deck)
    visual = _meta_selected_visual(deck, name, game, api_key)
    image = _safe_image_url(visual.get("image_url", ""))
    if image:
        art = f'<img src="{escape(image)}" alt="{escape(name)}">'
    else:
        icon = "⚡" if "energ" in _meta_norm(row.get("categoria", "")) else "🧩" if game != "Pokémon TCG" else "🃏"
        art = f'<div class="deck-card-placeholder"><span>{icon}</span><b>{escape(name)}</b><small>Vista visual pendiente</small></div>'
    price = 0
    if visual.get("price_usd") and clp_rate:
        price = int(round(float(visual.get("price_usd")) * float(clp_rate) / 100.0) * 100)
    matches = _meta_market_matches(name)
    availability = f"{len(matches)} publicación(es)" if matches else "Sin publicaciones"
    click_url = f"?deck_card={quote(name)}&deck_id={quote(str(deck.get('id','')))}"
    st.markdown(
        f'<div class="deck-card-art"><a class="deck-card-click" href="{escape(click_url)}" target="_self">{art}</a><span class="deck-owned-badge">{escape(str(row.get("categoria","—")))}</span><span class="deck-qty-badge">×{state["needed"]}</span></div>'
        f'<div class="deck-card-meta"><h4>{escape(name)}</h4><p>{escape(str(visual.get("set","Versión por definir")))}{(" · #"+escape(str(visual.get("number")))) if visual.get("number") else ""}</p>'
        f'<span class="deck-status {state["status"]}">{escape(state["label"])}</span><div class="deck-availability"><span>{availability}</span>'
        f'{f"<span>{_fmt_clp(price)} c/u</span>" if price else ""}'
        f'{f"<span>Faltante: {_fmt_clp(price * state["missing"])}</span>" if price and state["missing"] else ""}</div></div>', unsafe_allow_html=True)
    if manual_mode:
        manual_key = _meta_manual_owned_key(deck, name)
        st.caption(f"Tengo de {state['needed']} copias")
        m0, m1, m2 = st.columns([.75, 1.6, 1.05])
        m0.button("0", key=f"{prefix}_zero_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True,
                  on_click=_meta_set_manual_owned, args=(manual_key, 0))
        m1.number_input("Tengo", min_value=0, max_value=max(state["needed"], 0), step=1,
                        key=manual_key, label_visibility="collapsed")
        m2.button("✓ Todas", key=f"{prefix}_all_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True,
                  on_click=_meta_set_manual_owned, args=(manual_key, state["needed"]))
    b1, b2 = st.columns(2)
    if b1.button("Ver carta", key=f"{prefix}_view_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True):
        st.session_state["meta_selected_card"] = name; _track_event("abrir_carta_mazo", name, deck.get("nombre", "")); st.rerun()
    if state["in_cart"]:
        if b2.button("Quitar", key=f"{prefix}_remove_{deck.get('id')}_{_meta_norm(name)}", use_container_width=True):
            _meta_remove_from_cart(deck, name); st.rerun()
    elif state["missing"] > 0:
        if b2.button(f"+{state['missing']} carrito", key=f"{prefix}_add_{deck.get('id')}_{_meta_norm(name)}", type="primary", use_container_width=True):
            _meta_add_to_cart(deck, row, state["missing"], visual, clp_rate); _notify(f"Agregaste {state['missing']} × {name}.", "success"); st.rerun()
    else:
        b2.button("Completa", key=f"{prefix}_complete_{deck.get('id')}_{_meta_norm(name)}", disabled=True, use_container_width=True)


def _render_visual_decklist(deck: dict, game: str, api_key: str | None, clp_rate: float) -> None:
    rows = [dict(x) for x in deck.get("core", []) if _meta_is_real_card_row(x)]
    if not rows:
        _render_empty_state("🗂️", "La fuente no incluye decklist", "Puedes cargarla en la columna decklist del archivo administrado.")
        return
    actual = _inventory_from_results()
    mode_key = f"collection_mode_{deck.get('id')}"
    collection_mode = st.radio(
        "¿Cómo quieres indicar las cartas que ya tienes?",
        ["Checklist manual", "Inventario del tasador", "Colección de ejemplo"],
        horizontal=True, key=mode_key,
        help="El checklist manual permite marcar copias directamente mientras revisas el mazo.",
    )
    if collection_mode == "Checklist manual":
        inventory = _meta_manual_inventory(deck, rows)
        bulk1, bulk2, bulk3 = st.columns(3)
        if bulk1.button("Vaciar checklist", key=f"manual_empty_{deck.get('id')}", use_container_width=True):
            _meta_fill_manual_inventory(deck, rows, mode="empty"); st.rerun()
        if bulk2.button("Marcar mazo completo", key=f"manual_complete_{deck.get('id')}", use_container_width=True):
            _meta_fill_manual_inventory(deck, rows, mode="complete"); st.rerun()
        source_label = "Copiar inventario tasado" if actual else "Cargar colección de ejemplo"
        source_inventory = actual if actual else {_meta_norm(n): int(q) for n, q in META_DEMO_COLLECTION.items()}
        if bulk3.button(source_label, key=f"manual_source_{deck.get('id')}", use_container_width=True):
            _meta_fill_manual_inventory(deck, rows, source_inventory, mode="source"); st.rerun()
        st.caption("Ajusta la cantidad de cada carta con 0, el selector numérico o ✓ Todas. El progreso y el costo se recalculan con tu checklist.")
    elif collection_mode == "Inventario del tasador":
        inventory = dict(actual)
        if not inventory:
            st.info("Todavía no hay cartas procesadas en el Tasador. Puedes cambiar a Checklist manual para marcar este mazo ahora mismo.")
    else:
        inventory = {_meta_norm(n): int(q) for n, q in META_DEMO_COLLECTION.items()}
        st.caption("Colección demostrativa activa para enseñar el flujo del piloto.")
    pct, missing, owned, required = _deck_completion(deck, inventory)
    miss_units = sum(int(x.get("Faltan", 0)) for x in missing)
    cart_items = _meta_cart_items(deck)
    signature = _meta_inventory_signature(deck, rows, inventory)
    estimate = st.session_state.setdefault("meta_missing_estimates", {}).get(str(deck.get("id", "deck")), {})
    estimate_current = bool(estimate) and estimate.get("signature") == signature
    estimate_text = _fmt_clp(int(estimate.get("total", 0) or 0)) if estimate_current else "Sin calcular"
    top_left, top_right = st.columns([2.1, 1])
    with top_left:
        st.markdown(
            f'<div class="deck-builder-shell"><div class="deck-builder-title"><div><h3>Constructor visual del mazo</h3>'
            f'<p>Marca las copias que ya tienes, calcula cuánto te falta y prepara solo las cartas necesarias.</p></div><span class="meta-tier meta-tier-{str(deck.get("tier","B")).lower()}">TIER {escape(str(deck.get("tier","B")))}</span></div>'
            f'<div class="deck-progress-large"><span style="width:{pct}%"></span></div><div class="deck-availability"><span>{pct:.1f}% completo</span><span>{owned}/{required} copias cubiertas</span><span>{miss_units} faltantes</span><span>{sum(int(x.get("qty",0)) for x in cart_items)} en carrito</span><span>Costo faltante: {estimate_text}</span></div></div>',
            unsafe_allow_html=True)
        action1, action2, action3, action4 = st.columns(4)
        if action1.button("Sincronizar faltantes", type="primary", key=f"add_all_missing_{deck.get('id')}", use_container_width=True, disabled=not bool(missing)):
            prefix = f"{deck.get('id')}::"
            st.session_state["deck_cart"] = {k:v for k,v in st.session_state.get("deck_cart",{}).items() if not str(k).startswith(prefix)}
            missing_by_name = {_meta_norm(x["Carta"]): x for x in missing}
            for row in rows:
                miss = missing_by_name.get(_meta_norm(row.get("carta", "")))
                if miss:
                    visual = _meta_selected_visual(deck, row.get("carta", ""), game, api_key)
                    _meta_add_to_cart(deck, row, int(miss["Faltan"]), visual, clp_rate)
            _notify(f"El carrito quedó sincronizado con tus {miss_units} cartas faltantes.", "success")
            _track_event("sincronizar_faltantes", deck.get("id", ""), str(miss_units)); st.rerun()
        if action2.button("Calcular costo", key=f"estimate_missing_{deck.get('id')}", use_container_width=True, disabled=not bool(missing)):
            with st.spinner("Consultando precios de referencia para las cartas faltantes..."):
                result = _meta_estimate_missing_cost(deck, rows, inventory, game, api_key, clp_rate)
                st.session_state.setdefault("meta_missing_estimates", {})[str(deck.get("id", "deck"))] = result
            _track_event("calcular_costo_faltante", deck.get("id", ""), str(result.get("total", 0))); st.rerun()
        deck_text = _meta_decklist_text(deck)
        action3.download_button("Descargar lista", deck_text.encode("utf-8"), f"{_meta_norm(deck.get('nombre')).replace(' ','_')}.txt", "text/plain", use_container_width=True)
        if action4.button("Buscar faltante", key=f"search_first_missing_{deck.get('id')}", use_container_width=True, disabled=not bool(missing)):
            _meta_search_marketplace(missing[0]["Carta"], deck.get("nombre", ""))
        if estimate_current:
            known = int(estimate.get("priced_units", 0) or 0)
            unknown = int(estimate.get("unknown_units", 0) or 0)
            if unknown:
                st.warning(f"Estimación parcial: {_fmt_clp(estimate.get('total',0))} para {known} copia(s). Quedan {unknown} copia(s) sin precio disponible.")
            else:
                st.success(f"Completar las {estimate.get('missing_units',0)} copias faltantes se estima en {_fmt_clp(estimate.get('total',0))}.")
            with st.expander("Ver desglose de costo", expanded=False):
                detail_df = pd.DataFrame(estimate.get("details", []))
                if not detail_df.empty:
                    st.dataframe(detail_df, use_container_width=True, hide_index=True,
                                 column_config={"Precio unitario": st.column_config.NumberColumn(format="$%d"), "Subtotal": st.column_config.NumberColumn(format="$%d")})
                st.caption(f"Último cálculo: {estimate.get('updated_at','—')} · precios referenciales, no una cotización garantizada.")
        elif estimate:
            st.info("El checklist cambió desde el último cálculo. Presiona “Calcular costo” para actualizar el estimado.")
    with top_right:
        _meta_render_cart_panel(deck)

    try:
        query_card = st.query_params.get("deck_card")
        query_deck = st.query_params.get("deck_id")
        if query_card and str(query_deck or "") == str(deck.get("id", "")):
            st.session_state["meta_selected_card"] = str(query_card)
            st.query_params.clear()
    except Exception:
        pass
    selected_name = st.session_state.get("meta_selected_card")
    selected_row = next((x for x in rows if _meta_norm(x.get("carta")) == _meta_norm(selected_name)), None)
    if selected_row:
        _meta_render_selected_card(deck, selected_row, game, api_key, clp_rate, inventory, collection_mode == "Checklist manual")

    st.markdown("#### Explorar cartas")
    f1, f2, f3, f4 = st.columns([1.2, 1.6, 1, 1])
    view = f1.radio("Vista", ["Galería visual", "Checklist rápido", "Lista compacta", "Análisis"], horizontal=True, key=f"deck_view_{deck.get('id')}")
    categories = list(dict.fromkeys([str(x.get("categoria", "Otros")) for x in rows]))
    selected_categories = f2.multiselect("Categorías", categories, default=categories, key=f"deck_categories_{deck.get('id')}")
    only_missing = f3.toggle("Solo faltantes", value=st.session_state.get("meta_only_missing", False), key=f"only_missing_{deck.get('id')}")
    only_cart = f4.toggle("Solo carrito", value=False, key=f"only_cart_{deck.get('id')}")
    st.session_state["meta_only_missing"] = only_missing
    filtered = []
    for row in rows:
        state = _meta_card_state(row, inventory, deck)
        if row.get("categoria", "Otros") not in selected_categories:
            continue
        if only_missing and state["missing"] <= 0:
            continue
        if only_cart and not state["in_cart"]:
            continue
        filtered.append(row)
    if not filtered:
        _render_empty_state("✨", "No hay cartas con esos filtros", "Activa más categorías o desmarca “Solo faltantes / Solo carrito”.")
        return

    if view == "Checklist rápido":
        quick_rows = []
        for row in filtered:
            state = _meta_card_state(row, inventory, deck)
            quick_rows.append({
                "Carta": row.get("carta"), "Categoría": row.get("categoria"),
                "Necesitas": state["needed"], "Tengo": state["owned"],
                "Faltan": state["missing"], "Completa": state["missing"] == 0,
            })
        quick_df = pd.DataFrame(quick_rows)
        if collection_mode != "Checklist manual":
            st.info("La vista es editable cuando seleccionas “Checklist manual” en la parte superior.")
            st.dataframe(quick_df, use_container_width=True, hide_index=True)
        else:
            edited = st.data_editor(
                quick_df, use_container_width=True, hide_index=True, num_rows="fixed",
                disabled=["Carta", "Categoría", "Necesitas", "Faltan", "Completa"],
                column_config={
                    "Tengo": st.column_config.NumberColumn("Copias que tengo", min_value=0, step=1),
                    "Completa": st.column_config.CheckboxColumn("Lista"),
                }, key=f"quick_checklist_{deck.get('id')}"
            )
            changed = False
            for _, qrow in edited.iterrows():
                source_row = next((x for x in rows if _meta_norm(x.get("carta")) == _meta_norm(qrow.get("Carta"))), None)
                if not source_row:
                    continue
                needed = max(int(source_row.get("cantidad", 0) or 0), 0)
                try:
                    qty = max(0, min(int(qrow.get("Tengo", 0) or 0), needed))
                except Exception:
                    qty = 0
                key = _meta_manual_owned_key(deck, source_row.get("carta", ""))
                if int(st.session_state.get(key, 0) or 0) != qty:
                    st.session_state[key] = qty; changed = True
            if changed:
                st.session_state.setdefault("meta_missing_estimates", {}).pop(str(deck.get("id", "deck")), None)
                st.rerun()
            st.caption("Edita la columna “Copias que tengo”. Faltantes, progreso y costo se actualizan automáticamente.")
        return
    if view == "Lista compacta":
        with st.spinner("Preparando versiones de las cartas..."):
            _meta_prefetch_candidates([x.get("carta", "") for x in filtered], game, api_key)
        compact = []
        for row in filtered:
            state = _meta_card_state(row, inventory, deck)
            visual = _meta_selected_visual(deck, row.get("carta", ""), game, api_key)
            compact.append({
                "Cantidad": state["needed"], "Carta": row.get("carta"), "Categoría": row.get("categoria"),
                "En colección": state["owned"], "Faltan": state["missing"], "En carrito": "Sí" if state["in_cart"] else "No",
                "Versión": visual.get("set", "Por definir"), "Reg.": visual.get("regulation", "—"),
            })
        st.dataframe(pd.DataFrame(compact), use_container_width=True, hide_index=True)
        st.caption(deck.get("lista_fuente", "Lista representativa para validar el módulo."))
        return
    if view == "Análisis":
        by_category = pd.DataFrame(filtered).groupby("categoria", as_index=False)["cantidad"].sum().rename(columns={"categoria":"Categoría", "cantidad":"Copias"})
        a1, a2 = st.columns(2)
        with a1:
            st.markdown("##### Composición"); st.bar_chart(by_category.set_index("Categoría"))
        with a2:
            analysis_rows = []
            for row in filtered:
                state = _meta_card_state(row, inventory, deck)
                analysis_rows.append({"Carta":row.get("carta"),"Necesitas":state["needed"],"Tienes":state["owned"],"Faltan":state["missing"]})
            st.markdown("##### Brecha de colección"); st.dataframe(pd.DataFrame(analysis_rows),use_container_width=True,hide_index=True)
        return

    page_size = 12 if api_key or _DB_LOADED else 8
    pages = max(1, (len(filtered) + page_size - 1) // page_size)
    page = st.selectbox("Página", list(range(1, pages + 1)), format_func=lambda x: f"{x} de {pages}", key=f"deck_gallery_page_{deck.get('id')}") if pages > 1 else 1
    visible = filtered[(page - 1) * page_size: page * page_size]
    with st.spinner("Preparando imágenes del mazo..." if game == "Pokémon TCG" else "Preparando galería..."):
        _meta_prefetch_candidates([x.get("carta", "") for x in visible], game, api_key)
    columns_count = 4
    for start in range(0, len(visible), columns_count):
        cols = st.columns(columns_count, gap="medium")
        for col, row in zip(cols, visible[start:start + columns_count]):
            with col:
                with st.container(border=True):
                    _meta_render_card_tile(deck, row, game, api_key, clp_rate, inventory, f"deck_gallery_{start}", collection_mode == "Checklist manual")
    st.caption(deck.get("lista_fuente", "Lista representativa para validar el módulo."))


def _render_meta_rank_card(deck: dict, rank: int, key_prefix: str) -> None:
    tier = str(deck.get("tier","B")).lower()
    best = escape(str(deck.get("mejor","—")))
    st.markdown(
        f'<div class="meta-rank-card"><div style="display:flex;justify-content:space-between;align-items:center">'
        f'<span class="meta-rank">{rank}</span><span class="meta-tier meta-tier-{tier}">TIER {escape(str(deck.get("tier","B")))}</span></div>'
        f'<h4>{escape(str(deck.get("nombre","Mazo")))}</h4><p>{escape(str(deck.get("descripcion","")))}</p>'
        f'<div class="meta-stat-line"><span><b>{deck.get("uso",0):.1f}%</b> meta</span><span><b>{deck.get("win_rate",0):.1f}%</b> WR</span><span title="Mejor resultado">🏆 {best}</span></div></div>',
        unsafe_allow_html=True,
    )
    if st.button("Abrir análisis", key=f"{key_prefix}_{deck.get('id')}", use_container_width=True):
        st.session_state["selected_meta_deck"] = deck.get("id")
        _track_event("abrir_mazo_meta", deck.get("id",""), deck.get("nombre",""))
        st.rerun()


def _render_meta_deck_detail(game: str, fmt: str, dataset: dict, source: str, deck: dict, api_key: str | None = None, clp_rate: float = 950) -> None:
    if st.button("← Volver a Meta Lab", key="meta_back_to_lab"):
        st.session_state["selected_meta_deck"] = None
        st.session_state["meta_selected_card"] = None
        st.rerun()
    st.markdown(
        f'<div class="meta-deck-head"><span class="meta-tier meta-tier-{str(deck.get("tier","B")).lower()}">TIER {escape(str(deck.get("tier","B")))}</span>'
        f'<h2>{escape(str(deck.get("nombre","Mazo")))}</h2><p>{escape(str(deck.get("descripcion","")))}</p>'
        f'<div style="margin-top:10px"><span class="meta-card-pill">{escape(game)}</span><span class="meta-card-pill">{escape(fmt)}</span>'
        f'<span class="meta-card-pill">{escape(source)}</span></div></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Presencia", f"{deck.get('uso',0):.1f}%")
    m2.metric("Win rate", f"{deck.get('win_rate',0):.1f}%")
    conversion_label = dataset.get("conversion_label", "Conversión Top Cut")
    m3.metric(conversion_label, f"{deck.get('top_cut',0):.1f}%")
    m4.metric("Entradas", int(deck.get("entradas",0)))
    m5.metric("Costo estimado", _fmt_clp(deck.get("costo",0)) if deck.get("costo",0) else "Sin datos")
    tabs = st.tabs(["Constructor visual", "Rendimiento", "Matchups", "Resultados", "Completar mazo"])
    with tabs[0]:
        _render_visual_decklist(deck, game, api_key, clp_rate)
    with tabs[1]:
        perf = pd.DataFrame({
            "Indicador":["Presencia en meta", "Win rate", dataset.get("conversion_label", "Conversión Top Cut")],
            "Porcentaje":[deck.get("uso",0), deck.get("win_rate",0), deck.get("top_cut",0)]
        }).set_index("Indicador")
        st.bar_chart(perf)
        st.markdown(f"**Mejor resultado registrado:** {deck.get('mejor','—')}  ")
        st.markdown(f"**Muestra:** {int(deck.get('entradas',0))} jugadores del evento de referencia.")
    with tabs[2]:
        others = [d for d in dataset.get("mazos",[]) if d.get("id") != deck.get("id")]
        matchups = sorted([{"Mazo":d.get("nombre"), "Probabilidad":_meta_matchup_value(deck,d)} for d in others], key=lambda x:x["Probabilidad"], reverse=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Enfrentamientos favorables")
            for row in matchups[:4]:
                st.markdown(f'<p class="meta-match-good">{row["Probabilidad"]:.1f}% · {escape(str(row["Mazo"]))}</p>', unsafe_allow_html=True)
        with c2:
            st.markdown("#### Enfrentamientos difíciles")
            for row in matchups[-4:][::-1]:
                st.markdown(f'<p class="meta-match-bad">{row["Probabilidad"]:.1f}% · {escape(str(row["Mazo"]))}</p>', unsafe_allow_html=True)
        st.caption("En datos demo la matriz es ilustrativa. Con resultados cargados puede reemplazarse por estadísticas reales por ronda.")
    with tabs[3]:
        events = [e for e in dataset.get("torneos",[]) if deck.get("nombre") in e.get("top8",[]) or e.get("mazo") == deck.get("nombre")]
        if not events:
            _render_empty_state("🏆", "Sin resultados asociados", "La fuente actual no contiene eventos para este arquetipo.")
        for event in events:
            st.markdown(f"**{event.get('evento')}** · {event.get('fecha')} · {event.get('jugadores')} participantes  ")
            st.caption(f"Ganador: {event.get('ganador')} · {event.get('mazo')}")
    with tabs[4]:
        inventory = _inventory_from_results()
        if not inventory and st.session_state.get("meta_use_demo_collection", True):
            inventory = {_meta_norm(k):v for k,v in META_DEMO_COLLECTION.items()}
        pct, missing, owned, required = _deck_completion(deck, inventory)
        st.markdown(f"### Completitud: {pct:.1f}%")
        st.markdown(f'<div class="meta-progress"><span style="width:{pct}%"></span></div>', unsafe_allow_html=True)
        st.caption(f"{owned} de {required} copias clave detectadas en tu inventario.")
        if missing:
            st.dataframe(pd.DataFrame(missing), use_container_width=True, hide_index=True)
            c1, c2 = st.columns(2)
            first = missing[0]["Carta"]
            if c1.button(f"Buscar {first} en Marketplace", type="primary", key="meta_detail_missing_market", use_container_width=True):
                _meta_search_marketplace(first, deck.get("nombre", ""))
            if c2.button("Abrir constructor visual", key="meta_detail_open_builder", use_container_width=True):
                st.session_state["meta_selected_card"] = first
                st.rerun()
        else:
            st.success("Tienes todas las cartas clave consideradas en esta lista.")


def _render_meta_panorama(game: str, fmt: str, dataset: dict, source: str) -> None:
    decks=dataset.get("mazos",[]);events=dataset.get("torneos",[])
    total_entries=sum(int(d.get("entradas",0)) for d in decks)
    p1,p2,p3,p4=st.columns(4)
    p1.metric("Mazos analizados",len(decks));p2.metric("Entradas",total_entries);p3.metric("Eventos",len(events));p4.metric("Formato",fmt)
    source_text = f'{source} · {dataset.get("actualizado","")}'
    if dataset.get("legalidad"):
        source_text += f' · {dataset.get("legalidad")}'
    st.markdown(f'<span class="meta-source"><i></i>{escape(source_text)}</span>',unsafe_allow_html=True)
    if dataset.get("source_url"):
        st.link_button("Abrir fuente de referencia", dataset.get("source_url"), use_container_width=False)
    st.markdown("<br>",unsafe_allow_html=True)
    c1,c2=st.columns([1.35,1])
    with c1:
        st.markdown("#### Top 10 del metagame")
        rows=[]
        conversion_col = dataset.get("conversion_label", "Top Cut %")
        for i,d in enumerate(decks[:10],1):
            rows.append({"#":i,"Mazo":d.get("nombre"),"Tier":d.get("tier"),"Meta %":d.get("uso"),"Win %":d.get("win_rate"),conversion_col:d.get("top_cut"),"Muestra":d.get("entradas"),"Mejor resultado":d.get("mejor")})
        st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    with c2:
        st.markdown("#### Distribución")
        chart=pd.DataFrame({"Mazo":[d.get("nombre") for d in decks[:8]],"Meta":[d.get("uso",0) for d in decks[:8]]}).set_index("Mazo")
        st.bar_chart(chart)
    st.markdown("#### Mazos que lideran la conversación")
    cols=st.columns(min(4,len(decks[:4])))
    for i,(col,deck) in enumerate(zip(cols,decks[:4]),1):
        with col:_render_meta_rank_card(deck,i,"meta_panorama")
    if events:
        event=events[0]
        st.markdown("<br>",unsafe_allow_html=True)
        st.markdown(f'<div class="meta-event"><span class="meta-beta">ÚLTIMO EVENTO</span><h4>{escape(str(event.get("evento")))}</h4><p>{escape(str(event.get("fecha")))} · {escape(str(event.get("region")))} · {event.get("jugadores",0)} participantes</p><div class="meta-winner">🏆 <strong>{escape(str(event.get("ganador")))}</strong> ganó con {escape(str(event.get("mazo")))}</div></div>',unsafe_allow_html=True)
        winner=_meta_find_deck(dataset,event.get("mazo"))
        if winner and st.button("Ver mazo ganador",key="meta_latest_winner",type="primary"):
            st.session_state["selected_meta_deck"]=winner.get("id");_track_event("ver_mazo_ganador",winner.get("id",""),event.get("evento",""));st.rerun()


def _render_meta_decks(dataset: dict) -> None:
    decks=dataset.get("mazos",[])
    f1,f2=st.columns([2,1]);query=f1.text_input("Buscar arquetipo",placeholder="Dragapult, Doflamingo, Izzet...",key="meta_deck_query");tier=f2.selectbox("Tier",["Todos","S","A","B"],key="meta_tier_filter")
    filtered=[d for d in decks if (not query or _meta_norm(query) in _meta_norm(d.get("nombre"))) and (tier=="Todos" or d.get("tier")==tier)]
    if not filtered:_render_empty_state("🔎","No encontramos ese mazo","Prueba otro nombre o elimina el filtro de tier.")
    for start in range(0,len(filtered),3):
        cols=st.columns(3)
        for rank,(col,deck) in enumerate(zip(cols,filtered[start:start+3]),start+1):
            with col:_render_meta_rank_card(deck,decks.index(deck)+1,"meta_decks")


def _render_meta_matchups(dataset: dict) -> None:
    decks=dataset.get("mazos",[])
    if len(decks)<2:_render_empty_state("⚔️","Datos insuficientes","Se necesitan al menos dos arquetipos.");return
    top=decks[:min(7,len(decks))]
    selected_name=st.selectbox("Analizar mazo",[d.get("nombre") for d in decks],key="meta_matchup_select")
    selected=_meta_find_deck(dataset,selected_name)
    matrix=[]
    for a in top:
        row={"Mazo":a.get("nombre")}
        for b in top:row[b.get("nombre")]=f"{_meta_matchup_value(a,b):.1f}%"
        matrix.append(row)
    st.dataframe(pd.DataFrame(matrix).set_index("Mazo"),use_container_width=True)
    if selected:
        values=sorted([(d.get("nombre"),_meta_matchup_value(selected,d)) for d in decks if d.get("id")!=selected.get("id")],key=lambda x:x[1],reverse=True)
        c1,c2=st.columns(2)
        with c1:
            st.markdown("#### Mejores matchups")
            for name,val in values[:4]:st.success(f"{val:.1f}% contra {name}")
        with c2:
            st.markdown("#### Matchups a preparar")
            for name,val in values[-4:][::-1]:st.error(f"{val:.1f}% contra {name}")
    st.caption("La matriz demo sirve para validar la interfaz. Al importar rondas con rival y resultado, el siguiente paso es calcular datos observados.")


def _render_meta_tournaments(dataset: dict) -> None:
    events=dataset.get("torneos",[])
    if not events:_render_empty_state("🏆","No hay eventos cargados","Importa resultados desde el panel anfitrión.");return
    for start in range(0,len(events),2):
        cols=st.columns(2)
        for col,event in zip(cols,events[start:start+2]):
            with col:
                st.markdown(f'<div class="meta-event"><span class="meta-beta">{escape(str(event.get("region","EVENTO"))).upper()}</span><h4>{escape(str(event.get("evento")))}</h4><p>{escape(str(event.get("fecha")))} · {event.get("jugadores",0)} participantes</p><div class="meta-winner">🏆 <strong>{escape(str(event.get("ganador")))}</strong><br>{escape(str(event.get("mazo")))}</div></div>',unsafe_allow_html=True)
                with st.expander("Ver Top 8"):
                    top8=event.get("top8",[]);st.dataframe(pd.DataFrame({"Posición":range(1,len(top8)+1),"Arquetipo":top8}),use_container_width=True,hide_index=True)
                winner=_meta_find_deck(dataset,event.get("mazo"))
                if winner and st.button("Abrir lista ganadora",key=f"event_winner_{event.get('id')}",use_container_width=True):
                    st.session_state["selected_meta_deck"]=winner.get("id");_track_event("ver_evento_meta",event.get("id",""),event.get("evento",""));st.rerun()


def _render_meta_collection(dataset: dict) -> None:
    actual=_inventory_from_results();has_actual=bool(actual)
    use_demo=st.toggle("Usar colección de ejemplo" if not has_actual else "Comparar también con colección de ejemplo",value=st.session_state.get("meta_use_demo_collection",not has_actual),key="meta_collection_demo_toggle")
    st.session_state["meta_use_demo_collection"]=use_demo
    inventory=dict(actual)
    if use_demo:
        for name,qty in META_DEMO_COLLECTION.items():inventory[_meta_norm(name)]=max(inventory.get(_meta_norm(name),0),qty)
    if not inventory:
        _render_empty_state("🗃️","Aún no hay inventario","Usa el Tasador para cargar tus cartas o activa la colección de ejemplo.")
        if st.button("Ir al Tasador",type="primary",key="meta_go_tasador"):_go_to("Tasador")
        return
    results=[]
    for deck in dataset.get("mazos",[]):
        pct,missing,owned,required=_deck_completion(deck,inventory)
        results.append({"deck":deck,"pct":pct,"missing":missing,"owned":owned,"required":required})
    results=sorted(results,key=lambda x:x["pct"],reverse=True)
    st.markdown("#### Mazos más cercanos a tu colección")
    for row in results[:6]:
        deck=row["deck"]
        with st.container(border=True):
            c1,c2,c3=st.columns([3,2,1])
            c1.markdown(f"**{deck.get('nombre')}**  \n{row['owned']} de {row['required']} cartas clave")
            c2.markdown(f'<div class="meta-progress"><span style="width:{row["pct"]}%"></span></div><small>{row["pct"]:.1f}% completo</small>',unsafe_allow_html=True)
            if c3.button("Ver faltantes",key=f"meta_collection_open_{deck.get('id')}",use_container_width=True):st.session_state["selected_meta_deck"]=deck.get("id");_track_event("comparar_coleccion_meta",deck.get("id",""),str(row["pct"]));st.rerun()
    st.caption("La comparación utiliza nombres y cantidades. Más adelante puede considerar edición, idioma, legalidad y condición.")


def _meta_template_bytes() -> bytes:
    df=pd.DataFrame([
        {"game":"Pokémon TCG","format":"Standard","event":"Regional ejemplo","date":"2026-06-15","player":"Jugador A","placement":1,"archetype":"Dragapult","opponent_archetype":"N's Zoroark","result":"W","players":512,"decklist":"4 Dreepy; 4 Drakloak; 3 Dragapult ex"},
        {"game":"Pokémon TCG","format":"Standard","event":"Regional ejemplo","date":"2026-06-15","player":"Jugador B","placement":2,"archetype":"N's Zoroark","opponent_archetype":"Dragapult","result":"L","players":512,"decklist":"4 N's Zorua; 4 N's Zoroark ex; 2 N's Zekrom"},
    ])
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as writer:df.to_excel(writer,index=False,sheet_name="Resultados")
    return buf.getvalue()


def _render_meta_admin() -> None:
    st.markdown('<div class="meta-admin"><strong>Carga administrada</strong><br>Reemplaza los datos demo con resultados propios sin modificar el código.</div>',unsafe_allow_html=True)
    st.download_button("Descargar plantilla Excel",_meta_template_bytes(),"plantilla_resultados_meta.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    upload=st.file_uploader("CSV o Excel de resultados",type=["csv","xlsx","xls"],key="meta_results_uploader")
    if upload is not None:
        try:
            df=pd.read_csv(upload) if upload.name.lower().endswith(".csv") else pd.read_excel(upload)
            st.dataframe(df.head(30),use_container_width=True,hide_index=True)
            normalized={_meta_norm(c).replace(" ","_") for c in df.columns}
            has_event=bool({"event","evento"}&normalized);has_player=bool({"player","jugador"}&normalized);has_deck=bool({"archetype","arquetipo","mazo"}&normalized)
            if not (has_event and has_player and has_deck):st.error("Faltan columnas equivalentes a evento, jugador y arquetipo/mazo.")
            elif st.button("Usar estos resultados",type="primary",key="meta_apply_upload",use_container_width=True):
                st.session_state["meta_uploaded_df"]=df.copy();st.session_state["meta_upload_name"]=upload.name;st.session_state["selected_meta_deck"]=None
                _track_event("cargar_resultados_meta",metadata=f"{upload.name}|{len(df)} filas");st.success("Resultados activados para Meta Lab.");st.rerun()
        except Exception as exc:st.error(f"No se pudo leer el archivo: {exc}")
    if isinstance(st.session_state.get("meta_uploaded_df"),pd.DataFrame):
        st.success(f"Fuente activa: {st.session_state.get('meta_upload_name')} · {len(st.session_state['meta_uploaded_df'])} filas")
        if st.button("Volver a datos demostrativos",key="meta_clear_upload",use_container_width=True):
            st.session_state["meta_uploaded_df"]=None;st.session_state["meta_upload_name"]="";st.session_state["selected_meta_deck"]=None;st.rerun()
    st.markdown("**Columnas admitidas:** juego/game, formato/format, evento/event, fecha/date, jugador/player, posición/placement, arquetipo/archetype, rival/opponent_archetype, resultado/result, participantes/players y decklist.")


def render_meta_lab(api_key: str | None = None, clp_rate: float = 950) -> None:
    _track_event("visita_meta_lab",once=True)
    st.markdown('<div class="meta-hero"><span class="meta-beta">META LAB · BETA</span><h1>Del resultado del torneo a tu próxima partida.</h1><p>Explora el metagame, abre listas ganadoras, estudia enfrentamientos y descubre qué mazo puedes completar con las cartas que ya tienes.</p></div>',unsafe_allow_html=True)
    games=list(META_DEMO.keys())
    c1,c2,c3=st.columns([1.35,1,1.5])
    game=c1.selectbox("Juego",games,index=games.index(st.session_state.get("meta_game","Pokémon TCG")) if st.session_state.get("meta_game") in games else 0,key="meta_game_selector")
    formats=META_DEMO[game].get("formatos",["Standard"])
    fmt=c2.selectbox("Formato",formats,index=0,key="meta_format_selector")
    st.session_state["meta_game"]=game;st.session_state["meta_format"]=fmt
    dataset,source=_meta_dataset(game,fmt)
    c3.markdown(f'<div class="compact-note"><b>Fuente activa:</b> {escape(source)}<br>{escape(str(dataset.get("actualizado","")))}</div>',unsafe_allow_html=True)
    if dataset.get("tipo_datos") == "snapshot_real":
        st.info("Snapshot competitivo real y fechado. No se actualiza automáticamente todavía; revisa la fecha y la fuente antes de usarlo para preparar un torneo.")
    else:
        st.warning("Los datos de este juego son demostrativos para validar el producto. El módulo acepta archivos administrados; la automatización con fuentes oficiales/API viene después.")
    selected=_meta_find_deck(dataset,st.session_state.get("selected_meta_deck"))
    if selected:
        _render_meta_deck_detail(game,fmt,dataset,source,selected,api_key,clp_rate);return
    tab_labels=["Panorama","Mazos","Matchups","Torneos","Mi colección"]
    if st.session_state.get("admin_unlocked"):tab_labels.append("Administrar datos")
    tabs=st.tabs(tab_labels)
    with tabs[0]:_render_meta_panorama(game,fmt,dataset,source)
    with tabs[1]:_render_meta_decks(dataset)
    with tabs[2]:_render_meta_matchups(dataset)
    with tabs[3]:_render_meta_tournaments(dataset)
    with tabs[4]:_render_meta_collection(dataset)
    if st.session_state.get("admin_unlocked"):
        with tabs[5]:_render_meta_admin()
    _render_feature_vote("meta_lab","¿Usarías Meta Lab para elegir, estudiar o completar un mazo competitivo?")


# Extiende el recorrido piloto con el uso de Meta Lab.
def _tour_status() -> tuple[int, int, list[tuple[str, bool]]]:
    actions=st.session_state.get("_journey_actions",set())
    milestones=[
        ("Explorar el marketplace","visita_marketplace" in actions),
        ("Abrir una ficha","abrir_ficha" in actions),
        ("Guardar un favorito","favorito" in actions),
        ("Explorar Meta Lab","visita_meta_lab" in actions),
        ("Abrir un mazo competitivo","abrir_mazo_meta" in actions),
        ("Simular una reserva","reserva" in actions),
        ("Visitar el tasador o publicar",bool({"visita_tasador","publicar"}&set(actions))),
    ]
    return sum(done for _,done in milestones),len(milestones),milestones


# Navegación V4 con Meta Lab integrado.
def _render_top_navigation() -> str:
    options=[("Inicio","Inicio"),("Marketplace","Mercado"),("Vender","Vender"),("Tasador","Tasador"),("Meta Lab","Meta Lab"),("Subastas","Subastas"),("Servicios","Servicios"),("Feedback","Feedback")]
    pending=st.session_state.pop("pending_nav",None);current=pending or st.session_state.get("page","Inicio")
    valid={x[0] for x in options}|{"Detalle"}
    if current not in valid:current="Inicio"
    active="Marketplace" if current=="Detalle" else current
    nav_cols=st.columns([2.05,.8,.9,.76,.78,.92,.88,.9,.88],gap="small")
    with nav_cols[0]:
        st.markdown("<div class='nexo-brand'><div class='nexo-brand-mark'>✦</div><div><div class='nexo-brand-name'>NexoGeek</div><div class='nexo-brand-tag'>colección · juego · comunidad</div></div></div>",unsafe_allow_html=True)
    for column,(page,label) in zip(nav_cols[1:],options):
        with column:
            if st.button(label,key=f"top_nav_v4_{page.lower().replace(' ','_')}",use_container_width=True,type="primary" if active==page else "secondary"):
                if page!=current:_track_event("navegacion",page);st.session_state["page"]=page;st.rerun()
    st.session_state["page"]=current
    st.markdown("<hr style='margin:9px 0 20px;border:none;border-top:2px solid rgba(40,22,58,.12);'>",unsafe_allow_html=True)
    return current


# Inicio V4: presenta Meta Lab como un pilar, sin desplazar el marketplace.
def render_home() -> None:
    _track_event("visita_inicio",once=True);_render_onboarding()
    st.markdown("""
    <div class="nexo-hero"><div class="nexo-hero-copy"><span class="nexo-eyebrow">TU PORTAL GEEK, HECHO EN COMUNIDAD</span>
    <h1>Compra, juega y entiende tu <em>universo</em>.</h1>
    <p>Descubre piezas, publica tu colección, tasa cartas y estudia el metagame competitivo desde un solo lugar.</p>
    <div class="nexo-chip-row"><span>Marketplace</span><span>Tasador</span><span>Meta competitivo</span><span>Torneos</span><span>Subastas</span><span>Servicios</span></div></div></div>
    <div class="universe-ribbon"><span>Pokémon</span><span>One Piece</span><span>Magic</span><span>Juegos de mesa</span><span>Manga</span><span>Figuras</span></div>
    """,unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    if c1.button("Explorar Marketplace",type="primary",use_container_width=True,key="home_market_v4"):_go_to("Marketplace")
    if c2.button("Abrir Meta Lab",use_container_width=True,key="home_meta_v4"):_go_to("Meta Lab")
    if c3.button("Tasar mi colección",use_container_width=True,key="home_tasar_v4"):_go_to("Tasador")
    listings=[x for x in st.session_state["marketplace_db"] if x.get("active",True)]
    hm1,hm2,hm3,hm4=st.columns(4);hm1.metric("Piezas activas",len(listings));hm2.metric("Mazos en Meta Lab",sum(len(x["mazos"]) for x in META_DEMO.values()));hm3.metric("Eventos demo",sum(len(x["torneos"]) for x in META_DEMO.values()));hm4.metric("Guardados",len(st.session_state.get("favorites",[])))
    st.markdown("<br>",unsafe_allow_html=True)
    _section_header("Meta de la semana","Del torneo a tu colección","Consulta tendencias, abre el mazo ganador y descubre cuánto te falta para completarlo.")
    dataset,_=_meta_dataset("Pokémon TCG","Standard");top=dataset.get("mazos",[])[:3]
    cols=st.columns(3)
    for i,(col,deck) in enumerate(zip(cols,top),1):
        with col:_render_meta_rank_card(deck,i,"home_meta_rank")
    if st.button("Explorar todo el Meta Lab",key="home_meta_all",type="primary"):_go_to("Meta Lab")
    st.markdown("<br>",unsafe_allow_html=True)
    _section_header("Radar del Nexo","Hallazgos que están llamando la atención","Una selección para descubrir piezas, vendedores y categorías.")
    featured=sorted(listings,key=lambda x:(x.get("likes",0),x.get("views",0)),reverse=True)[:4];cols=st.columns(4)
    for col,item in zip(cols,featured):
        with col:_render_listing_card(item,"home_v4")
    st.markdown('<div class="trust-strip"><strong>Un solo recorrido</strong><br>Tasa tu colección · descubre un mazo · revisa qué te falta · encuentra cartas en el marketplace</div>',unsafe_allow_html=True)
    _render_feature_vote("inicio","¿La propuesta se entiende claramente desde el inicio?")




# ══════════════════════════════════════════════════════════════════════════════
# NEXOGEEK · EDICIÓN MARKETPLACE GEEK
# La navegación pública prioriza el marketplace geek. Meta Lab y el constructor
# visual quedan disponibles como herramientas TCG secundarias dentro de un hub propio.
# ══════════════════════════════════════════════════════════════════════════════

GEEK_MARKETPLACE_CSS = r"""
<style>
:root{
  --ng-ink:#28163A;
  --ng-purple:#5B2A86;
  --ng-coral:#FF686B;
  --ng-gold:#FFC857;
  --ng-mint:#35C7B4;
  --ng-cream:#FFF8ED;
  --ng-paper:#FFFCF7;
}
.geek-market-hero{
  position:relative;overflow:hidden;border:3px solid var(--ng-ink);border-radius:28px;
  padding:2.35rem 2.45rem;background:
  radial-gradient(circle at 88% 42%,rgba(255,200,87,.98) 0 9%,transparent 9.5%),
  radial-gradient(circle at 88% 42%,transparent 0 18%,rgba(53,199,180,.92) 18.5% 25%,transparent 25.5%),
  radial-gradient(circle at 88% 42%,transparent 0 34%,rgba(255,104,107,.9) 34.5% 41%,transparent 41.5%),
  linear-gradient(135deg,#351649 0%,#5E2C86 58%,#7B46D9 100%);
  color:white;box-shadow:12px 12px 0 var(--ng-gold);margin-bottom:1.45rem;
}
.geek-market-hero:after{
  content:'✦';position:absolute;right:8.5%;top:16%;font-size:2rem;color:white;
  text-shadow:-56px 68px 0 var(--ng-mint),-110px 20px 0 var(--ng-coral),25px 100px 0 var(--ng-gold);
}
.geek-market-hero h1{
  max-width:760px;font-size:clamp(2.2rem,4.2vw,4.2rem);line-height:.98;margin:.8rem 0 1rem;
  color:white !important;letter-spacing:-.045em;
}
.geek-market-hero h1 em{color:var(--ng-gold);font-style:normal;border-bottom:6px solid var(--ng-coral);}
.geek-market-hero p{max-width:720px;font-size:1.04rem;color:#FFF9EE;margin:0 0 1.15rem;line-height:1.62;}
.geek-eyebrow{
  display:inline-flex;background:var(--ng-gold);color:var(--ng-ink);border:2px solid var(--ng-ink);
  border-radius:10px 10px 3px 10px;padding:.38rem .7rem;font-size:.7rem;font-weight:900;
  letter-spacing:.11em;text-transform:uppercase;box-shadow:4px 4px 0 var(--ng-coral);
}
.geek-chip-row{display:flex;gap:.55rem;flex-wrap:wrap;margin-top:1rem}
.geek-chip-row span{
  border:1px solid rgba(255,255,255,.6);background:rgba(255,255,255,.1);color:white;
  border-radius:999px;padding:.38rem .66rem;font-size:.72rem;font-weight:800;
}
.category-tile{
  min-height:150px;border:2px solid var(--ng-ink);border-radius:22px;background:var(--ng-paper);
  box-shadow:7px 7px 0 var(--ng-gold);padding:1.15rem;display:flex;flex-direction:column;
  justify-content:space-between;margin-bottom:.65rem;
}
.category-tile .icon{font-size:2rem}
.category-tile h4{margin:.45rem 0 .3rem;color:var(--ng-ink);font-size:1rem}
.category-tile p{margin:0;color:#75677E;font-size:.78rem;line-height:1.45}
.community-banner{
  border:3px solid var(--ng-ink);border-radius:25px;padding:1.7rem 1.85rem;
  background:linear-gradient(135deg,#DDF8F3,#FFF8ED 60%,#FFE1D7);
  box-shadow:10px 10px 0 var(--ng-coral);margin:1rem 0 1.4rem;
}
.community-banner h2{margin:.25rem 0;color:var(--ng-ink)}
.community-banner p{margin:.35rem 0 0;color:#62526B;max-width:860px}
.tool-hub-card{
  border:2px solid var(--ng-ink);border-radius:22px;padding:1.3rem;background:white;
  box-shadow:7px 7px 0 var(--ng-mint);min-height:175px;
}
.tool-hub-card.secondary{box-shadow:7px 7px 0 var(--ng-gold)}
.tool-hub-card h3{margin:.35rem 0;color:var(--ng-ink)}
.tool-hub-card p{color:#706278;font-size:.86rem;line-height:1.5}
.geek-pulse{
  border:2px dashed rgba(40,22,58,.4);background:#FFF8ED;border-radius:18px;padding:1rem 1.15rem;
  color:#5E4E67;margin:.8rem 0 1rem;
}
.geek-nav-note{
  border-left:5px solid var(--ng-coral);background:#FFF0ED;border-radius:0 14px 14px 0;
  padding:.75rem 1rem;color:#5B405F;font-size:.82rem;margin-bottom:1rem;
}
.event-card{
  border:2px solid var(--ng-ink);border-radius:20px;padding:1.1rem;background:white;
  box-shadow:6px 6px 0 var(--ng-gold);height:100%;
}
.event-card .date{font-size:.7rem;font-weight:900;color:var(--ng-purple);letter-spacing:.08em;text-transform:uppercase}
.event-card h4{color:var(--ng-ink);margin:.35rem 0}
.event-card p{color:#726278;font-size:.82rem;line-height:1.45}
.geek-market-strip{
  display:grid;grid-template-columns:repeat(4,1fr);gap:.75rem;margin:1rem 0 1.25rem;
}
.geek-market-strip div{
  border:2px solid var(--ng-ink);border-radius:16px;background:white;padding:.85rem;
  box-shadow:5px 5px 0 var(--ng-mint);
}
.geek-market-strip strong{display:block;color:var(--ng-ink)}
.geek-market-strip span{display:block;color:#76677D;font-size:.74rem;margin-top:.25rem}
@media(max-width:900px){
  .geek-market-hero{padding:1.65rem 1.35rem;box-shadow:7px 7px 0 var(--ng-gold)}
  .geek-market-hero:after{opacity:.35}
  .geek-market-strip{grid-template-columns:1fr 1fr}
}
@media(max-width:600px){
  .geek-market-strip{grid-template-columns:1fr}
}
</style>
"""


def _tour_status() -> tuple[int, int, list[tuple[str, bool]]]:
    """Recorrido centrado en el marketplace y la comunidad geek."""
    actions = st.session_state.get("_journey_actions", set())
    milestones = [
        ("Explorar el marketplace", "visita_marketplace" in actions),
        ("Abrir una publicación", "abrir_ficha" in actions),
        ("Guardar un favorito", "favorito" in actions),
        ("Simular una reserva", bool({"reserva", "checkout"} & set(actions))),
        ("Publicar, subastar o contratar", bool({"publicar", "visita_subastas", "visita_servicios"} & set(actions))),
    ]
    return sum(done for _, done in milestones), len(milestones), milestones


def _render_onboarding() -> None:
    if st.session_state.get("onboarding_complete"):
        return
    st.markdown(
        '<div class="welcome-deck"><span class="tour-badge">BIENVENIDA · EXPLORA EL NEXO</span>'
        '<h2>¿Qué parte del mundo geek quieres descubrir?</h2>'
        '<p>Elige un punto de entrada. El marketplace es el centro y las herramientas TCG quedan disponibles cuando las necesites.</p></div>',
        unsafe_allow_html=True,
    )
    cols = st.columns(4)
    missions = [
        (cols[0], "Comprador", "Buscar una pieza", "Explora cartas, mangas, figuras, juegos y accesorios.", "Marketplace"),
        (cols[1], "Vendedor", "Publicar algo", "Crea un anuncio y revisa cómo se verá frente a la comunidad.", "Vender"),
        (cols[2], "Comunidad", "Descubrir personas", "Encuentra servicios, actividades, creadores y subastas.", "Comunidad"),
        (cols[3], "Jugador TCG", "Usar herramientas", "Tasa cartas, revisa el meta y abre el constructor visual.", "Herramientas"),
    ]
    for col, role, title, copy, page in missions:
        with col:
            st.markdown(
                f'<div class="role-card"><div class="role-icon">✦</div><h3>{title}</h3><p>{copy}</p></div>',
                unsafe_allow_html=True,
            )
            if st.button(
                f"Entrar como {role.lower()}",
                key=f"onboard_geek_{role}",
                use_container_width=True,
                type="primary" if role == "Comprador" else "secondary",
            ):
                st.session_state["pilot_role"] = role
                st.session_state["onboarding_complete"] = True
                _track_event("onboarding", role, page)
                _go_to(page)
    if st.button("Explorar libremente", key="skip_onboarding_geek"):
        st.session_state["pilot_role"] = "Explorador"
        st.session_state["onboarding_complete"] = True
        _track_event("onboarding", "Explorador", "sin guía")
        st.rerun()


def _render_sidebar(api_key_default: str = "") -> tuple[str | None, float, float]:
    """Sidebar pública enfocada en compra, venta y comunidad geek."""
    api_key_input = st.session_state.get("api_key_guardada", api_key_default) or ""
    clp_rate = float(st.session_state.get("clp_rate_sidebar", 950))
    comision = float(st.session_state.get("commission_sidebar", 5.0))
    _CFG["max_workers"] = int(st.session_state.get("workers_sidebar", 4))
    _CFG["throttle"] = float(st.session_state.get("throttle_sidebar", 0.2))

    if not _DB_LOADED:
        for cp in ("card_data", os.path.join("pokemon-tcg-data", "cards", "en")):
            if os.path.isdir(cp) and cargar_base_local(cp):
                break

    nav_pages = ["Inicio", "Marketplace", "Vender", "Subastas", "Servicios", "Comunidad", "Herramientas", "Feedback"]
    with st.sidebar:
        st.markdown(
            "<div class='side-brand'><div class='side-logo'>NG</div>"
            "<div><strong>NexoGeek</strong><small>marketplace geek y coleccionables</small></div></div>",
            unsafe_allow_html=True,
        )
        st.markdown('<span class="pilot-pill"><span class="live-dot"></span>Piloto activo</span>', unsafe_allow_html=True)
        st.caption("Compra, vende, descubre y conecta con la comunidad geek.")

        current = st.session_state.get("page", "Inicio")
        if current == "Detalle":
            current_nav = "Marketplace"
        elif current in {"Tasador", "Meta Lab"}:
            current_nav = "Herramientas"
        else:
            current_nav = current if current in nav_pages else "Inicio"

        if (
            st.session_state.get("_sidebar_nav_page") != current_nav
            or st.session_state.get("sidebar_public_nav") not in nav_pages
        ):
            st.session_state["sidebar_public_nav"] = current_nav
            st.session_state["_sidebar_nav_page"] = current_nav

        selected_nav = st.selectbox(
            "Ir a",
            nav_pages,
            index=nav_pages.index(current_nav),
            key="sidebar_public_nav",
        )
        if selected_nav != current_nav:
            st.session_state["page"] = selected_nav
            st.session_state["pending_nav"] = selected_nav
            st.session_state["_sidebar_nav_page"] = selected_nav
            st.rerun()

        with st.expander("Mi espacio geek", expanded=True):
            alias = st.text_input(
                "Alias",
                value=st.session_state.get("pilot_alias", "Usuario_Piloto"),
                key="profile_alias_input_geek",
            )
            location = st.selectbox(
                "Ubicación",
                UBICACIONES_DEMO,
                index=UBICACIONES_DEMO.index(st.session_state.get("pilot_location", "Santiago"))
                if st.session_state.get("pilot_location", "Santiago") in UBICACIONES_DEMO else 0,
                key="profile_location_input_geek",
            )
            universes = [
                "De todo un poco", "Pokémon TCG", "One Piece", "Magic",
                "Manga y cómics", "Figuras y animé", "Juegos de mesa", "Accesorios",
            ]
            current_interest = st.session_state.get("favorite_universe", universes[0])
            interest = st.selectbox(
                "Interés principal",
                universes,
                index=universes.index(current_interest) if current_interest in universes else 0,
                key="profile_interest_geek",
            )
            st.session_state["pilot_alias"] = alias.strip() or "Usuario_Piloto"
            st.session_state["pilot_location"] = location
            st.session_state["favorite_universe"] = interest
            st.caption(f"Explorando como: {st.session_state.get('pilot_role') or 'Explorador'}")

        f1, f2, f3 = st.columns(3)
        f1.metric("Favoritos", len(st.session_state.get("favorites", [])))
        f2.metric("Reserva", len(st.session_state.get("cart", [])))
        f3.metric("Avisos", len(st.session_state.get("notifications", [])))

        completed, total, milestones = _tour_status()
        pct = int(completed / total * 100) if total else 0
        st.markdown(
            f'<div class="journey-card"><strong>Recorrido marketplace · {completed}/{total}</strong>'
            f'<div class="ng-progress"><span style="width:{pct}%"></span></div>'
            f'<small>{"Siguiente: " + next((name for name, done in milestones if not done), "Recorrido completado")}</small></div>',
            unsafe_allow_html=True,
        )

        with st.expander("Reserva", expanded=False):
            cart_ids = list(st.session_state.get("cart", []))
            if not cart_ids:
                st.caption("Tu reserva está vacía.")
            total_price = 0
            for lid in cart_ids:
                item = _listing_by_id(lid)
                if not item:
                    continue
                total_price += int(item.get("price", 0))
                c1, c2 = st.columns([4, 1])
                c1.caption(f"{item.get('title')} · {_fmt_clp(item.get('price'))}")
                if c2.button("×", key=f"sidebar_remove_geek_{lid}"):
                    st.session_state["cart"].remove(lid)
                    st.rerun()
            if cart_ids:
                st.markdown(f"**Total referencial: {_fmt_clp(total_price)}**")
                if st.button("Simular checkout", use_container_width=True, type="primary", key="sidebar_checkout_geek"):
                    _track_event("checkout", metadata=str(total_price))
                    st.success("Flujo completado sin cobro.")

        with st.expander("Actividad reciente", expanded=False):
            notifications = st.session_state.get("notifications", [])
            if not notifications:
                st.caption("Aún no hay actividad.")
            for note in notifications[:6]:
                st.caption(f"{note['time']} · {note['message']}")

        with st.expander("Acceso anfitrión", expanded=False):
            if not st.session_state.get("admin_unlocked"):
                st.markdown(
                    '<div class="admin-lock">Configuración técnica, API y descargas del piloto.</div>',
                    unsafe_allow_html=True,
                )
                pin = st.text_input("PIN", type="password", key="admin_pin_input_geek")
                if st.button("Desbloquear", key="admin_unlock_geek", use_container_width=True):
                    if pin == ADMIN_PIN:
                        st.session_state["admin_unlocked"] = True
                        st.rerun()
                    else:
                        st.error("PIN incorrecto.")
            else:
                st.markdown('<span class="host-badge">MODO ANFITRIÓN</span>', unsafe_allow_html=True)
                api_key_input = st.text_input(
                    "PokémonTCG API Key",
                    type="password",
                    value=api_key_input,
                    key="api_key_sidebar_geek",
                )
                clp_rate = st.number_input(
                    "USD → CLP", min_value=0, value=int(clp_rate), step=10,
                    key="clp_rate_sidebar_geek",
                )
                comision = st.number_input(
                    "Comisión simulada %", min_value=0.0, max_value=50.0,
                    value=float(comision), step=0.5, key="commission_sidebar_geek",
                )
                workers = st.slider(
                    "Hilos", 1, 8, int(_CFG.get("max_workers", 4)),
                    key="workers_sidebar_geek",
                )
                throttle = st.slider(
                    "Delay anti rate-limit", 0.0, 2.0,
                    float(_CFG.get("throttle", .2)), .1,
                    key="throttle_sidebar_geek",
                )
                st.session_state["api_key_guardada"] = api_key_input
                st.session_state["clp_rate_sidebar"] = int(clp_rate)
                st.session_state["commission_sidebar"] = float(comision)
                st.session_state["workers_sidebar"] = int(workers)
                st.session_state["throttle_sidebar"] = float(throttle)
                _CFG["max_workers"] = int(workers)
                _CFG["throttle"] = float(throttle)
                st.caption(f"Fuente de cartas: {'DB local + API de precios' if _DB_LOADED else 'API cloud'}")
                if api_key_input:
                    st.success("API Key cargada desde Secrets o panel anfitrión.")
                else:
                    st.warning("Sin API Key: habrá imágenes locales, pero no precios automáticos.")

                api_c1, api_c2 = st.columns(2)
                if api_c1.button("Probar API", key="test_api_geek", use_container_width=True):
                    ok, message = diagnostico_api_pokemon(api_key_input or None)
                    (st.success if ok else st.error)(message)
                if api_c2.button("Vaciar caché precios", key="clear_price_cache_geek", use_container_width=True):
                    limpiar_cache_precios()
                    st.success("Caché de precios vaciada.")

                fb1 = leer_feedback()
                fb2 = _read_csv_safe(EXTENDED_FEEDBACK_FILE)
                interactions = _read_csv_safe(INTERACTIONS_FILE)
                journey = _read_csv_safe(JOURNEY_FILE)
                if not fb1.empty:
                    st.download_button(
                        "Feedback básico", fb1.to_csv(index=False).encode("utf-8-sig"),
                        "feedback_nexogeek.csv", "text/csv", use_container_width=True,
                    )
                if not fb2.empty:
                    st.download_button(
                        "Feedback completo", fb2.to_csv(index=False).encode("utf-8-sig"),
                        "feedback_nexogeek_extendido.csv", "text/csv", use_container_width=True,
                    )
                if not interactions.empty:
                    st.download_button(
                        "Reacciones", interactions.to_csv(index=False).encode("utf-8-sig"),
                        "interacciones_nexogeek.csv", "text/csv", use_container_width=True,
                    )
                if not journey.empty:
                    st.download_button(
                        "Recorrido de usuarios", journey.to_csv(index=False).encode("utf-8-sig"),
                        "recorrido_nexogeek.csv", "text/csv", use_container_width=True,
                    )
                if st.button("Cerrar modo anfitrión", key="admin_lock_geek", use_container_width=True):
                    st.session_state["admin_unlocked"] = False
                    st.rerun()

        if st.session_state.get("admin_unlocked"):
            if st.button("Reiniciar demo", use_container_width=True, key="reset_demo_geek"):
                for key in [
                    "marketplace_db", "favorites", "cart", "compare", "notifications",
                    "selected_listing", "deck_cart", "meta_selected_card", "meta_print_choices",
                    "auction_watchlist", "subastas_db", "servicios_db", "feature_votes",
                    "df_result", "candidatos", "onboarding_complete", "pilot_role",
                    "_journey_seen", "_journey_actions", "favorite_universe",
                ]:
                    st.session_state.pop(key, None)
                _init_demo_state()
                _init_meta_state()
                st.rerun()

    return (api_key_input or None), float(clp_rate), float(comision)


def _render_top_navigation() -> str:
    options = [
        ("Inicio", "Inicio"),
        ("Marketplace", "Mercado"),
        ("Vender", "Vender"),
        ("Subastas", "Subastas"),
        ("Servicios", "Servicios"),
        ("Comunidad", "Comunidad"),
        ("Herramientas", "Herramientas"),
        ("Feedback", "Feedback"),
    ]
    pending = st.session_state.pop("pending_nav", None)
    current = pending or st.session_state.get("page", "Inicio")
    valid = {x[0] for x in options} | {"Detalle", "Tasador", "Meta Lab"}
    if current not in valid:
        current = "Inicio"
    if current == "Detalle":
        active = "Marketplace"
    elif current in {"Tasador", "Meta Lab"}:
        active = "Herramientas"
    else:
        active = current

    nav_cols = st.columns([1.8, .72, .88, .72, .78, .78, .82, .94, .76], gap="small")
    with nav_cols[0]:
        st.markdown(
            "<div class='nexo-brand'><div class='nexo-brand-mark'>✦</div>"
            "<div><div class='nexo-brand-name'>NexoGeek</div>"
            "<div class='nexo-brand-tag'>marketplace · comunidad · coleccionables</div></div></div>",
            unsafe_allow_html=True,
        )
    for column, (page, label) in zip(nav_cols[1:], options):
        with column:
            if st.button(
                label,
                key=f"top_nav_geek_{page.lower().replace(' ', '_')}",
                use_container_width=True,
                type="primary" if active == page else "secondary",
            ):
                if page != current:
                    _track_event("navegacion", page)
                    st.session_state["page"] = page
                    st.rerun()

    st.session_state["page"] = current
    st.markdown(
        "<hr style='margin:9px 0 20px;border:none;border-top:2px solid rgba(40,22,58,.12);'>",
        unsafe_allow_html=True,
    )
    return current


def _open_marketplace_search(term: str = "") -> None:
    st.session_state["market_search_v3"] = term
    _track_event("categoria_portada", term or "todos")
    _go_to("Marketplace")


def render_home() -> None:
    _track_event("visita_inicio", once=True)
    _render_onboarding()

    st.markdown(
        """
        <div class="geek-market-hero">
          <span class="geek-eyebrow">MARKETPLACE GEEK · HECHO EN COMUNIDAD</span>
          <h1>Compra, vende y descubre tu próximo <em>hallazgo geek</em>.</h1>
          <p>Cartas TCG, mangas, figuras, juegos de mesa, accesorios, productos sellados y creaciones de la comunidad reunidos en un solo lugar.</p>
          <div class="geek-chip-row">
            <span>Cartas TCG</span><span>Mangas y cómics</span><span>Figuras</span>
            <span>Juegos de mesa</span><span>Accesorios</span><span>Servicios creativos</span>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns([1.15, 1, 1])
    if c1.button("Explorar Marketplace", type="primary", use_container_width=True, key="home_market_geek"):
        _go_to("Marketplace")
    if c2.button("Publicar una pieza", use_container_width=True, key="home_sell_geek"):
        _go_to("Vender")
    if c3.button("Ver subastas", use_container_width=True, key="home_auctions_geek"):
        _go_to("Subastas")

    listings = [x for x in st.session_state["marketplace_db"] if x.get("active", True)]
    categories = len(set(x.get("product_type", "") for x in listings))
    sellers = len(set(x.get("seller", "") for x in listings))
    hm1, hm2, hm3, hm4 = st.columns(4)
    hm1.metric("Piezas activas", len(listings))
    hm2.metric("Categorías", categories)
    hm3.metric("Vendedores demo", sellers)
    hm4.metric("Favoritos guardados", len(st.session_state.get("favorites", [])))

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header(
        "Explora por categoría",
        "Todo el mundo geek en un solo Nexo",
        "Entra por el tipo de pieza que estás buscando y deja que la comunidad haga el resto.",
    )
    category_data = [
        ("🃏", "TCG y cartas", "Cartas sueltas, sellados y accesorios.", "TCG"),
        ("📚", "Mangas y cómics", "Tomos, colecciones y ediciones especiales.", "manga"),
        ("🗿", "Figuras y animé", "Figuras, estatuas y piezas de vitrina.", "figura"),
        ("🎲", "Juegos de mesa", "Juegos modernos, clásicos y expansiones.", "juego de mesa"),
        ("🧰", "Accesorios", "Playmats, deck boxes, dados y exhibidores.", "accesorio"),
        ("✨", "Creaciones", "Impresión 3D, arte y productos personalizados.", "diseño"),
    ]
    cols = st.columns(6)
    for idx, (col, (icon, title, copy, term)) in enumerate(zip(cols, category_data)):
        with col:
            st.markdown(
                f'<div class="category-tile"><div class="icon">{icon}</div>'
                f'<div><h4>{title}</h4><p>{copy}</p></div></div>',
                unsafe_allow_html=True,
            )
            if st.button("Explorar", key=f"home_category_{idx}", use_container_width=True):
                _open_marketplace_search(term)

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header(
        "Radar del Nexo",
        "Hallazgos de la comunidad",
        "Publicaciones con mayor interés, guardados y actividad dentro de la demostración.",
    )
    featured = sorted(
        listings,
        key=lambda x: (x.get("likes", 0), x.get("views", 0)),
        reverse=True,
    )[:4]
    cols = st.columns(4)
    for col, item in zip(cols, featured):
        with col:
            _render_listing_card(item, "home_geek")

    st.markdown(
        '<div class="community-banner"><span class="section-kicker">COMUNIDAD EN MOVIMIENTO</span>'
        '<h2>El valor no está solo en el producto.</h2>'
        '<p>Descubre servicios, creadores, encargos, torneos, subastas y espacios para conectar con otras personas del hobby.</p></div>',
        unsafe_allow_html=True,
    )
    cm1, cm2, cm3 = st.columns(3)
    if cm1.button("Explorar comunidad", type="primary", use_container_width=True, key="home_community_geek"):
        _go_to("Comunidad")
    if cm2.button("Ver servicios", use_container_width=True, key="home_services_geek"):
        _go_to("Servicios")
    if cm3.button("Explorar subastas", use_container_width=True, key="home_auction_geek"):
        _go_to("Subastas")

    st.markdown("<br>", unsafe_allow_html=True)
    _section_header(
        "Herramientas opcionales",
        "Para quienes también juegan TCG",
        "El marketplace sigue siendo el centro. El tasador, Meta Lab y constructor visual aparecen como utilidades complementarias.",
    )
    t1, t2 = st.columns(2)
    with t1:
        st.markdown(
            '<div class="tool-hub-card"><div class="section-kicker">TASACIÓN</div>'
            '<h3>Identifica y valoriza tus cartas</h3>'
            '<p>Reconoce versiones, consulta referencias de precio y convierte el resultado en una publicación.</p></div>',
            unsafe_allow_html=True,
        )
        if st.button("Abrir Tasador", key="home_tasador_geek", use_container_width=True):
            _go_to("Tasador")
    with t2:
        st.markdown(
            '<div class="tool-hub-card secondary"><div class="section-kicker">JUEGO COMPETITIVO</div>'
            '<h3>Meta Lab y constructor visual</h3>'
            '<p>Revisa mazos, listas, matchups y calcula cuánto te falta para completar un deck.</p></div>',
            unsafe_allow_html=True,
        )
        if st.button("Abrir Herramientas TCG", key="home_tools_geek", use_container_width=True):
            _go_to("Herramientas")

    st.markdown(
        '<div class="trust-strip"><strong>Compra y vende con contexto</strong><br>'
        'Estado visible · reputación · moderación · entrega coordinada · referencias de precio · comunidad local</div>',
        unsafe_allow_html=True,
    )
    _render_feature_vote("inicio", "¿La portada comunica claramente que NexoGeek es un marketplace geek amplio?")


_MARKETPLACE_CORE_GEEK = render_marketplace


def render_marketplace() -> None:
    st.markdown(
        '<div class="geek-market-strip">'
        '<div><strong>TCG</strong><span>Cartas, sellados y accesorios</span></div>'
        '<div><strong>Coleccionables</strong><span>Figuras, manga y piezas especiales</span></div>'
        '<div><strong>Juegos</strong><span>Mesa, rol y productos para jugar</span></div>'
        '<div><strong>Comunidad</strong><span>Creadores, encargos y servicios</span></div>'
        '</div>',
        unsafe_allow_html=True,
    )
    quick = st.columns(7)
    quick_terms = [
        ("Todo", ""), ("TCG", "TCG"), ("Manga", "manga"), ("Figuras", "figura"),
        ("Juegos", "juego de mesa"), ("Accesorios", "accesorio"), ("Sellados", "sellado"),
    ]
    for col, (label, term) in zip(quick, quick_terms):
        with col:
            if st.button(label, key=f"market_quick_{label.lower()}", use_container_width=True):
                st.session_state["market_search_v3"] = term
                st.rerun()
    _MARKETPLACE_CORE_GEEK()


def render_community() -> None:
    _track_event("visita_comunidad", once=True)
    st.markdown(
        '<div class="community-banner"><span class="section-kicker">COMUNIDAD GEEK</span>'
        '<h2>Descubre personas, actividades y proyectos.</h2>'
        '<p>NexoGeek no solo conecta compradores con vendedores: también visibiliza creadores, organizadores, tiendas, ligas y servicios del hobby.</p></div>',
        unsafe_allow_html=True,
    )

    listings = [x for x in st.session_state.get("marketplace_db", []) if x.get("active", True)]
    services = st.session_state.get("servicios_db", [])
    auctions = st.session_state.get("subastas_db", [])
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Vendedores activos", len(set(x.get("seller") for x in listings)))
    c2.metric("Servicios disponibles", len(services))
    c3.metric("Subastas abiertas", len(auctions))
    c4.metric("Universos representados", len(set(x.get("game") for x in listings)))

    tabs = st.tabs(["Agenda geek", "Creadores y servicios", "Subastas destacadas", "Participa"])
    with tabs[0]:
        events = [
            ("SÁBADO · 16:00", "Liga TCG de la comunidad", "Torneo amistoso y espacio de intercambio para jugadores nuevos y experimentados.", "Santiago"),
            ("DOMINGO · 12:00", "Tarde de juegos de mesa", "Mesas abiertas, demostraciones y búsqueda de grupos para jugar.", "Providencia"),
            ("PRÓXIMO MES", "Feria de ilustración y coleccionables", "Artistas, impresión 3D, accesorios y piezas hechas por la comunidad.", "Ñuñoa"),
            ("MENSUAL", "Club manga, animé y cosplay", "Encuentro temático, intercambio de recomendaciones y venta de colecciones.", "Online + presencial"),
        ]
        cols = st.columns(4)
        for col, (date, title, copy, place) in zip(cols, events):
            with col:
                st.markdown(
                    f'<div class="event-card"><div class="date">{date}</div><h4>{title}</h4>'
                    f'<p>{copy}</p><span class="active-filter">{place}</span></div>',
                    unsafe_allow_html=True,
                )
        st.caption("Agenda demostrativa para validar interés. Los eventos reales requerirán verificación y fecha confirmada.")

    with tabs[1]:
        if not services:
            _render_empty_state("✨", "Aún no hay servicios", "La comunidad podrá publicar encargos, diseño, impresión, grading y organización.")
        else:
            cols = st.columns(min(4, len(services)))
            for col, service in zip(cols, services[:4]):
                with col:
                    st.markdown(
                        f'<div class="soft-card"><div class="section-kicker">{_safe_text(service.get("type","Servicio"))}</div>'
                        f'<h4>{_safe_text(service.get("title",""))}</h4>'
                        f'<p>{_safe_text(service.get("description",""))}</p>'
                        f'<strong>{_fmt_clp(service.get("price",0))}</strong><br>'
                        f'<small>{_safe_text(service.get("provider",""))} · ⭐ {service.get("rating","-")}</small></div>',
                        unsafe_allow_html=True,
                    )
            if st.button("Ver todos los servicios", type="primary", key="community_all_services"):
                _go_to("Servicios")

    with tabs[2]:
        if not auctions:
            _render_empty_state("🔨", "Sin subastas activas", "Las subastas verificadas aparecerán aquí.")
        else:
            for auction in auctions[:3]:
                with st.container(border=True):
                    a1, a2, a3 = st.columns([3, 1, 1])
                    a1.markdown(
                        f"**{auction.get('name','')}**  \n"
                        f"{auction.get('game','')} · {auction.get('seller','')}"
                    )
                    a2.metric("Oferta", _fmt_clp(auction.get("current_bid", 0)))
                    a3.metric("Pujas", auction.get("bids", 0))
            if st.button("Entrar a Subastas", key="community_all_auctions", use_container_width=True):
                _go_to("Subastas")

    with tabs[3]:
        st.markdown("#### ¿Cómo quieres participar?")
        p1, p2, p3 = st.columns(3)
        with p1:
            st.markdown("**Vende una pieza**  \nPublica cartas, figuras, mangas, juegos o accesorios.")
            if st.button("Crear publicación", key="community_publish", use_container_width=True):
                _go_to("Vender")
        with p2:
            st.markdown("**Ofrece un servicio**  \nDiseño, impresión, grading, encargos u organización.")
            if st.button("Publicar servicio", key="community_service", use_container_width=True):
                _go_to("Servicios")
        with p3:
            st.markdown("**Propón una actividad**  \nAyuda a construir una agenda local del hobby.")
            if st.button("Registrar interés", key="community_event_interest", use_container_width=True):
                _track_event("interes_evento_comunidad")
                st.success("Interés registrado para el análisis del piloto.")

    _render_feature_vote("comunidad", "¿La comunidad agrega valor más allá de comprar y vender?")


def render_tools_hub() -> None:
    _track_event("visita_herramientas", once=True)
    _section_header(
        "Herramientas TCG",
        "Utilidades para jugar, coleccionar y vender mejor",
        "Son funciones complementarias del marketplace; no reemplazan la experiencia principal de compra y venta geek.",
    )
    st.markdown(
        '<div class="geek-nav-note"><strong>Enfoque del producto:</strong> el marketplace es el centro. '
        'Estas herramientas ayudan a identificar cartas, entender el juego y convertir una necesidad en una búsqueda dentro del mercado.</div>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(
            '<div class="tool-hub-card"><div class="section-kicker">TASADOR</div>'
            '<h3>De carta física a publicación</h3>'
            '<p>Busca la impresión correcta, consulta precios de referencia, administra tu inventario y publica en el marketplace.</p></div>',
            unsafe_allow_html=True,
        )
        if st.button("Abrir Tasador", type="primary", use_container_width=True, key="tools_open_tasador"):
            _go_to("Tasador")
    with c2:
        st.markdown(
            '<div class="tool-hub-card secondary"><div class="section-kicker">META LAB</div>'
            '<h3>Mazos, torneos y constructor visual</h3>'
            '<p>Explora listas competitivas, marca las cartas que ya tienes y calcula el costo estimado de las faltantes.</p></div>',
            unsafe_allow_html=True,
        )
        if st.button("Abrir Meta Lab", type="primary", use_container_width=True, key="tools_open_meta"):
            _go_to("Meta Lab")

    st.markdown("<br>", unsafe_allow_html=True)
    m1, m2, m3 = st.columns(3)
    m1.metric("Juegos con datos demo", len(META_DEMO))
    m2.metric("Mazos disponibles", sum(len(data.get("mazos", [])) for data in META_DEMO.values()))
    m3.metric("Cartas tasadas", len(st.session_state.get("df_result", [])) if "df_result" in st.session_state else 0)

    st.markdown("#### Cómo se conectan con el marketplace")
    steps = [
        ("1", "Identifica", "El tasador reconoce la carta y su impresión."),
        ("2", "Decide", "Meta Lab muestra una lista o necesidad concreta."),
        ("3", "Busca", "NexoGeek lleva la carta faltante al Marketplace."),
        ("4", "Publica", "Las cartas repetidas pueden transformarse en anuncios."),
    ]
    cols = st.columns(4)
    for col, (num, title, copy) in zip(cols, steps):
        with col:
            st.markdown(
                f'<div class="soft-card"><div class="step-number">{num}</div><h4>{title}</h4><p>{copy}</p></div>',
                unsafe_allow_html=True,
            )
    _render_feature_vote("herramientas", "¿Las herramientas TCG se sienten útiles sin quitar protagonismo al marketplace?")



# ══════════════════════════════════════════════════════════════════════════════
# COMUNIDAD ACTIVA — RSVP, LISTA DE ESPERA Y COORDINACIÓN
# ══════════════════════════════════════════════════════════════════════════════
EVENT_RSVP_FILE = os.getenv("NEXOGEEK_EVENT_RSVP_FILE", "eventos_nexogeek_rsvp.csv")
_EVENT_RSVP_LOCK = threading.Lock()

COMMUNITY_EVENTS_CSS = """
<style>
.event-shell{
  background:#FFFDF8;border:2px solid var(--ng-ink);border-radius:22px 22px 8px 22px;
  padding:18px;box-shadow:6px 6px 0 var(--ng-sun);margin-bottom:14px;min-height:260px
}
.event-shell .event-date{font-size:.68rem;font-weight:950;letter-spacing:.09em;text-transform:uppercase;color:var(--ng-purple)}
.event-shell h3{margin:.35rem 0 .45rem;color:var(--ng-ink);font-size:1.25rem;line-height:1.18}
.event-shell p{color:#726278;font-size:.82rem;line-height:1.45;margin:.2rem 0 .8rem}
.event-meta-row{display:flex;gap:7px;flex-wrap:wrap;margin:.45rem 0 .75rem}
.event-chip{display:inline-flex;align-items:center;background:#FFF8ED;border:1.5px solid #8A55B5;color:#512176;
  border-radius:999px;padding:4px 8px;font-size:.67rem;font-weight:850}
.event-status{display:inline-flex;align-items:center;border:2px solid var(--ng-ink);border-radius:10px 10px 4px 10px;
  padding:5px 9px;font-size:.7rem;font-weight:950;box-shadow:2px 2px 0 var(--ng-mint);margin-bottom:8px}
.event-status.confirmed{background:#DDF8F3;color:#185D56}.event-status.interested{background:#FFF0CC;color:#7C4B00}
.event-status.waitlist{background:#F3E9FF;color:#512176}.event-status.cancelled{background:#FFE1D7;color:#8B341E}
.event-counts{color:var(--ng-muted);font-size:.75rem;margin:.25rem 0 .4rem}
.event-person-list{background:#F9F2FF;border:1px solid #D7B9EE;border-radius:12px;padding:9px 11px;color:#5F4B69;font-size:.73rem}
.boardgame-spotlight{background:linear-gradient(135deg,#DDF8F3 0%,#FFF0CC 100%);border:2px solid var(--ng-ink);
  border-radius:20px 20px 7px 20px;padding:17px;box-shadow:5px 5px 0 var(--ng-coral);margin:12px 0 18px}
.boardgame-spotlight h3{margin:.1rem 0 .35rem;color:var(--ng-ink)}
.boardgame-spotlight p{margin:0;color:#66566E;font-size:.82rem}
@media (max-width:780px){.event-shell{min-height:auto}.event-shell h3{font-size:1.1rem}}
</style>
"""


def _board_game_demo_listings() -> list[dict]:
    """Publicaciones demo para representar con claridad al público de juegos de mesa."""
    return [
        {
            "id":"board-catan-001","title":"Catan · juego base en español","game":"Juegos de mesa",
            "product_type":"Juego de mesa","condition":"Usado","price":25990,"location":"Ñuñoa",
            "shipping":"Envío y retiro","seller":"MeepleSantiago","verified":True,"rating":4.9,"sales":58,
            "image":"","stock":1,"negotiable":True,
            "description":"Juego base completo en español. Incluye todas las losetas, cartas, caminos, poblados y dados. Caja con desgaste leve de almacenamiento.",
            "tags":["Completo","3-4 jugadores","Conversable"],"views":118,"likes":24,"active":True,"owner":False,
            "response_time":"Responde en 40 minutos","member_since":"Miembro desde 2023","photo_count":8,"protected":True,
            "authenticity":"Componentes revisados y fotografiados por el vendedor",
        },
        {
            "id":"board-dixit-001","title":"Dixit Odyssey · edición en español","game":"Juegos de mesa",
            "product_type":"Juego de mesa","condition":"Como nuevo","price":22990,"location":"Providencia",
            "shipping":"Envío y retiro","seller":"LaMesaSecreta","verified":True,"rating":4.8,"sales":41,
            "image":"","stock":1,"negotiable":False,
            "description":"Edición Odyssey para grupos grandes. Cartas y componentes completos, usados en dos sesiones y guardados en bolsas.",
            "tags":["Completo","Hasta 12 jugadores","Familiar"],"views":94,"likes":19,"active":True,"owner":False,
            "response_time":"Responde en 1 hora","member_since":"Miembro desde 2024","photo_count":7,"protected":True,
            "authenticity":"Inventario de cartas y componentes disponible en la ficha",
        },
        {
            "id":"board-ticket-001","title":"Ticket to Ride Europa · edición en español","game":"Juegos de mesa",
            "product_type":"Juego de mesa","condition":"Usado","price":29990,"location":"Santiago",
            "shipping":"Solo retiro","seller":"MesaCritica","verified":True,"rating":4.8,"sales":35,
            "image":"","stock":1,"negotiable":True,
            "description":"Juego completo, cartas enfundadas y trenes separados por color. Ideal para iniciar una colección moderna de juegos de mesa.",
            "tags":["Completo","Familiar","Conversable"],"views":82,"likes":13,"active":True,"owner":False,
            "response_time":"Responde el mismo día","member_since":"Miembro desde 2024","photo_count":9,"protected":True,
            "authenticity":"Componentes contabilizados antes de publicar",
        },
    ]


_legacy_demo_marketplace_seed_events = _demo_marketplace_seed
_legacy_init_demo_state_events = _init_demo_state


def _demo_marketplace_seed() -> list[dict]:
    items = _legacy_demo_marketplace_seed_events()
    existing = {str(item.get("id")) for item in items}
    for item in _board_game_demo_listings():
        if item["id"] not in existing:
            items.append(item)
    return items


def _demo_community_events_seed() -> list[dict]:
    return [
        {
            "id":"community-tcg-league","date_label":"SÁBADO 20 JUN · 16:00",
            "title":"Liga TCG de la comunidad",
            "description":"Torneo amistoso, intercambio y mesas para Pokémon, One Piece y Magic. Apto para jugadores nuevos.",
            "location":"Santiago Centro","modality":"Presencial","capacity":24,"base_confirmed":17,"base_interested":31,
            "organizer":"LigaCentral","start":"20260620T200000Z","end":"20260620T230000Z",
            "whatsapp_secret":"WHATSAPP_EVENT_TCG",
        },
        {
            "id":"community-boardgame-afternoon","date_label":"DOMINGO 21 JUN · 12:00",
            "title":"Tarde abierta de juegos de mesa",
            "description":"Mesas de Catan, Dixit, Ticket to Ride y juegos cooperativos. Puedes llegar solo y unirte a un grupo.",
            "location":"Providencia","modality":"Presencial","capacity":32,"base_confirmed":19,"base_interested":38,
            "organizer":"MesaCritica","start":"20260621T160000Z","end":"20260621T210000Z",
            "whatsapp_secret":"WHATSAPP_EVENT_BOARDGAMES",
        },
        {
            "id":"community-art-fair","date_label":"SÁBADO 4 JUL · 11:00",
            "title":"Feria de ilustración y coleccionables",
            "description":"Artistas, impresión 3D, accesorios, figuras y piezas realizadas por creadores de la comunidad.",
            "location":"Ñuñoa","modality":"Presencial","capacity":60,"base_confirmed":34,"base_interested":72,
            "organizer":"PortalCraft","start":"20260704T150000Z","end":"20260704T220000Z",
            "whatsapp_secret":"WHATSAPP_EVENT_FAIR",
        },
        {
            "id":"community-manga-club","date_label":"MENSUAL · 19:30",
            "title":"Club de manga, animé y cosplay",
            "description":"Encuentro temático, recomendaciones, intercambio de colecciones y planificación de actividades.",
            "location":"Online + presencial","modality":"Híbrido","capacity":40,"base_confirmed":21,"base_interested":49,
            "organizer":"NexoOtaku","start":"20260711T233000Z","end":"20260712T013000Z",
            "whatsapp_secret":"WHATSAPP_EVENT_MANGA",
        },
    ]


def _init_demo_state() -> None:
    _legacy_init_demo_state_events()
    st.session_state.setdefault("community_events_db", _demo_community_events_seed())
    st.session_state.setdefault("event_rsvp_status", {})
    # Si Streamlit recarga el código sin limpiar session_state, añadimos los juegos demo faltantes.
    listings = st.session_state.setdefault("marketplace_db", [])
    existing = {str(item.get("id")) for item in listings}
    for item in _board_game_demo_listings():
        if item["id"] not in existing:
            listings.append(item)


def _read_event_rsvps() -> pd.DataFrame:
    columns = ["fecha", "event_id", "session_id", "alias", "status"]
    if not os.path.exists(EVENT_RSVP_FILE):
        return pd.DataFrame(columns=columns)
    try:
        df = pd.read_csv(EVENT_RSVP_FILE)
        for column in columns:
            if column not in df.columns:
                df[column] = ""
        return df[columns]
    except Exception:
        return pd.DataFrame(columns=columns)


def _write_event_rsvp(event_id: str, status: str) -> None:
    session_id = str(st.session_state.get("session_id") or _new_id("session"))
    st.session_state["session_id"] = session_id
    alias = str(st.session_state.get("pilot_alias") or "Participante").strip() or "Participante"
    row = {
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "event_id": str(event_id), "session_id": session_id,
        "alias": alias[:60], "status": str(status),
    }
    try:
        with _EVENT_RSVP_LOCK:
            df = _read_event_rsvps()
            if not df.empty:
                df = df[~((df["event_id"].astype(str) == str(event_id)) &
                          (df["session_id"].astype(str) == session_id))]
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
            df.to_csv(EVENT_RSVP_FILE, index=False, encoding="utf-8-sig")
    except Exception:
        pass
    st.session_state.setdefault("event_rsvp_status", {})[str(event_id)] = str(status)
    _track_event(f"evento_{status}", str(event_id), alias)


def _current_event_status(event_id: str) -> str:
    cached = st.session_state.setdefault("event_rsvp_status", {}).get(str(event_id))
    if cached:
        return str(cached)
    session_id = str(st.session_state.get("session_id", ""))
    if not session_id:
        return ""
    df = _read_event_rsvps()
    if df.empty:
        return ""
    rows = df[(df["event_id"].astype(str) == str(event_id)) &
              (df["session_id"].astype(str) == session_id)]
    if rows.empty:
        return ""
    status = str(rows.iloc[-1].get("status", ""))
    st.session_state["event_rsvp_status"][str(event_id)] = status
    return status


def _event_snapshot(event: dict) -> dict:
    df = _read_event_rsvps()
    event_rows = df[df["event_id"].astype(str) == str(event.get("id"))] if not df.empty else pd.DataFrame()
    confirmed_rows = event_rows[event_rows["status"] == "confirmed"] if not event_rows.empty else pd.DataFrame()
    interested_rows = event_rows[event_rows["status"].isin(["interested", "confirmed", "waitlist"])] if not event_rows.empty else pd.DataFrame()
    wait_rows = event_rows[event_rows["status"] == "waitlist"] if not event_rows.empty else pd.DataFrame()
    confirmed = int(event.get("base_confirmed", 0)) + len(confirmed_rows)
    interested = int(event.get("base_interested", 0)) + len(interested_rows)
    aliases = [str(x) for x in confirmed_rows.get("alias", pd.Series(dtype=str)).dropna().tolist() if str(x).strip()]
    return {
        "confirmed": confirmed,
        "interested": interested,
        "waitlist": len(wait_rows),
        "aliases": aliases,
    }


def _event_waitlist_position(event_id: str) -> int | None:
    session_id = str(st.session_state.get("session_id", ""))
    df = _read_event_rsvps()
    if df.empty:
        return None
    rows = df[(df["event_id"].astype(str) == str(event_id)) & (df["status"] == "waitlist")].copy()
    if rows.empty:
        return None
    rows = rows.sort_values("fecha").reset_index(drop=True)
    matches = rows.index[rows["session_id"].astype(str) == session_id].tolist()
    return matches[0] + 1 if matches else None


def _secret_or_env(name: str) -> str:
    try:
        value = st.secrets.get(name, "")
        if value:
            return str(value).strip()
    except Exception:
        pass
    return str(os.getenv(name, "") or "").strip()


def _event_whatsapp_url(event: dict) -> str:
    direct = str(event.get("whatsapp_url", "") or "").strip()
    if direct.startswith("https://"):
        return direct
    secret_name = str(event.get("whatsapp_secret", "") or "").strip()
    value = _secret_or_env(secret_name) if secret_name else ""
    return value if value.startswith("https://") else ""


def _event_calendar_url(event: dict) -> str:
    start = str(event.get("start", "") or "")
    end = str(event.get("end", "") or "")
    if not start or not end:
        return ""
    title = quote(str(event.get("title", "Evento NexoGeek")))
    details = quote(str(event.get("description", "")) + "\nInscripción gestionada desde NexoGeek.")
    location = quote(str(event.get("location", "")))
    return f"https://calendar.google.com/calendar/render?action=TEMPLATE&text={title}&dates={start}/{end}&details={details}&location={location}"


def _status_badge(status: str) -> str:
    labels = {
        "confirmed": ("confirmed", "✓ Asistencia confirmada"),
        "interested": ("interested", "♡ Marcado como interesante"),
        "waitlist": ("waitlist", "⏳ En lista de espera"),
        "cancelled": ("cancelled", "Cupo cancelado"),
    }
    css, label = labels.get(status, ("interested", "Sin inscripción"))
    return f'<span class="event-status {css}">{label}</span>'


def _render_event_card(event: dict, prefix: str) -> None:
    event_id = str(event.get("id"))
    snapshot = _event_snapshot(event)
    capacity = max(1, int(event.get("capacity", 1)))
    confirmed = int(snapshot["confirmed"])
    available = max(0, capacity - confirmed)
    status = _current_event_status(event_id)
    progress = min(1.0, confirmed / capacity)

    st.markdown(
        f'<div class="event-shell"><div class="event-date">{_safe_text(event.get("date_label",""))}</div>'
        f'<h3>{_safe_text(event.get("title",""))}</h3><p>{_safe_text(event.get("description",""))}</p>'
        f'<div class="event-meta-row"><span class="event-chip">📍 {_safe_text(event.get("location",""))}</span>'
        f'<span class="event-chip">{_safe_text(event.get("modality",""))}</span>'
        f'<span class="event-chip">Organiza {_safe_text(event.get("organizer","Comunidad"))}</span></div>'
        f'{_status_badge(status) if status else ""}'
        f'<div class="event-counts"><strong>{confirmed}/{capacity}</strong> confirmados · '
        f'{snapshot["interested"]} interesados · {available} cupos disponibles</div></div>',
        unsafe_allow_html=True,
    )
    st.progress(progress)

    b1, b2, b3 = st.columns(3)
    with b1:
        if st.button("♡ Me interesa", key=f"{prefix}_interest_{event_id}", use_container_width=True,
                     disabled=status in {"interested", "confirmed", "waitlist"}):
            _write_event_rsvp(event_id, "interested")
            st.rerun()
    with b2:
        label = "✓ Confirmado" if status == "confirmed" else ("⏳ En espera" if status == "waitlist" else "Confirmar asistencia")
        if st.button(label, key=f"{prefix}_confirm_{event_id}", type="primary", use_container_width=True,
                     disabled=status in {"confirmed", "waitlist"}):
            current = _event_snapshot(event)
            new_status = "confirmed" if int(current["confirmed"]) < capacity else "waitlist"
            _write_event_rsvp(event_id, new_status)
            if new_status == "confirmed":
                _notify(f"Cupo confirmado para {event.get('title','el evento')}", "success")
            else:
                _notify(f"Ingresaste a la lista de espera de {event.get('title','el evento')}", "info")
            st.rerun()
    with b3:
        if st.button("Cancelar", key=f"{prefix}_cancel_{event_id}", use_container_width=True,
                     disabled=status not in {"interested", "confirmed", "waitlist"}):
            _write_event_rsvp(event_id, "cancelled")
            _notify(f"Cancelaste tu participación en {event.get('title','el evento')}", "info")
            st.rerun()

    if status == "waitlist":
        position = _event_waitlist_position(event_id)
        st.info(f"Estás en la lista de espera" + (f" en la posición {position}." if position else "."))

    link1, link2 = st.columns(2)
    calendar_url = _event_calendar_url(event)
    whatsapp_url = _event_whatsapp_url(event)
    with link1:
        if calendar_url:
            st.link_button(f"📅 Agregar · {str(event.get('title',''))[:18]}", calendar_url, use_container_width=True)
    with link2:
        if status == "confirmed" and whatsapp_url:
            st.link_button(f"💬 Grupo · {str(event.get('title',''))[:20]}", whatsapp_url, use_container_width=True)
        elif status == "confirmed":
            st.caption("El organizador habilitará aquí el grupo de coordinación.")
        else:
            st.caption("El grupo de coordinación se habilita después de confirmar.")

    aliases = snapshot.get("aliases", [])
    if aliases:
        preview = ", ".join(aliases[:6])
        extra = len(aliases) - 6
        st.markdown(
            f'<div class="event-person-list"><strong>Participantes del piloto:</strong> {_safe_text(preview)}'
            f'{" y " + str(extra) + " más" if extra > 0 else ""}</div>', unsafe_allow_html=True,
        )


def render_community() -> None:
    _track_event("visita_comunidad", once=True)
    st.markdown(
        '<div class="community-banner"><span class="section-kicker">COMUNIDAD GEEK</span>'
        '<h2>Descubre personas, actividades y proyectos.</h2>'
        '<p>Confirma tu asistencia dentro de NexoGeek y usa WhatsApp solo como canal opcional de coordinación.</p></div>',
        unsafe_allow_html=True,
    )

    listings = [x for x in st.session_state.get("marketplace_db", []) if x.get("active", True)]
    services = st.session_state.get("servicios_db", [])
    auctions = st.session_state.get("subastas_db", [])
    events = st.session_state.get("community_events_db", _demo_community_events_seed())
    rsvps = _read_event_rsvps()
    confirmed_total = int((rsvps.get("status", pd.Series(dtype=str)) == "confirmed").sum()) if not rsvps.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Vendedores activos", len(set(x.get("seller") for x in listings)))
    c2.metric("Servicios disponibles", len(services))
    c3.metric("Eventos en agenda", len(events))
    c4.metric("Confirmaciones piloto", confirmed_total)

    st.markdown(
        '<div class="boardgame-spotlight"><div class="section-kicker">MÁS QUE TCG</div>'
        '<h3>Juegos de mesa también tienen espacio en NexoGeek</h3>'
        '<p>Explora publicaciones de Catan, Dixit y Ticket to Ride, o confirma asistencia a una tarde abierta para encontrar grupo.</p></div>',
        unsafe_allow_html=True,
    )
    if st.button("🎲 Ver juegos de mesa en Marketplace", type="primary", key="community_boardgames_market", use_container_width=True):
        st.session_state["market_search_v3"] = "Juegos de mesa"
        _track_event("categoria_comunidad", "juegos_de_mesa")
        _go_to("Marketplace")

    tabs = st.tabs(["Agenda e inscripciones", "Creadores y servicios", "Subastas destacadas", "Participa"])
    with tabs[0]:
        active_statuses = []
        for event in events:
            status = _current_event_status(str(event.get("id")))
            if status in {"interested", "confirmed", "waitlist"}:
                active_statuses.append((event, status))
        if active_statuses:
            with st.expander(f"Mis eventos ({len(active_statuses)})", expanded=False):
                for event, status in active_statuses:
                    st.markdown(f"**{event.get('title')}** · {_safe_text(status)}")

        for start in range(0, len(events), 2):
            cols = st.columns(2)
            for col, event in zip(cols, events[start:start+2]):
                with col:
                    _render_event_card(event, "community")
        st.caption(
            "Las inscripciones del piloto se guardan en un archivo compartido del servidor. "
            "En producción deberían persistirse en una base de datos con usuarios autenticados."
        )

    with tabs[1]:
        if not services:
            _render_empty_state("✨", "Aún no hay servicios", "La comunidad podrá publicar encargos, diseño, impresión, grading y organización.")
        else:
            cols = st.columns(min(4, len(services)))
            for col, service in zip(cols, services[:4]):
                with col:
                    st.markdown(
                        f'<div class="soft-card"><div class="section-kicker">{_safe_text(service.get("type","Servicio"))}</div>'
                        f'<h4>{_safe_text(service.get("title",""))}</h4>'
                        f'<p>{_safe_text(service.get("description",""))}</p>'
                        f'<strong>{_fmt_clp(service.get("price",0))}</strong><br>'
                        f'<small>{_safe_text(service.get("provider",""))} · ⭐ {service.get("rating","-")}</small></div>',
                        unsafe_allow_html=True,
                    )
            if st.button("Ver todos los servicios", type="primary", key="community_all_services_events"):
                _go_to("Servicios")

    with tabs[2]:
        if not auctions:
            _render_empty_state("🔨", "Sin subastas activas", "Las subastas verificadas aparecerán aquí.")
        else:
            for auction in auctions[:3]:
                with st.container(border=True):
                    a1, a2, a3 = st.columns([3, 1, 1])
                    a1.markdown(f"**{auction.get('name','')}**  \n{auction.get('game','')} · {auction.get('seller','')}")
                    a2.metric("Oferta", _fmt_clp(auction.get("current_bid", 0)))
                    a3.metric("Pujas", auction.get("bids", 0))
            if st.button("Entrar a Subastas", key="community_all_auctions_events", use_container_width=True):
                _go_to("Subastas")

    with tabs[3]:
        st.markdown("#### Propón una actividad para la comunidad")
        with st.form("community_create_event_form", clear_on_submit=True):
            e1, e2 = st.columns([2, 1])
            title = e1.text_input("Nombre de la actividad *", placeholder="Ej. Noche abierta de juegos de mesa")
            date_label = e2.text_input("Fecha u horario", placeholder="Sábado · 18:00")
            e3, e4, e5 = st.columns(3)
            location = e3.text_input("Lugar", placeholder="Ñuñoa")
            modality = e4.selectbox("Modalidad", ["Presencial", "Online", "Híbrido"])
            capacity = e5.number_input("Cupos", min_value=2, max_value=500, value=20)
            description = st.text_area("Descripción", placeholder="Qué harán, a quién está dirigida y qué debe llevar cada persona.")
            whatsapp = st.text_input("Enlace del grupo de WhatsApp (opcional)", placeholder="https://chat.whatsapp.com/...")
            create = st.form_submit_button("Publicar actividad demo", type="primary", use_container_width=True)
        if create:
            if not title.strip() or not location.strip():
                st.error("Completa al menos el nombre y el lugar.")
            else:
                event = {
                    "id": _new_id("community-event"), "date_label": date_label.strip() or "FECHA POR CONFIRMAR",
                    "title": title.strip(), "description": description.strip() or "Actividad propuesta durante el piloto.",
                    "location": location.strip(), "modality": modality, "capacity": int(capacity),
                    "base_confirmed": 0, "base_interested": 0,
                    "organizer": st.session_state.get("pilot_alias", "Usuario_Piloto"),
                    "start": "", "end": "", "whatsapp_url": whatsapp.strip(),
                }
                st.session_state.setdefault("community_events_db", []).insert(0, event)
                _track_event("crear_evento_comunidad", event["id"], event["title"])
                st.success("Actividad creada para esta sesión de demostración.")
                st.rerun()

        st.markdown("---")
        p1, p2 = st.columns(2)
        with p1:
            st.markdown("**Vende una pieza**  \nPublica cartas, figuras, mangas, juegos o accesorios.")
            if st.button("Crear publicación", key="community_publish_events", use_container_width=True):
                _go_to("Vender")
        with p2:
            st.markdown("**Ofrece un servicio**  \nDiseño, impresión, grading, encargos u organización.")
            if st.button("Publicar servicio", key="community_service_events", use_container_width=True):
                _go_to("Servicios")

        if st.session_state.get("admin_unlocked"):
            if st.button("🧹 Limpiar confirmaciones del piloto", key="clear_event_rsvps", use_container_width=True):
                try:
                    if os.path.exists(EVENT_RSVP_FILE):
                        os.remove(EVENT_RSVP_FILE)
                except Exception:
                    pass
                st.session_state["event_rsvp_status"] = {}
                st.success("Confirmaciones eliminadas.")
                st.rerun()

    _render_feature_vote("comunidad", "¿Usarías NexoGeek para descubrir y confirmar actividades de la comunidad?")


# Oculta completamente la barra lateral pública y el control para desplegarla.
# La lógica interna del sidebar se mantiene para conservar la API Key, la tasa
# USD/CLP y las preferencias del piloto, pero ya no se muestra a visitantes.
HIDE_SIDEBAR_CSS = """
<style>
section[data-testid="stSidebar"],
[data-testid="stSidebar"] {
    display: none !important;
}
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"] {
    display: none !important;
}
/* Permite que el contenido principal aproveche el ancho disponible. */
[data-testid="stAppViewContainer"] > .main,
[data-testid="stMain"] {
    margin-left: 0 !important;
}
</style>
"""

def main():
    st.set_page_config(
        page_title="NexoGeek · Marketplace Geek",
        page_icon="✦",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown(
        DARK_CSS + EXTRA_CSS + ORIGINAL_IDENTITY_CSS + PILOT_V3_CSS
        + META_LAB_CSS + META_DECK_BUILDER_CSS + GEEK_MARKETPLACE_CSS + COMMUNITY_EVENTS_CSS
        + HIDE_SIDEBAR_CSS,
        unsafe_allow_html=True,
    )
    _init_demo_state()
    _init_meta_state()

    api_key, clp_rate, comision = _render_sidebar(obtener_api_key_streamlit())
    page = _render_top_navigation()
    _track_event("visita_pagina", page, once=True)

    if page == "Inicio":
        render_home()
    elif page == "Marketplace":
        context = st.session_state.pop("meta_marketplace_context", None)
        if context:
            st.info(
                f"Buscando **{context.get('carta')}** para completar "
                f"**{context.get('mazo')}**."
            )
        render_marketplace()
    elif page == "Detalle":
        render_product_detail()
    elif page == "Vender":
        render_sell()
    elif page == "Subastas":
        render_auctions()
    elif page == "Servicios":
        render_services()
    elif page == "Comunidad":
        render_community()
    elif page == "Herramientas":
        render_tools_hub()
    elif page == "Tasador":
        render_catalogador(api_key, clp_rate, comision)
    elif page == "Meta Lab":
        render_meta_lab(api_key, clp_rate)
    elif page == "Feedback":
        render_feedback()

    st.markdown(
        "<br><hr style='border:none;border-top:2px solid rgba(40,22,58,.12);'>",
        unsafe_allow_html=True,
    )
    foot1, foot2 = st.columns([3, 1])
    foot1.caption(
        "NexoGeek · Marketplace geek para comprar, vender y conectar con la comunidad. "
        "Meta Lab y el constructor visual funcionan como herramientas complementarias."
    )
    foot2.caption("Mercado · Comunidad · Colección")


if __name__ == "__main__":
    main()
